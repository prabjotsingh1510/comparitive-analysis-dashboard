"""
Build new D object for index.html from Channel Economics Dashboard.xlsx
This is the primary data source for all channel visualizations.
"""
import openpyxl
import json

wb = openpyxl.load_workbook('Channel Economics Dashboard.xlsx', data_only=True)

def num(v):
    if v is None: return 0.0
    try: return float(v)
    except: return 0.0

def safe_f(v):
    try: return float(v) if v is not None else None
    except: return None

# ============================================================
# WEBSITE
# ============================================================
ws_web = wb['Raw Data - Website']
web_headers = [ws_web.cell(row=4, column=c).value for c in range(1, 50)]
W = []
for r in range(5, 200):
    code = ws_web.cell(row=r, column=1).value
    if code is None: continue
    row = {}
    for c, h in enumerate(web_headers[:42], 1):
        if h:
            v = ws_web.cell(row=r, column=c).value
            try: row[h] = float(v) if v is not None else None
            except: row[h] = v
    W.append(row)

# ============================================================
# AMAZON  
# ============================================================
ws_amz = wb['Raw Data - Amazon']
amz_headers = [ws_amz.cell(row=4, column=c).value for c in range(1, 50)]
A = []
for r in range(5, 200):
    prod = ws_amz.cell(row=r, column=1).value
    if prod is None: continue
    row = {}
    for c, h in enumerate(amz_headers[:45], 1):
        if h:
            v = ws_amz.cell(row=r, column=c).value
            try: row[h] = float(v) if v is not None else None
            except: row[h] = v
    A.append(row)

# ============================================================
# FIRSTCRY
# ============================================================
ws_fc = wb['Raw Data - FirstCry']
fc_headers = [ws_fc.cell(row=4, column=c).value for c in range(1, 35)]
FC_all = []
for r in range(5, 200):
    prod = ws_fc.cell(row=r, column=1).value
    if prod is None: continue
    row = {}
    for c, h in enumerate(fc_headers, 1):
        if h:
            v = ws_fc.cell(row=r, column=c).value
            try: row[h] = float(v) if v is not None else None
            except: row[h] = v
    FC_all.append(row)
FC = [p for p in FC_all if num(p.get('Units (3 months)')) > 0]

# ============================================================
# BLINKIT
# ============================================================
ws_bl = wb['Raw Data - Blinkit']
bl_headers = [ws_bl.cell(row=4, column=c).value for c in range(1, 35)]
BL = []
for r in range(5, 200):
    prod = ws_bl.cell(row=r, column=1).value
    if prod is None: continue
    row = {}
    for c, h in enumerate(bl_headers, 1):
        if h:
            v = ws_bl.cell(row=r, column=c).value
            try: row[h] = float(v) if v is not None else None
            except: row[h] = v
    BL.append(row)

# ============================================================
# COMPUTE AGGREGATES
# ============================================================
def web_prod_to_dict(p):
    """Convert website product row to standard dict for D.apr.products"""
    code = p.get('Product Name (Code)', '')
    disp = p.get('Product Name (Full)', code)
    units = num(p.get('Units sold'))
    netrev_u = num(p.get('Net Revenue without GST'))
    gm_u = num(p.get('Gross margin'))
    cm1_u = gm_u - num(p.get('Packaging')) - num(p.get('Razorpay Deduction')) - num(p.get('Shipping to Customer')) - num(p.get('COD Costs')) - num(p.get('Shopify Charges')) - num(p.get('Shopflo'))
    cm2_u = num(p.get('CM2'))
    oh_u = num(p.get('Overhead Cost'))
    ebitda_u = num(p.get('EBITA'))
    mktg_u = num(p.get('Marketing Cost'))
    ship_u = num(p.get('Shipping to Customer'))
    mrp = num(p.get('MRP Including GST'))
    
    cm2_t = cm2_u * units
    ebitda_t = ebitda_u * units
    
    if cm2_u > 0 and ebitda_u > 0:
        quad = 'star'
    elif cm2_u > 0:
        quad = 'overhead'
    else:
        quad = 'loss'
    
    return {
        'k': f"{code}||{disp}",
        'key': code,
        'disp': disp,
        'cat': p.get('Category', ''),
        'rev': round(num(p.get('3 month revenue with GST')), 2),
        'units': units,
        'mrp': mrp if mrp else None,
        'netrev_u': round(netrev_u, 2),
        'gm_u': round(gm_u, 2),
        'ship_u': round(ship_u, 2),
        'mktg_u': round(mktg_u, 2),
        'cm1_u': round(num(p.get('CM 1', 0)), 2),
        'cm2_u': round(cm2_u, 2),
        'oh_u': round(oh_u, 2),
        'ebitda_u': round(ebitda_u, 2),
        'cm2_t': round(cm2_t, 2),
        'ebitda_t': round(ebitda_t, 2),
        'cm2pct': round(num(p.get('CM2%', 0))*100, 2),
        'ebitdapct': round(num(p.get('EBITA%', 0))*100, 2),
        'quad': quad
    }

def amz_prod_to_dict(p):
    """Convert amazon product row to D format"""
    code = p.get('P. Breakdown', '')
    disp = p.get('Product', code)
    units = num(p.get('Units Sold'))
    netrev_u = num(p.get('Net Revenue without GST'))
    gm_u = num(p.get('Gross Margin'))
    cm2_u = num(p.get('CM2'))
    oh_u = num(p.get('Over heads'))
    ebitda_u = num(p.get('EBITDA'))
    mktg_u = num(p.get('Cost of Advertising'))
    
    cm2_t = cm2_u * units
    ebitda_t = ebitda_u * units
    net_rev_t = num(p.get('3 - Month Net Revenue without GST', netrev_u * units))
    
    if cm2_u > 0 and ebitda_u > 0:
        quad = 'star'
    elif cm2_u > 0:
        quad = 'overhead'
    else:
        quad = 'loss'
    
    return {
        'k': code,
        'key': code,
        'disp': disp,
        'cat': p.get('Category', ''),
        'units': units,
        'mrp': num(p.get('MRP')),
        'netrev_u': round(netrev_u, 3),
        'gm_u': round(gm_u, 3),
        'mktg_u': round(mktg_u, 2),
        'cm1_u': round(num(p.get('CM1')), 3),
        'cm2_u': round(cm2_u, 3),
        'oh_u': round(oh_u, 2),
        'ebitda_u': round(ebitda_u, 3),
        'cm2_t': round(cm2_t, 2),
        'ebitda_t': round(ebitda_t, 2),
        'net_rev_t': round(net_rev_t, 2),
        'cm2pct': round(num(p.get('CM2%', 0))*100, 2),
        'ebitdapct': round(num(p.get('EBITDA%', 0))*100, 2),
        'quad': quad
    }

def fc_prod_to_dict(p):
    code = p.get('P. Breakdown', '')
    disp = p.get('ProductName', code)
    units = num(p.get('Units (3 months)'))
    netrev_u = num(p.get('Net Revenue without GST'))
    gm_u = num(p.get('Gross Margin'))
    cm2_u = num(p.get('CM2'))
    ebitda_u = num(p.get('EBITDA'))
    total_rev = num(p.get('Total Net Revenue'))
    
    if cm2_u > 0 and ebitda_u > 0:
        quad = 'star'
    elif cm2_u > 0:
        quad = 'overhead'
    else:
        quad = 'loss'
    
    return {
        'key': code,
        'disp': disp,
        'cat': p.get('Category', ''),
        'units': units,
        'mrp': num(p.get('MRP')),
        'netrev_u': round(netrev_u, 2),
        'gm_u': round(gm_u, 2),
        'cm2_u': round(cm2_u, 2),
        'ebitda_u': round(ebitda_u, 2),
        'cm2_t': round(cm2_u*units, 2),
        'ebitda_t': round(ebitda_u*units, 2),
        'total_rev': round(total_rev, 2),
        'cm2pct': round(num(p.get('CM2%', 0))*100, 2),
        'ebitdapct': round(num(p.get('EBITDA%', 0))*100, 2),
        'quad': quad,
        'drr': round(num(p.get('Daily Run Rate(DRR)', 0)), 4)
    }

def bl_prod_to_dict(p):
    code = p.get('P. Breakdown', '')
    disp = p.get('item_name', code)
    units = num(p.get('Units (3 months)'))
    netrev_u = num(p.get('Net Revenue without GST'))
    gm_u = num(p.get('Gross Margin'))
    cm2_u = num(p.get('CM2'))
    ebitda_u = num(p.get('EBITDA'))
    mktg_u = num(p.get('Marketing Cost'))
    bl_margin_u = num(p.get('Blinkit Margin'))
    
    if cm2_u > 0 and ebitda_u > 0:
        quad = 'star'
    elif cm2_u > 0:
        quad = 'overhead'
    else:
        quad = 'loss'
    
    return {
        'key': code,
        'disp': disp,
        'cat': p.get('Category', ''),
        'units': units,
        'mrp': num(p.get('MRP')),
        'sp': num(p.get('SP')),
        'netrev_u': round(netrev_u, 2),
        'gm_u': round(gm_u, 2),
        'mktg_u': round(mktg_u, 2),
        'bl_margin_u': round(bl_margin_u, 2),
        'cm1_u': round(num(p.get('CM1')), 2),
        'cm2_u': round(cm2_u, 2),
        'ebitda_u': round(ebitda_u, 2),
        'cm2_t': round(cm2_u*units, 2),
        'ebitda_t': round(ebitda_u*units, 2),
        'cm2pct': round(num(p.get('CM2%', 0))*100, 2),
        'ebitdapct': round(num(p.get('EBITDA%', 0))*100, 2),
        'quad': quad
    }

# Build aggregates
def agg_channel(prods, rev_key, units_key, cm2_key, ebitda_key, mktg_key, gm_key, oh_key):
    total_rev = sum(num(p.get(rev_key)) for p in prods)
    total_units = sum(num(p.get(units_key)) for p in prods)
    total_net_rev = sum(num(p.get('netrev_u', 0)) * num(p.get(units_key)) for p in prods)
    total_cm2 = sum(num(p.get(cm2_key)) for p in prods)
    total_ebitda = sum(num(p.get(ebitda_key)) for p in prods)
    total_mktg = sum(num(p.get(mktg_key)) * num(p.get(units_key)) for p in prods)
    total_gm = sum(num(p.get(gm_key)) * num(p.get(units_key)) for p in prods)
    total_oh = sum(num(p.get(oh_key)) * num(p.get(units_key)) for p in prods)
    return {
        'units': total_units, 'rev': total_rev, 'net_rev': total_net_rev,
        'cm2': total_cm2, 'ebitda': total_ebitda, 'mktg': total_mktg,
        'gm': total_gm, 'oh': total_oh,
        'cm2pct': total_cm2/total_net_rev if total_net_rev else 0,
        'ebitdapct': total_ebitda/total_net_rev if total_net_rev else 0,
        'gmpct': total_gm/total_net_rev if total_net_rev else 0,
        'mktgpct': total_mktg/total_net_rev if total_net_rev else 0
    }

# Convert products
web_prods = [web_prod_to_dict(p) for p in W]
amz_prods = [amz_prod_to_dict(p) for p in A]
fc_prods = [fc_prod_to_dict(p) for p in FC]
bl_prods = [bl_prod_to_dict(p) for p in BL]

# Website aggregates
web_agg = {
    'units': sum(p['units'] for p in web_prods),
    'rev': sum(p['rev'] for p in web_prods),
    'net_rev': sum(num(p.get('netrev_u'))*p['units'] for p in web_prods),
    'cm2': sum(p['cm2_t'] for p in web_prods),
    'ebitda': sum(p['ebitda_t'] for p in web_prods),
    'mktg': sum(p['mktg_u']*p['units'] for p in web_prods),
    'gm': sum(p['gm_u']*p['units'] for p in web_prods),
}
web_agg['cm2pct'] = web_agg['cm2']/web_agg['net_rev'] if web_agg['net_rev'] else 0
web_agg['ebitdapct'] = web_agg['ebitda']/web_agg['net_rev'] if web_agg['net_rev'] else 0
web_agg['gmpct'] = web_agg['gm']/web_agg['net_rev'] if web_agg['net_rev'] else 0
web_agg['mktgpct'] = web_agg['mktg']/web_agg['net_rev'] if web_agg['net_rev'] else 0

# Amazon aggregates
amz_agg = {
    'units': sum(p['units'] for p in amz_prods),
    'net_rev': sum(num(p.get('netrev_u'))*p['units'] for p in amz_prods),
    'cm2': sum(p['cm2_t'] for p in amz_prods),
    'ebitda': sum(p['ebitda_t'] for p in amz_prods),
    'mktg': sum(p['mktg_u']*p['units'] for p in amz_prods),
    'gm': sum(p['gm_u']*p['units'] for p in amz_prods),
}
amz_agg['cm2pct'] = amz_agg['cm2']/amz_agg['net_rev'] if amz_agg['net_rev'] else 0
amz_agg['ebitdapct'] = amz_agg['ebitda']/amz_agg['net_rev'] if amz_agg['net_rev'] else 0
amz_agg['gmpct'] = amz_agg['gm']/amz_agg['net_rev'] if amz_agg['net_rev'] else 0
amz_agg['mktgpct'] = amz_agg['mktg']/amz_agg['net_rev'] if amz_agg['net_rev'] else 0

# FirstCry aggregates
fc_agg = {
    'units': sum(p['units'] for p in fc_prods),
    'net_rev': sum(p['total_rev'] for p in fc_prods),
    'cm2': sum(p['cm2_t'] for p in fc_prods),
    'ebitda': sum(p['ebitda_t'] for p in fc_prods),
    'gm': sum(p['gm_u']*p['units'] for p in fc_prods),
}
fc_agg['cm2pct'] = fc_agg['cm2']/fc_agg['net_rev'] if fc_agg['net_rev'] else 0
fc_agg['ebitdapct'] = fc_agg['ebitda']/fc_agg['net_rev'] if fc_agg['net_rev'] else 0
fc_agg['gmpct'] = fc_agg['gm']/fc_agg['net_rev'] if fc_agg['net_rev'] else 0

# Blinkit aggregates
bl_agg = {
    'units': sum(p['units'] for p in bl_prods),
    'net_rev': sum(p['netrev_u']*p['units'] for p in bl_prods),
    'cm2': sum(p['cm2_t'] for p in bl_prods),
    'ebitda': sum(p['ebitda_t'] for p in bl_prods),
    'mktg': sum(p['mktg_u']*p['units'] for p in bl_prods),
    'gm': sum(p['gm_u']*p['units'] for p in bl_prods),
    'bl_margin': sum(p['bl_margin_u']*p['units'] for p in bl_prods),
}
bl_agg['cm2pct'] = bl_agg['cm2']/bl_agg['net_rev'] if bl_agg['net_rev'] else 0
bl_agg['ebitdapct'] = bl_agg['ebitda']/bl_agg['net_rev'] if bl_agg['net_rev'] else 0
bl_agg['gmpct'] = bl_agg['gm']/bl_agg['net_rev'] if bl_agg['net_rev'] else 0
bl_agg['mktgpct'] = bl_agg['mktg']/bl_agg['net_rev'] if bl_agg['net_rev'] else 0
bl_agg['bl_margin_pct'] = bl_agg['bl_margin']/bl_agg['net_rev'] if bl_agg['net_rev'] else 0

# Marketing for website
total_web_rev_gst = sum(num(p.get('3 month revenue with GST')) for p in W)
total_web_rev_no_gst = sum(num(p.get('3 month revenue without GST')) for p in W)

# Category breakdown for website
cats_web = {}
for p in web_prods:
    cat = p.get('cat', 'Other') or 'Other'
    if cat not in cats_web:
        cats_web[cat] = {'cat': cat, 'rev': 0, 'units': 0, 'cm2': 0, 'ebitda': 0, 'n': 0, 'mktg': 0}
    cats_web[cat]['rev'] += p.get('rev', 0)
    cats_web[cat]['units'] += p.get('units', 0)
    cats_web[cat]['cm2'] += p.get('cm2_t', 0)
    cats_web[cat]['ebitda'] += p.get('ebitda_t', 0)
    cats_web[cat]['n'] += 1
    cats_web[cat]['mktg'] += p.get('mktg_u', 0) * p.get('units', 0)

cats_web_list = sorted(cats_web.values(), key=lambda x: -x['rev'])

# Build the new D object
D = {
    'web': {
        'agg': web_agg,
        'products': web_prods,
        'cats': cats_web_list,
        'rev_gst': round(total_web_rev_gst, 2),
        'rev_no_gst': round(total_web_rev_no_gst, 2),
        'mktg_pct_gst': round(web_agg['mktg']/total_web_rev_gst*100, 2) if total_web_rev_gst else 0,
        'mktg_pct_no_gst': round(web_agg['mktg']/total_web_rev_no_gst*100, 2) if total_web_rev_no_gst else 0
    },
    'amz': {
        'agg': amz_agg,
        'products': amz_prods
    },
    'fc': {
        'agg': fc_agg,
        'products': fc_prods
    },
    'bl': {
        'agg': bl_agg,
        'products': bl_prods
    }
}

# Write D object to JSON for injection
with open('new_D_data.json', 'w', encoding='utf-8') as f:
    json.dump(D, f, ensure_ascii=False, default=str)

print("new_D_data.json written")
print(f"Website: {len(web_prods)} prods, net_rev={web_agg['net_rev']:,.0f}, cm2={web_agg['cm2']:,.0f} ({web_agg['cm2pct']*100:.1f}%), ebitda={web_agg['ebitda']:,.0f}")
print(f"Amazon: {len(amz_prods)} prods, net_rev={amz_agg['net_rev']:,.0f}, cm2={amz_agg['cm2']:,.0f} ({amz_agg['cm2pct']*100:.1f}%), ebitda={amz_agg['ebitda']:,.0f}")
print(f"FirstCry: {len(fc_prods)} prods, net_rev={fc_agg['net_rev']:,.0f}, cm2={fc_agg['cm2']:,.0f} ({fc_agg['cm2pct']*100:.1f}%), ebitda={fc_agg['ebitda']:,.0f}")
print(f"Blinkit: {len(bl_prods)} prods, net_rev={bl_agg['net_rev']:,.0f}, cm2={bl_agg['cm2']:,.0f} ({bl_agg['cm2pct']*100:.1f}%), ebitda={bl_agg['ebitda']:,.0f}")
