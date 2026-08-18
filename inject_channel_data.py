"""
Build new D_AMZ, D_FC, D_BL objects from Channel Economics Dashboard.xlsx
and inject them into index.html, replacing the old values.
"""
import openpyxl, json, re

wb = openpyxl.load_workbook('Channel Economics Dashboard.xlsx', data_only=True)

def num(v):
    if v is None: return 0.0
    try: return float(v)
    except: return 0.0

# ============================================================
# AMAZON
# ============================================================
ws_amz = wb['Raw Data - Amazon']
amz_headers = [ws_amz.cell(row=4, column=c).value for c in range(1, 50)]

A = []
for r in range(5, 200):
    prod = ws_amz.cell(row=r, column=1).value
    if prod is None: continue
    row = {}
    for c, h in enumerate(amz_headers[:45], 1):
        if h:
            v = ws_amz.cell(row=r, column=c).value
            try: row[h] = float(v) if v is not None else None
            except: row[h] = v
    A.append(row)

amz_prods = []
for p in A:
    code = str(p.get('P. Breakdown', '') or '')
    disp = str(p.get('Product', code) or code)
    units = num(p.get('Units Sold'))
    netrev_u = num(p.get('Net Revenue without GST'))
    gm_u = num(p.get('Gross Margin'))
    cogs_u = num(p.get('COGS'))
    cm1_u = num(p.get('CM1'))
    cm2_u = num(p.get('CM2'))
    oh_u = num(p.get('Over heads'))
    ebitda_u = num(p.get('EBITDA'))
    mktg_u = num(p.get('Cost of Advertising'))
    evit_u = num(p.get('E-vitamin'))
    bss_u = num(p.get('Business Support Services'))
    listing_u = num(p.get('Listing Fee'))
    closing_u = num(p.get('Closing Fee'))
    shipping_u = num(p.get('Shipping charge'))
    storage_u = num(p.get('Storage Fee (FBA)'))
    pickpack_u = num(p.get('Pick and pack Fee (FBA)'))
    lts_u = num(p.get('Long Term Storage Fee (FBA)'))
    giftwrap_u = num(p.get('Gift Wrap Fee (FBA)'))
    inbound_u = num(p.get('Inbound Transportation Fee (FBA)'))
    mrp = num(p.get('MRP'))
    asin = str(p.get('ASIN', '') or '')
    cat = str(p.get('Category', '') or '')
    net_rev_t = num(p.get('3 - Month Net Revenue without GST', 0)) or netrev_u * units
    
    # per-unit totals
    cm2_t = cm2_u * units
    ebitda_t = ebitda_u * units
    gm_t = gm_u * units
    cm1_t = cm1_u * units
    mktg_t = mktg_u * units

    if cm2_u > 0 and ebitda_u > 0:
        quad = 'star'
        act = 'scale'
    elif cm2_u > 0:
        quad = 'overhead'
        act = 'overhead'
    else:
        quad = 'loss'
        act = 'cut'

    amz_prods.append({
        'key': code, 'disp': disp[:96], 'asin': asin,
        'cat': cat, 'subcat': cat,
        'units': units, 'mrp': mrp,
        'netrev_u': round(netrev_u, 4), 'netrev_t': round(net_rev_t, 4),
        'cogs_u': cogs_u, 'gm_u': round(gm_u, 4), 'gm_t': round(gm_t, 4),
        'gmpct': round(gm_u/netrev_u, 4) if netrev_u else 0,
        'ship_u': round(shipping_u, 4),
        'cm1_u': round(cm1_u, 4), 'cm1_t': round(cm1_t, 4),
        'cm1pct': round(cm1_u/netrev_u, 4) if netrev_u else 0,
        'bss_u': round(bss_u, 4), 'evit_u': round(evit_u, 4),
        'mktg_u': round(mktg_u, 4), 'mktg_t': round(mktg_t, 4),
        'cm2_u': round(cm2_u, 4), 'cm2_t': round(cm2_t, 4),
        'cm2pct': round(num(p.get('CM2%', 0))*100, 4),
        'oh_u': round(oh_u, 4), 'oh_t': round(oh_u*units, 4),
        'ebitda_u': round(ebitda_u, 4), 'ebitda_t': round(ebitda_t, 4),
        'ebitdapct': round(num(p.get('EBITDA%', 0))*100, 4),
        'quad': quad, 'act': act,
        'listing_u': round(listing_u, 4), 'closing_u': round(closing_u, 4),
        'storage_u': round(storage_u, 4), 'pickpack_u': round(pickpack_u, 4),
        'lts_u': round(lts_u, 4), 'giftwrap_u': round(giftwrap_u, 4),
        'inbound_u': round(inbound_u, 4),
        'sessions': None, 'pageviews': None, 'units_ordered': None,
        's2p': None, 'p2u': None, 's2u': None
    })

# Amazon totals
T_amz_units = sum(p['units'] for p in amz_prods)
T_amz_net_rev = sum(p['netrev_t'] for p in amz_prods)
T_amz_gm = sum(p['gm_t'] for p in amz_prods)
T_amz_cm1 = sum(p['cm1_t'] for p in amz_prods)
T_amz_cm2 = sum(p['cm2_t'] for p in amz_prods)
T_amz_mktg = sum(p['mktg_t'] for p in amz_prods)
T_amz_oh = sum(p['oh_t'] for p in amz_prods)
T_amz_ebitda = sum(p['ebitda_t'] for p in amz_prods)

amz_star = [p for p in amz_prods if p['quad']=='star']
amz_overhead = [p for p in amz_prods if p['quad']=='overhead']
amz_loss = [p for p in amz_prods if p['quad']=='loss']

# Category rollups for Amazon
amz_cats = {}
for p in amz_prods:
    cat = p['cat'] or 'Other'
    if cat not in amz_cats:
        amz_cats[cat] = {'cat': cat, 'n': 0, 'rev': 0, 'units': 0, 'netrev': 0, 'mktg': 0, 'cm2': 0, 'cm2pct': 0, 'ebitda': 0, 'ebitdapct': 0}
    amz_cats[cat]['n'] += 1
    amz_cats[cat]['rev'] += p['netrev_t']
    amz_cats[cat]['units'] += p['units']
    amz_cats[cat]['netrev'] += p['netrev_t']
    amz_cats[cat]['mktg'] += p['mktg_t']
    amz_cats[cat]['cm2'] += p['cm2_t']
    amz_cats[cat]['ebitda'] += p['ebitda_t']
amz_cats_list = sorted(amz_cats.values(), key=lambda x: -x['rev'])
for c in amz_cats_list:
    c['cm2pct'] = round(c['cm2']/c['netrev']*100, 4) if c['netrev'] else 0
    c['ebitdapct'] = round(c['ebitda']/c['netrev']*100, 4) if c['netrev'] else 0

# Pareto for Amazon
amz_pareto = sorted(amz_prods, key=lambda p: -p['netrev_t'])
cum = 0
for p in amz_pareto:
    cum += p['netrev_t']
    p['_cum_share'] = round(cum/T_amz_net_rev*100, 4)

D_AMZ = {
    'channel': 'Amazon',
    'periodLabel': 'Apr\u2013Jun 2026',
    'tot': {
        'rev': round(T_amz_net_rev, 2),
        'netrev': round(T_amz_net_rev, 2),
        'units': int(T_amz_units),
        'gm': round(T_amz_gm, 2),
        'cm1': round(T_amz_cm1, 2),
        'cm2': round(T_amz_cm2, 2),
        'mktg': round(T_amz_mktg, 2),
        'oh': round(T_amz_oh, 2),
        'ebitda': round(T_amz_ebitda, 2),
        'cm2pct': round(T_amz_cm2/T_amz_net_rev*100, 5) if T_amz_net_rev else 0,
        'ebitdapct': round(T_amz_ebitda/T_amz_net_rev*100, 5) if T_amz_net_rev else 0,
        'gmpct': round(T_amz_gm/T_amz_net_rev, 7) if T_amz_net_rev else 0,
        'n': len(amz_prods)
    },
    'quad': {
        'star': {'n': len(amz_star), 'rev': round(sum(p['netrev_t'] for p in amz_star), 2), 'units': int(sum(p['units'] for p in amz_star)), 'ebitda': round(sum(p['ebitda_t'] for p in amz_star), 5)},
        'overhead': {'n': len(amz_overhead), 'rev': round(sum(p['netrev_t'] for p in amz_overhead), 2), 'units': int(sum(p['units'] for p in amz_overhead)), 'ebitda': round(sum(p['ebitda_t'] for p in amz_overhead), 5)},
        'loss': {'n': len(amz_loss), 'rev': round(sum(p['netrev_t'] for p in amz_loss), 2), 'units': int(sum(p['units'] for p in amz_loss)), 'ebitda': round(sum(p['ebitda_t'] for p in amz_loss), 5)}
    },
    'products': amz_prods,
    'cats': amz_cats_list,
    'pareto': [{'key': p['key'], 'disp': p['disp'][:80], 'rev': round(p['netrev_t'], 2), 'cum_share': p['_cum_share']} for p in amz_pareto],
    'fun': {'sessions': None, 'pageviews': None, 'units_ordered': None, 's2p': None, 'p2u': None, 's2u': None}
}

print(f"D_AMZ: {len(amz_prods)} prods, stars={len(amz_star)}, overhead={len(amz_overhead)}, loss={len(amz_loss)}")
print(f"  Net Rev={T_amz_net_rev:,.0f}, CM2={T_amz_cm2:,.0f} ({T_amz_cm2/T_amz_net_rev*100:.1f}%), EBITDA={T_amz_ebitda:,.0f}")

# ============================================================
# FIRSTCRY
# ============================================================
ws_fc = wb['Raw Data - FirstCry']
fc_headers = [ws_fc.cell(row=4, column=c).value for c in range(1, 35)]

FC_all = []
for r in range(5, 200):
    prod = ws_fc.cell(row=r, column=1).value
    if prod is None: continue
    row = {}
    for c, h in enumerate(fc_headers, 1):
        if h:
            v = ws_fc.cell(row=r, column=c).value
            try: row[h] = float(v) if v is not None else None
            except: row[h] = v
    FC_all.append(row)

FC = [p for p in FC_all if num(p.get('Units (3 months)')) > 0]

fc_prods = []
for p in FC:
    code = str(p.get('P. Breakdown', '') or '')
    disp = str(p.get('ProductName', code) or code)
    units = num(p.get('Units (3 months)'))
    netrev_u = num(p.get('Net Revenue without GST'))
    gm_u = num(p.get('Gross Margin'))
    fcm_u = num(p.get('Firstcry Margin'))
    cm2_u = num(p.get('CM2'))
    oh_u = num(p.get('Overhead'))
    ebitda_u = num(p.get('EBITDA'))
    total_rev = num(p.get('Total Net Revenue'))
    mrp = num(p.get('MRP'))
    drr = num(p.get('Daily Run Rate(DRR)'))
    
    netrev_t = total_rev if total_rev else netrev_u * units
    
    if cm2_u > 0 and ebitda_u > 0:
        quad = 'star'
    elif cm2_u > 0:
        quad = 'overhead'
    else:
        quad = 'loss'

    fc_prods.append({
        'key': code, 'disp': disp,
        'cat': str(p.get('Category', '') or ''),
        'units': units, 'mrp': mrp,
        'netrev_u': round(netrev_u, 2),
        'netrev_t': round(netrev_t, 2),
        'rev_t': round(netrev_t, 2),
        'cogs_u': num(p.get('COGS')),
        'gm_u': round(gm_u, 2), 'gm_t': round(gm_u*units, 2),
        'gmpct': round(gm_u/netrev_u*100, 2) if netrev_u else 0,
        'fcm_u': round(fcm_u, 2), 'fcm_t': round(fcm_u*units, 2),
        'cm2_u': round(cm2_u, 2), 'cm2_t': round(cm2_u*units, 2),
        'cm2pct': round(num(p.get('CM2%', 0))*100, 2),
        'oh_u': round(oh_u, 2), 'oh_t': round(oh_u*units, 2),
        'ebitda_u': round(ebitda_u, 2), 'ebitda_t': round(ebitda_u*units, 2),
        'ebitdapct': round(num(p.get('EBITDA%', 0))*100, 2),
        'quad': quad, 'drr': round(drr, 4)
    })

T_fc_units = sum(p['units'] for p in fc_prods)
T_fc_net_rev = sum(p['netrev_t'] for p in fc_prods)
T_fc_gm = sum(p['gm_t'] for p in fc_prods)
T_fc_fcm = sum(p['fcm_t'] for p in fc_prods)
T_fc_cm2 = sum(p['cm2_t'] for p in fc_prods)
T_fc_oh = sum(p['oh_t'] for p in fc_prods)
T_fc_ebitda = sum(p['ebitda_t'] for p in fc_prods)
T_fc_rev = sum(p['rev_t'] for p in fc_prods)

fc_star = [p for p in fc_prods if p['quad']=='star']
fc_overhead = [p for p in fc_prods if p['quad']=='overhead']
fc_loss = [p for p in fc_prods if p['quad']=='loss']

# FC cats
fc_cats = {}
for p in fc_prods:
    cat = p['cat'] or 'Other'
    if cat not in fc_cats:
        fc_cats[cat] = {'cat': cat, 'n': 0, 'rev': 0, 'units': 0, 'netrev': 0, 'cm2': 0, 'cm2pct': 0, 'ebitda': 0, 'ebitdapct': 0}
    fc_cats[cat]['n'] += 1; fc_cats[cat]['units'] += p['units']
    fc_cats[cat]['rev'] += p['netrev_t']; fc_cats[cat]['netrev'] += p['netrev_t']
    fc_cats[cat]['cm2'] += p['cm2_t']; fc_cats[cat]['ebitda'] += p['ebitda_t']
fc_cats_list = sorted(fc_cats.values(), key=lambda x: -x['rev'])
for c in fc_cats_list:
    c['cm2pct'] = round(c['cm2']/c['netrev']*100, 4) if c['netrev'] else 0
    c['ebitdapct'] = round(c['ebitda']/c['netrev']*100, 4) if c['netrev'] else 0

D_FC = {
    'channel': 'FirstCry', 'periodLabel': 'Apr\u2013Jun 2026',
    'tot': {
        'rev': round(T_fc_rev, 2), 'netrev': round(T_fc_net_rev, 2),
        'units': int(T_fc_units), 'gm': round(T_fc_gm, 2),
        'fcm': round(T_fc_fcm, 2), 'cm2': round(T_fc_cm2, 2),
        'oh': round(T_fc_oh, 2), 'ebitda': round(T_fc_ebitda, 2),
        'cm2pct': round(T_fc_cm2/T_fc_net_rev*100, 5) if T_fc_net_rev else 0,
        'ebitdapct': round(T_fc_ebitda/T_fc_net_rev*100, 5) if T_fc_net_rev else 0,
        'gmpct': round(T_fc_gm/T_fc_net_rev*100, 5) if T_fc_net_rev else 0,
        'n': len(fc_prods)
    },
    'quad': {
        'star': {'n': len(fc_star), 'rev': round(sum(p['rev_t'] for p in fc_star), 2), 'units': int(sum(p['units'] for p in fc_star)), 'ebitda': round(sum(p['ebitda_t'] for p in fc_star), 5)},
        'overhead': {'n': len(fc_overhead), 'rev': round(sum(p['rev_t'] for p in fc_overhead), 2), 'units': int(sum(p['units'] for p in fc_overhead)), 'ebitda': round(sum(p['ebitda_t'] for p in fc_overhead), 5)},
        'loss': {'n': len(fc_loss), 'rev': 0, 'units': 0, 'ebitda': 0}
    },
    'products': fc_prods, 'cats': fc_cats_list,
    'pareto': sorted([{'key': p['key'], 'disp': p['disp'][:80], 'rev': round(p['netrev_t'], 2)} for p in fc_prods], key=lambda x: -x['rev'])
}

print(f"D_FC: {len(fc_prods)} prods, stars={len(fc_star)}, overhead={len(fc_overhead)}, loss={len(fc_loss)}")
print(f"  Net Rev={T_fc_net_rev:,.0f}, CM2={T_fc_cm2:,.0f} ({T_fc_cm2/T_fc_net_rev*100:.1f}%), EBITDA={T_fc_ebitda:,.0f}")

# ============================================================
# BLINKIT
# ============================================================
ws_bl = wb['Raw Data - Blinkit']
bl_headers = [ws_bl.cell(row=4, column=c).value for c in range(1, 35)]

BL = []
for r in range(5, 200):
    prod = ws_bl.cell(row=r, column=1).value
    if prod is None: continue
    row = {}
    for c, h in enumerate(bl_headers, 1):
        if h:
            v = ws_bl.cell(row=r, column=c).value
            try: row[h] = float(v) if v is not None else None
            except: row[h] = v
    BL.append(row)

bl_prods = []
for p in BL:
    code = str(p.get('P. Breakdown', '') or '')
    disp = str(p.get('item_name', code) or code)
    units = num(p.get('Units (3 months)'))
    netrev_u = num(p.get('Net Revenue without GST'))
    gm_u = num(p.get('Gross Margin'))
    blmargin_u = num(p.get('Blinkit Margin'))
    mktg_u = num(p.get('Marketing Cost'))
    cm1_u = num(p.get('CM1'))
    cm2_u = num(p.get('CM2'))
    oh_u = num(p.get('Overhead'))
    ebitda_u = num(p.get('EBITDA'))
    mrp = num(p.get('MRP'))
    sp = num(p.get('SP'))

    if cm2_u > 0 and ebitda_u > 0:
        quad = 'star'; act = 'scale'
    elif cm2_u > 0:
        quad = 'overhead'; act = 'overhead'
    else:
        quad = 'loss'; act = 'cut'

    bl_prods.append({
        'key': code, 'disp': disp,
        'cat': str(p.get('Category', '') or ''),
        'units': units, 'mrp': mrp, 'sp': sp,
        'netrev_u': round(netrev_u, 2), 'netrev_t': round(netrev_u*units, 2),
        'gm_u': round(gm_u, 2), 'gm_t': round(gm_u*units, 2),
        'blmargin_u': round(blmargin_u, 2), 'blmargin_t': round(blmargin_u*units, 2),
        'mktg_u': round(mktg_u, 2), 'mktg_t': round(mktg_u*units, 2),
        'cm1_u': round(cm1_u, 2), 'cm1_t': round(cm1_u*units, 2),
        'cm2_u': round(cm2_u, 2), 'cm2_t': round(cm2_u*units, 2),
        'cm2pct': round(num(p.get('CM2%', 0))*100, 2),
        'oh_u': round(oh_u, 2), 'oh_t': round(oh_u*units, 2),
        'ebitda_u': round(ebitda_u, 2), 'ebitda_t': round(ebitda_u*units, 2),
        'ebitdapct': round(num(p.get('EBITDA%', 0))*100, 2),
        'quad': quad, 'act': act
    })

T_bl_units = sum(p['units'] for p in bl_prods)
T_bl_net_rev = sum(p['netrev_t'] for p in bl_prods)
T_bl_gm = sum(p['gm_t'] for p in bl_prods)
T_bl_bl_margin = sum(p['blmargin_t'] for p in bl_prods)
T_bl_mktg = sum(p['mktg_t'] for p in bl_prods)
T_bl_cm2 = sum(p['cm2_t'] for p in bl_prods)
T_bl_oh = sum(p['oh_t'] for p in bl_prods)
T_bl_ebitda = sum(p['ebitda_t'] for p in bl_prods)

bl_star = [p for p in bl_prods if p['quad']=='star']
bl_overhead = [p for p in bl_prods if p['quad']=='overhead']
bl_loss = [p for p in bl_prods if p['quad']=='loss']

D_BL = {
    'channel': 'Blinkit', 'periodLabel': 'Apr\u2013Jun 2026',
    'tot': {
        'rev': round(T_bl_net_rev, 2), 'netrev': round(T_bl_net_rev, 2),
        'units': int(T_bl_units), 'gm': round(T_bl_gm, 2),
        'mktg': round(T_bl_mktg, 2), 'blmargin': round(T_bl_bl_margin, 2),
        'cm2': round(T_bl_cm2, 2), 'oh': round(T_bl_oh, 2),
        'ebitda': round(T_bl_ebitda, 2),
        'cm2pct': round(T_bl_cm2/T_bl_net_rev*100, 5) if T_bl_net_rev else 0,
        'ebitdapct': round(T_bl_ebitda/T_bl_net_rev*100, 5) if T_bl_net_rev else 0,
        'gmpct': round(T_bl_gm/T_bl_net_rev*100, 5) if T_bl_net_rev else 0,
        'mktgpct': round(T_bl_mktg/T_bl_net_rev*100, 5) if T_bl_net_rev else 0,
        'n': len(bl_prods)
    },
    'quad': {
        'star': {'n': len(bl_star), 'rev': round(sum(p['netrev_t'] for p in bl_star), 2), 'units': 0, 'ebitda': 0},
        'overhead': {'n': len(bl_overhead), 'rev': round(sum(p['netrev_t'] for p in bl_overhead), 2), 'units': int(sum(p['units'] for p in bl_overhead)), 'ebitda': round(sum(p['ebitda_t'] for p in bl_overhead), 5)},
        'loss': {'n': len(bl_loss), 'rev': round(sum(p['netrev_t'] for p in bl_loss), 2), 'units': int(sum(p['units'] for p in bl_loss)), 'ebitda': round(sum(p['ebitda_t'] for p in bl_loss), 5)}
    },
    'products': bl_prods
}

print(f"D_BL: {len(bl_prods)} prods, stars={len(bl_star)}, overhead={len(bl_overhead)}, loss={len(bl_loss)}")
print(f"  Net Rev={T_bl_net_rev:,.0f}, CM2={T_bl_cm2:,.0f} ({T_bl_cm2/T_bl_net_rev*100:.1f}%), EBITDA={T_bl_ebitda:,.0f}")

# ============================================================
# INJECT INTO HTML
# ============================================================
with open('index.html', encoding='utf-8') as f:
    html = f.read()

# Build JSON strings
amz_json = json.dumps(D_AMZ, ensure_ascii=False, default=str)
fc_json = json.dumps(D_FC, ensure_ascii=False, default=str)
bl_json = json.dumps(D_BL, ensure_ascii=False, default=str)

# Replace D_AMZ
amz_start = html.find('const D_AMZ = {')
if amz_start != -1:
    amz_end = html.find(';\nconst D_FC', amz_start)
    if amz_end == -1:
        amz_end = html.find(';\r\nconst D_FC', amz_start)
    if amz_end != -1:
        html = html[:amz_start] + 'const D_AMZ = ' + amz_json + html[amz_end:]
        print("Replaced D_AMZ")
    else:
        print("ERROR: Could not find end of D_AMZ")
else:
    print("ERROR: Could not find D_AMZ")

# Replace D_FC
fc_start = html.find('const D_FC = {')
if fc_start != -1:
    fc_end = html.find(';\nconst D_BL', fc_start)
    if fc_end == -1:
        fc_end = html.find(';\r\nconst D_BL', fc_start)
    if fc_end != -1:
        html = html[:fc_start] + 'const D_FC = ' + fc_json + html[fc_end:]
        print("Replaced D_FC")
    else:
        print("ERROR: Could not find end of D_FC")
else:
    print("ERROR: Could not find D_FC")

# Replace D_BL
bl_start = html.find('const D_BL = {')
if bl_start != -1:
    # Find end - look for ; followed by const or \n\n
    bl_end = html.find(';\nconst ', bl_start)
    if bl_end == -1:
        bl_end = html.find(';\r\nconst ', bl_start)
    if bl_end == -1:
        # Try finding the end by counting braces
        i = bl_start + len('const D_BL = ')
        depth = 0
        in_str = False; esc = False
        while i < len(html):
            c = html[i]
            if esc: esc = False
            elif c == '\\' and in_str: esc = True
            elif c == '"' and not esc: in_str = not in_str
            elif not in_str:
                if c == '{': depth += 1
                elif c == '}':
                    depth -= 1
                    if depth == 0: bl_end = i + 1; break
            i += 1
    if bl_end != -1:
        # bl_end might point to right after }, need to include the semicolon
        # Check if bl_end is where ; is
        while bl_end < len(html) and html[bl_end] in ' \r\n': bl_end += 1
        if html[bl_end] == ';': bl_end += 1
        html = html[:bl_start] + 'const D_BL = ' + bl_json + ';' + html[bl_end:]
        print("Replaced D_BL")
    else:
        print("ERROR: Could not find end of D_BL")
else:
    print("ERROR: Could not find D_BL")

# Save
with open('index.html', 'w', encoding='utf-8') as f:
    f.write(html)
print("index.html saved.")
