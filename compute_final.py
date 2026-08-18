"""
Final complete rebuild of index.html with corrected data from Channel Economics Dashboard.xlsx
"""
import openpyxl
import json
import re

wb = openpyxl.load_workbook('Channel Economics Dashboard.xlsx', data_only=True)

def num(v):
    if v is None: return 0.0
    try:
        return float(v)
    except:
        return 0.0

def fmt_inr(v, sign=True):
    if v is None: return '—'
    s = '+' if (v > 0 and sign) else ''
    if abs(v) >= 100000:
        return f"{s}₹{v/100000:.2f} L"
    elif abs(v) >= 1000:
        return f"{s}₹{v:,.0f}"
    else:
        return f"{s}₹{v:.0f}"

def fmt_pct(v, decimals=1):
    if v is None: return '—'
    return f"{v*100:.{decimals}f}%"

def up_down(v, text=None):
    if text is None:
        text = fmt_inr(v)
    cls = 'up' if v >= 0 else 'down'
    return f'<span class="{cls}">{text}</span>'

# ============================================================
# WEBSITE
# ============================================================
ws_web = wb['Raw Data - Website']
web_headers = [ws_web.cell(row=4, column=c).value for c in range(1, 50)]

website_products = []
for r in range(5, 200):
    code = ws_web.cell(row=r, column=1).value
    if code is None: continue
    row = {}
    for c, h in enumerate(web_headers[:42], 1):
        if h:
            v = ws_web.cell(row=r, column=c).value
            try:
                row[h] = float(v) if v is not None else None
            except:
                row[h] = v
    website_products.append(row)

W = website_products
total_web_units = sum(num(p.get('Units sold')) for p in W)
total_web_rev_gst = sum(num(p.get('3 month revenue with GST')) for p in W)
total_web_rev_no_gst = sum(num(p.get('3 month revenue without GST')) for p in W)
total_web_net_rev = sum(num(p.get('Net Revenue without GST')) * num(p.get('Units sold')) for p in W)
total_web_cm2 = sum(num(p.get('CM2')) * num(p.get('Units sold')) for p in W)
total_web_ebita = sum(num(p.get('EBITA')) * num(p.get('Units sold')) for p in W)
total_web_mktg = sum(num(p.get('Marketing Cost')) * num(p.get('Units sold')) for p in W)
total_web_gm = sum(num(p.get('Gross margin')) * num(p.get('Units sold')) for p in W)
total_web_oh = sum(num(p.get('Overhead Cost')) * num(p.get('Units sold')) for p in W)

web_cm2pct = total_web_cm2/total_web_net_rev if total_web_net_rev else 0
web_ebitapct = total_web_ebita/total_web_net_rev if total_web_net_rev else 0
web_gmpct = total_web_gm/total_web_net_rev if total_web_net_rev else 0
web_mktg_pct = total_web_mktg/total_web_net_rev if total_web_net_rev else 0

web_stars = [p for p in W if num(p.get('CM2')) > 0 and num(p.get('EBITA')) > 0]
web_overhead_h = [p for p in W if num(p.get('CM2')) > 0 and num(p.get('EBITA')) <= 0]
web_loss = [p for p in W if num(p.get('CM2')) <= 0]

web_star_count = len(web_stars)
web_star_rev = sum(num(p.get('Net Revenue without GST'))*num(p.get('Units sold')) for p in web_stars)
web_star_ebitda_t = sum(num(p.get('EBITA'))*num(p.get('Units sold')) for p in web_stars)

# Top web stars by total EBITDA
web_stars_sorted = sorted(web_stars, key=lambda p: -num(p.get('EBITA'))*num(p.get('Units sold')))
# Top web overhead by closest to breakeven
web_oh_sorted = sorted(web_overhead_h, key=lambda p: abs(num(p.get('EBITA'))))

# ============================================================
# AMAZON
# ============================================================
ws_amz = wb['Raw Data - Amazon']
amz_headers = [ws_amz.cell(row=4, column=c).value for c in range(1, 50)]

amazon_products = []
for r in range(5, 200):
    prod = ws_amz.cell(row=r, column=1).value
    if prod is None: continue
    row = {}
    for c, h in enumerate(amz_headers[:45], 1):
        if h:
            v = ws_amz.cell(row=r, column=c).value
            try:
                row[h] = float(v) if v is not None else None
            except:
                row[h] = v
    amazon_products.append(row)

A = amazon_products
total_amz_units = sum(num(p.get('Units Sold')) for p in A)
total_amz_net_rev = sum(num(p.get('Net Revenue without GST'))*num(p.get('Units Sold')) for p in A)
total_amz_cm2 = sum(num(p.get('CM2'))*num(p.get('Units Sold')) for p in A)
total_amz_ebitda = sum(num(p.get('EBITDA'))*num(p.get('Units Sold')) for p in A)
total_amz_mktg = sum(num(p.get('Cost of Advertising'))*num(p.get('Units Sold')) for p in A)
total_amz_gm = sum(num(p.get('Gross Margin'))*num(p.get('Units Sold')) for p in A)

amz_cm2pct = total_amz_cm2/total_amz_net_rev if total_amz_net_rev else 0
amz_ebitdapct = total_amz_ebitda/total_amz_net_rev if total_amz_net_rev else 0
amz_gmpct = total_amz_gm/total_amz_net_rev if total_amz_net_rev else 0
amz_mktg_pct = total_amz_mktg/total_amz_net_rev if total_amz_net_rev else 0

amz_stars = [p for p in A if num(p.get('CM2')) > 0 and num(p.get('EBITDA')) > 0]
amz_overhead_h = [p for p in A if num(p.get('CM2')) > 0 and num(p.get('EBITDA')) <= 0]
amz_loss = [p for p in A if num(p.get('CM2')) <= 0]

amz_star_count = len(amz_stars)
amz_star_rev = sum(num(p.get('Net Revenue without GST'))*num(p.get('Units Sold')) for p in amz_stars)
amz_star_ebitda_t = sum(num(p.get('EBITDA'))*num(p.get('Units Sold')) for p in amz_stars)

amz_stars_sorted = sorted(amz_stars, key=lambda p: -num(p.get('EBITDA'))*num(p.get('Units Sold')))
amz_oh_sorted = sorted(amz_overhead_h, key=lambda p: abs(num(p.get('EBITDA'))))

# ============================================================
# FIRSTCRY
# ============================================================
ws_fc = wb['Raw Data - FirstCry']
fc_headers = [ws_fc.cell(row=4, column=c).value for c in range(1, 35)]

firstcry_products = []
for r in range(5, 200):
    prod = ws_fc.cell(row=r, column=1).value
    if prod is None: continue
    row = {}
    for c, h in enumerate(fc_headers, 1):
        if h:
            v = ws_fc.cell(row=r, column=c).value
            try:
                row[h] = float(v) if v is not None else None
            except:
                row[h] = v
    firstcry_products.append(row)

FC = [p for p in firstcry_products if num(p.get('Units (3 months)')) > 0]
total_fc_units = sum(num(p.get('Units (3 months)')) for p in FC)
total_fc_net_rev = sum(num(p.get('Total Net Revenue')) for p in FC)
total_fc_cm2 = sum(num(p.get('CM2'))*num(p.get('Units (3 months)')) for p in FC)
total_fc_ebitda = sum(num(p.get('EBITDA'))*num(p.get('Units (3 months)')) for p in FC)
total_fc_gm = sum(num(p.get('Gross Margin'))*num(p.get('Units (3 months)')) for p in FC)
total_fc_platform = sum(num(p.get('Firstcry Margin'))*num(p.get('Units (3 months)')) for p in FC)

fc_cm2pct = total_fc_cm2/total_fc_net_rev if total_fc_net_rev else 0
fc_ebitdapct = total_fc_ebitda/total_fc_net_rev if total_fc_net_rev else 0
fc_gmpct = total_fc_gm/total_fc_net_rev if total_fc_net_rev else 0
fc_platform_pct = total_fc_platform/total_fc_net_rev if total_fc_net_rev else 0

fc_stars = [p for p in FC if num(p.get('CM2')) > 0 and num(p.get('EBITDA')) > 0]
fc_overhead_h = [p for p in FC if num(p.get('CM2')) > 0 and num(p.get('EBITDA')) <= 0]

fc_star_count = len(fc_stars)
fc_star_rev = sum(num(p.get('Total Net Revenue')) for p in fc_stars)
fc_star_ebitda_t = sum(num(p.get('EBITDA'))*num(p.get('Units (3 months)')) for p in fc_stars)

fc_stars_sorted = sorted(fc_stars, key=lambda p: -num(p.get('EBITDA'))*num(p.get('Units (3 months)')))
fc_oh_sorted = sorted(fc_overhead_h, key=lambda p: abs(num(p.get('EBITDA'))))

# ============================================================
# BLINKIT
# ============================================================
ws_bl = wb['Raw Data - Blinkit']
bl_headers = [ws_bl.cell(row=4, column=c).value for c in range(1, 35)]

blinkit_products = []
for r in range(5, 200):
    prod = ws_bl.cell(row=r, column=1).value
    if prod is None: continue
    row = {}
    for c, h in enumerate(bl_headers, 1):
        if h:
            v = ws_bl.cell(row=r, column=c).value
            try:
                row[h] = float(v) if v is not None else None
            except:
                row[h] = v
    blinkit_products.append(row)

BL = blinkit_products
total_bl_units = sum(num(p.get('Units (3 months)')) for p in BL)
total_bl_net_rev = sum(num(p.get('Net Revenue without GST'))*num(p.get('Units (3 months)')) for p in BL)
total_bl_cm2 = sum(num(p.get('CM2'))*num(p.get('Units (3 months)')) for p in BL)
total_bl_ebitda = sum(num(p.get('EBITDA'))*num(p.get('Units (3 months)')) for p in BL)
total_bl_mktg = sum(num(p.get('Marketing Cost'))*num(p.get('Units (3 months)')) for p in BL)
total_bl_gm = sum(num(p.get('Gross Margin'))*num(p.get('Units (3 months)')) for p in BL)
total_bl_blmargin = sum(num(p.get('Blinkit Margin'))*num(p.get('Units (3 months)')) for p in BL)

bl_cm2pct = total_bl_cm2/total_bl_net_rev if total_bl_net_rev else 0
bl_ebitdapct = total_bl_ebitda/total_bl_net_rev if total_bl_net_rev else 0
bl_gmpct = total_bl_gm/total_bl_net_rev if total_bl_net_rev else 0
bl_mktg_pct = total_bl_mktg/total_bl_net_rev if total_bl_net_rev else 0
bl_platform_pct = total_bl_blmargin/total_bl_net_rev if total_bl_net_rev else 0

bl_stars = [p for p in BL if num(p.get('CM2')) > 0 and num(p.get('EBITDA')) > 0]
bl_overhead_h = [p for p in BL if num(p.get('CM2')) > 0 and num(p.get('EBITDA')) <= 0]

bl_star_count = len(bl_stars)

# Blinkit only star for overhead?
bl_cm2pos = [p for p in BL if num(p.get('CM2')) > 0]  # Only B8

print("=== DATA VERIFIED ===")
print(f"Website: {len(W)} SKUs, units={int(total_web_units)}")
print(f"  Net Rev={total_web_net_rev:,.0f}, CM2={total_web_cm2:,.0f} ({web_cm2pct*100:.1f}%), EBITA={total_web_ebita:,.0f} ({web_ebitapct*100:.1f}%)")
print(f"  Stars={web_star_count}/{len(W)}")
print(f"Amazon: {len(A)} SKUs, units={int(total_amz_units)}")
print(f"  Net Rev={total_amz_net_rev:,.0f}, CM2={total_amz_cm2:,.0f} ({amz_cm2pct*100:.1f}%), EBITDA={total_amz_ebitda:,.0f} ({amz_ebitdapct*100:.1f}%)")
print(f"  Stars={amz_star_count}/{len(A)}")
print(f"FirstCry: {len(FC)} active SKUs, units={int(total_fc_units)}")
print(f"  Net Rev={total_fc_net_rev:,.0f}, CM2={total_fc_cm2:,.0f} ({fc_cm2pct*100:.1f}%), EBITDA={total_fc_ebitda:,.0f} ({fc_ebitdapct*100:.1f}%)")
print(f"  Stars={fc_star_count}/{len(FC)}")
print(f"Blinkit: {len(BL)} SKUs, units={int(total_bl_units)}")
print(f"  Net Rev={total_bl_net_rev:,.0f}, CM2={total_bl_cm2:,.0f} ({bl_cm2pct*100:.1f}%), EBITDA={total_bl_ebitda:,.0f} ({bl_ebitdapct*100:.1f}%)")
print(f"  Stars={bl_star_count}/{len(BL)}")

print("\nBlinkit CM2+ products:")
for p in bl_cm2pos:
    print(f"  {p.get('P. Breakdown')}: CM2={num(p.get('CM2')):.0f}, EBITDA={num(p.get('EBITDA')):.0f}")

print("\nWeb Stars (top5):")
for p in web_stars_sorted[:5]:
    code = p.get('Product Name (Code)','')
    cm2 = num(p.get('CM2'))
    ebita = num(p.get('EBITA'))
    units = num(p.get('Units sold'))
    print(f"  {code}: CM2/u={cm2:.0f} EBITA/u={ebita:.0f} units={units:.0f}")

print("\nAMZ Stars (top5):")
for p in amz_stars_sorted[:5]:
    code = p.get('P. Breakdown','')
    cm2 = num(p.get('CM2'))
    ebitda = num(p.get('EBITDA'))
    units = num(p.get('Units Sold'))
    print(f"  {code}: CM2/u={cm2:.0f} EBITDA/u={ebitda:.0f} units={units:.0f}")

print("\nFC Stars (top5):")
for p in fc_stars_sorted[:5]:
    code = p.get('P. Breakdown','')
    cm2 = num(p.get('CM2'))
    ebitda = num(p.get('EBITDA'))
    units = num(p.get('Units (3 months)'))
    print(f"  {code}: CM2/u={cm2:.0f} EBITDA/u={ebitda:.0f} units={units:.0f}")

print("\nWeb overhead-heavy (closest to breakeven):")
for p in web_oh_sorted[:4]:
    code = p.get('Product Name (Code)','')
    cm2 = num(p.get('CM2'))
    ebita = num(p.get('EBITA'))
    print(f"  {code}: CM2/u={cm2:.0f} EBITA/u={ebita:.0f} gap={-ebita:.0f}")

print("\nAMZ overhead-heavy (closest):")
for p in amz_oh_sorted[:4]:
    code = p.get('P. Breakdown','')
    cm2 = num(p.get('CM2'))
    ebitda = num(p.get('EBITDA'))
    print(f"  {code}: CM2/u={cm2:.0f} EBITDA/u={ebitda:.0f} gap={-ebitda:.0f}")

print("\nFC overhead-heavy (closest):")
for p in fc_oh_sorted[:4]:
    code = p.get('P. Breakdown','')
    cm2 = num(p.get('CM2'))
    ebitda = num(p.get('EBITDA'))
    print(f"  {code}: CM2/u={cm2:.0f} EBITDA/u={ebitda:.0f} gap={-ebitda:.0f}")

# --- Website worst performers (discontinue candidates)
web_worst = sorted(W, key=lambda p: num(p.get('EBITA'))*num(p.get('Units sold')))
print("\nWeb worst (discontinue candidates):")
for p in web_worst[:4]:
    code = p.get('Product Name (Code)','')
    units = num(p.get('Units sold'))
    ebita_total = num(p.get('EBITA'))*units
    net_rev_total = num(p.get('Net Revenue without GST'))*units
    loss_pct = -ebita_total/net_rev_total*100 if net_rev_total > 0 else 0
    print(f"  {code}: units={units:.0f}, EBITA total={ebita_total:,.0f}, loss%={loss_pct:.0f}%")

# --- Amazon worst performers
amz_worst = sorted(A, key=lambda p: num(p.get('EBITDA'))*num(p.get('Units Sold')))
print("\nAMZ worst (discontinue candidates):")
for p in amz_worst[:4]:
    code = p.get('P. Breakdown','')
    units = num(p.get('Units Sold'))
    ebitda_total = num(p.get('EBITDA'))*units
    net_rev_total = num(p.get('Net Revenue without GST'))*units
    loss_pct = -ebitda_total/net_rev_total*100 if net_rev_total > 0 else 0
    print(f"  {code}: units={units:.0f}, EBITDA total={ebitda_total:,.0f}, loss%={loss_pct:.0f}%")

# FC worst
fc_worst = sorted(FC, key=lambda p: num(p.get('EBITDA'))*num(p.get('Units (3 months)')))
print("\nFC worst (discontinue):")
for p in fc_worst[:3]:
    code = p.get('P. Breakdown','')
    units = num(p.get('Units (3 months)'))
    ebitda_total = num(p.get('EBITDA'))*units
    net_rev = num(p.get('Total Net Revenue'))
    loss_pct = -ebitda_total/net_rev*100 if net_rev > 0 else 0
    print(f"  {code}: units={units:.0f}, EBITDA total={ebitda_total:,.0f}, loss%={loss_pct:.0f}%")

# BL all 
print("\nBlinkit all:")
for p in sorted(BL, key=lambda p: num(p.get('EBITDA'))*num(p.get('Units (3 months)'))):
    code = p.get('P. Breakdown','')
    units = num(p.get('Units (3 months)'))
    ebitda = num(p.get('EBITDA'))
    ebitda_t = ebitda*units
    net_rev = num(p.get('Net Revenue without GST'))*units
    loss_pct = -ebitda_t/net_rev*100 if net_rev > 0 else 0
    print(f"  {code}: units={units:.0f}, EBITDA/u={ebitda:.0f}, total={ebitda_t:,.0f}, loss%={loss_pct:.0f}%")

# Marketing spend comparison (for new tab)
print("\n=== MARKETING SPEND TAB ===")
print(f"Website 3-month Rev WITH GST: ₹{total_web_rev_gst:,.0f}")
print(f"Website 3-month Rev WITHOUT GST: ₹{total_web_rev_no_gst:,.0f}")
print(f"Website Marketing Spend: ₹{total_web_mktg:,.0f}")
mktg_pct_with_gst = total_web_mktg/total_web_rev_gst*100
mktg_pct_without_gst = total_web_mktg/total_web_rev_no_gst*100
print(f"Mktg as % of Rev WITH GST: {mktg_pct_with_gst:.1f}%")
print(f"Mktg as % of Rev WITHOUT GST: {mktg_pct_without_gst:.1f}%")
print(f"GST impact on mktg%: +{mktg_pct_without_gst - mktg_pct_with_gst:.1f}pp")
