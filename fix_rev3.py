"""
Final fix: pull revenue from BOTH FBA (Combined sheet col35) AND FBM (sheet col8*col4)
for every P.Breakdown key. Sum them where the same key appears in both.
"""
import openpyxl, warnings, json, re
warnings.filterwarnings('ignore')

def flt(v, d=0.0):
    try: return float(v) if v is not None else d
    except: return d
def sint(v):
    try: return int(float(v)) if v is not None else 0
    except: return 0

# ════ AMAZON: Combined (FBA) + FBM sheets ════════════════════════════════════
wb_a = openpyxl.load_workbook('Unit cost economics AMAZON April-June (2).xlsx', data_only=True)

# FBA side: Combined sheet col35 = Total Net Revenue (already 3-month total)
# Using MRP*units as gross since col35 is net; actual gross = MRP*units
ws_fba = wb_a['Final Sheet (Combined)']
amz_rev   = {}
amz_units = {}
for r in range(2, ws_fba.max_row+1):
    pb    = ws_fba.cell(r, 2).value
    mrp   = flt(ws_fba.cell(r, 6).value)
    units = sint(ws_fba.cell(r, 33).value)
    tot_nr= flt(ws_fba.cell(r, 35).value)   # total net revenue
    if not pb or units == 0: continue
    pb = str(pb).strip()
    # Use Total Net Revenue (col35) — this is what the Excel shows as actual revenue
    amz_rev[pb]   = amz_rev.get(pb, 0)   + tot_nr
    amz_units[pb] = amz_units.get(pb, 0) + units

# FBM side: col4=units, col8=MRP, col11=MRP-GST*units (net revenue ex-GST total)
# Revenue = MRP * units (gross) is cleanest; col11 = net ex-GST total
ws_fbm = wb_a['Final Sheet (FBM)']
fbm_rev   = {}
fbm_units = {}
for r in range(2, ws_fbm.max_row+1):
    pb    = ws_fbm.cell(r, 2).value
    mrp   = flt(ws_fbm.cell(r, 8).value)
    units = sint(ws_fbm.cell(r, 4).value)
    nr_t  = flt(ws_fbm.cell(r, 11).value)   # MRP-GST*units (net ex-gst total)
    if not pb or units == 0: continue
    pb = str(pb).strip()
    # Gross = MRP * units for FBM products
    fbm_rev[pb]   = fbm_rev.get(pb, 0)   + mrp * units
    fbm_units[pb] = fbm_units.get(pb, 0) + units

# Merge: for keys that appear only in FBM, use FBM gross
all_keys = set(amz_rev.keys()) | set(fbm_rev.keys())
merged_rev   = {}
merged_units = {}
for k in all_keys:
    merged_rev[k]   = amz_rev.get(k, 0) + fbm_rev.get(k, 0)
    merged_units[k] = amz_units.get(k, 0) + fbm_units.get(k, 0)

print('Amazon merged revenue map (top 10):')
for k, v in sorted(merged_rev.items(), key=lambda x: -x[1])[:10]:
    print(f'  {k!r:30s}  rev=₹{v:>12,.0f}  units={merged_units[k]:>5}')
print(f'Total keys: {len(merged_rev)}')

# ════ FIRSTCRY ════════════════════════════════════════════════════════════════
wb_fc = openpyxl.load_workbook('Firstcry_Unit Cost Economic(Apr-June) (1).xlsx', data_only=True)
ws_fc = wb_fc['SP Vendor Format']
fc_rev = {}
for r in range(2, ws_fc.max_row+1):
    pb    = ws_fc.cell(r, 4).value
    units = sint(ws_fc.cell(r, 10).value)
    sp    = flt(ws_fc.cell(r, 11).value)
    if not pb or units == 0: continue
    pb = str(pb).strip()
    fc_rev[pb] = round(fc_rev.get(pb, 0) + sp * units, 2)

# ════ BLINKIT ════════════════════════════════════════════════════════════════
wb_bl = openpyxl.load_workbook('Blinkit_Unit Cost Economic(Apr-June) (1).xlsx', data_only=True)
ws_bl = wb_bl['Final(Unit Cost Format)']
bl_rev = {}
for r in range(2, ws_bl.max_row+1):
    pb    = ws_bl.cell(r, 3).value
    units = sint(ws_bl.cell(r, 5).value)
    sp    = flt(ws_bl.cell(r, 7).value)
    if not pb or units == 0: continue
    pb = str(pb).strip()
    bl_rev[pb] = round(bl_rev.get(pb, 0) + sp * units, 2)

# ════ Patch index.html ═══════════════════════════════════════════════════════
with open('index.html', encoding='utf-8') as f:
    html = f.read()

def load_var(html, var_name):
    m = re.search('const ' + var_name + r' = (\{.*?\});', html, re.DOTALL)
    return json.loads(m.group(1))

def save_var(html, var_name, obj):
    new_js = json.dumps(obj, ensure_ascii=False, separators=(',', ':'))
    start_tok = 'const ' + var_name + ' = {'
    idx_start = html.find(start_tok)
    depth = 0
    i = idx_start + len('const ' + var_name + ' = ')
    while i < len(html):
        c = html[i]
        if c == '{': depth += 1
        elif c == '}':
            depth -= 1
            if depth == 0:
                idx_end = i + 1
                break
        i += 1
    s = html.find(';', idx_end)
    if s != -1 and s < idx_end + 3:
        idx_end = s + 1
    new = 'const ' + var_name + ' = ' + new_js + ';'
    old = html[idx_start:idx_end]
    print(f'{var_name}: {len(old):,} -> {len(new):,} chars')
    return html[:idx_start] + new + html[idx_end:]

# ── D_AMZ ────────────────────────────────────────────────────────────────────
amz = load_var(html, 'D_AMZ')
unmatched_amz = []
for p in amz['products']:
    k = p['key']
    if k in merged_rev:
        p['rev_t'] = round(merged_rev[k], 2)
        p['units'] = merged_units[k]   # sync units with Excel
    else:
        # last fallback: mrp * stored units
        fallback = round(flt(p.get('mrp', 0)) * flt(p.get('units', 0)), 2)
        p['rev_t'] = fallback
        unmatched_amz.append(f'{k}(fallback={fallback:,.0f})')

print(f'\nAMZ: {len(amz["products"])} products')
if unmatched_amz:
    print(f'  Still unmatched (using MRP×units fallback): {unmatched_amz}')

# Recompute AMZ totals
prods = amz['products']
amz['tot']['rev']    = round(sum(p['rev_t']            for p in prods), 2)
amz['tot']['netrev'] = round(sum(p.get('netrev_t', p['rev_t']) for p in prods), 2)
amz['tot']['units']  = sum(p.get('units', 0)           for p in prods)
amz['tot']['cm2']    = round(sum(p.get('cm2_t', 0)     for p in prods), 2)
amz['tot']['ebitda'] = round(sum(p.get('ebitda_t', 0)  for p in prods), 2)
amz['tot']['mktg']   = round(sum(p.get('mktg_t', 0)    for p in prods), 2)
amz['tot']['gm']     = round(sum(p.get('gm_t', 0)      for p in prods), 2)
amz['tot']['oh']     = round(sum(p.get('oh_t', 0)      for p in prods), 2)
amz['tot']['n']      = len(prods)
nr = amz['tot']['netrev']
amz['tot']['cm2pct']    = round(amz['tot']['cm2'] / nr * 100, 5) if nr else 0
amz['tot']['ebitdapct'] = round(amz['tot']['ebitda'] / nr * 100, 5) if nr else 0
amz['tot']['gmpct']     = round(amz['tot']['gm'] / nr, 7) if nr else 0

for q in ['star', 'overhead', 'loss']:
    qp = [p for p in prods if p.get('quad') == q]
    amz['quad'][q] = {
        'n': len(qp),
        'rev': round(sum(p['rev_t'] for p in qp), 2),
        'units': sum(p.get('units', 0) for p in qp),
        'ebitda': round(sum(p.get('ebitda_t', 0) for p in qp), 2),
    }

# ── D_FC ─────────────────────────────────────────────────────────────────────
fc = load_var(html, 'D_FC')
unmatched_fc = []
for p in fc['products']:
    k = p.get('key', '')
    if k in fc_rev:
        p['rev_t'] = fc_rev[k]
    else:
        unmatched_fc.append(k)
fc['tot']['rev'] = round(sum(p.get('rev_t', 0) for p in fc['products']), 2)
if unmatched_fc:
    print(f'FC unmatched: {unmatched_fc}')

# ── D_BL ─────────────────────────────────────────────────────────────────────
bl = load_var(html, 'D_BL')
unmatched_bl = []
for p in bl['products']:
    k = p.get('key', '')
    if k in bl_rev:
        p['rev_t'] = bl_rev[k]
    else:
        unmatched_bl.append(k)
bl['tot']['rev'] = round(sum(p.get('rev_t', 0) for p in bl['products']), 2)
if unmatched_bl:
    print(f'BL unmatched: {unmatched_bl}')

# ── Save ─────────────────────────────────────────────────────────────────────
html = save_var(html, 'D_AMZ', amz)
html = save_var(html, 'D_FC',  fc)
html = save_var(html, 'D_BL',  bl)

with open('index.html', 'w', encoding='utf-8') as f:
    f.write(html)
print('\nindex.html written. Size:', len(html))

# ════ Final verification ═════════════════════════════════════════════════════
print('\n=== FINAL REVENUE ===')
for var in ['D_AMZ', 'D_FC', 'D_BL']:
    m = re.search('const ' + var + r' = (\{.*?\});', html, re.DOTALL)
    d = json.loads(m.group(1))
    prods = d['products']
    zero = [p['key'] for p in prods if p.get('rev_t', 0) <= 0]
    print(f'\n{var}:  {len(prods)} products   tot.rev = ₹{d["tot"]["rev"]:>12,.0f}')
    for p in sorted(prods, key=lambda x: -x.get('rev_t', 0))[:5]:
        print(f'  {p["key"]!r:28s}  ₹{p["rev_t"]:>12,.0f}  ({p.get("units",0)} units × ₹{p.get("mrp",0)} MRP)')
    if zero:
        print(f'  ⚠ zero rev_t: {zero}')

m4 = re.search(r'const D = (\{.*?\});', html, re.DOTALL)
d4 = json.loads(m4.group(1))
print(f'\nWebsite D intact: apr.rev=₹{d4["periods"]["apr"]["tot"]["rev"]:,.2f}')
print('ALL OK' if not any(
    p.get('rev_t', 0) <= 0
    for var in ['D_AMZ','D_FC','D_BL']
    for p in json.loads(re.search('const '+var+r' = (\{.*?\});', html, re.DOTALL).group(1))['products']
) else 'SOME PRODUCTS STILL HAVE ZERO REV')
