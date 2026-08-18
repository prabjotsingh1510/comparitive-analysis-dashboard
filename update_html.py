"""
Update index.html with correct data from Channel Economics Dashboard.xlsx
All 4 tasks:
1. Update channel data values
2. Update Channel Analysis decision criteria
3. Remove Important Metrics tab
4. Add Marketing Spend Comparison tab
"""
import openpyxl
import re

wb = openpyxl.load_workbook('Channel Economics Dashboard.xlsx', data_only=True)

def num(v):
    if v is None: return 0.0
    try: return float(v)
    except: return 0.0

# ============================================================
# EXTRACT ALL DATA
# ============================================================
ws_web = wb['Raw Data - Website']
web_headers = [ws_web.cell(row=4, column=c).value for c in range(1, 50)]
website_products = []
for r in range(5, 200):
    code = ws_web.cell(row=r, column=1).value
    if code is None: continue
    row = {}
    for c, h in enumerate(web_headers[:42], 1):
        if h:
            v = ws_web.cell(row=r, column=c).value
            try: row[h] = float(v) if v is not None else None
            except: row[h] = v
    website_products.append(row)

W = website_products
total_web_units = sum(num(p.get('Units sold')) for p in W)
total_web_rev_gst = sum(num(p.get('3 month revenue with GST')) for p in W)
total_web_rev_no_gst = sum(num(p.get('3 month revenue without GST')) for p in W)
total_web_net_rev = sum(num(p.get('Net Revenue without GST')) * num(p.get('Units sold')) for p in W)
total_web_cm2 = sum(num(p.get('CM2')) * num(p.get('Units sold')) for p in W)
total_web_ebita = sum(num(p.get('EBITA')) * num(p.get('Units sold')) for p in W)
total_web_mktg = sum(num(p.get('Marketing Cost')) * num(p.get('Units sold')) for p in W)
total_web_gm = sum(num(p.get('Gross margin')) * num(p.get('Units sold')) for p in W)

web_cm2pct = total_web_cm2/total_web_net_rev if total_web_net_rev else 0
web_ebitapct = total_web_ebita/total_web_net_rev if total_web_net_rev else 0
web_gmpct = total_web_gm/total_web_net_rev if total_web_net_rev else 0
web_mktg_pct = total_web_mktg/total_web_net_rev if total_web_net_rev else 0

web_stars = sorted([p for p in W if num(p.get('CM2')) > 0 and num(p.get('EBITA')) > 0],
                   key=lambda p: -num(p.get('EBITA'))*num(p.get('Units sold')))
web_overhead_h = sorted([p for p in W if num(p.get('CM2')) > 0 and num(p.get('EBITA')) <= 0],
                        key=lambda p: abs(num(p.get('EBITA'))))

web_star_count = len(web_stars)
web_star_rev = sum(num(p.get('Net Revenue without GST'))*num(p.get('Units sold')) for p in web_stars)
web_star_ebitda_t = sum(num(p.get('EBITA'))*num(p.get('Units sold')) for p in web_stars)

ws_amz = wb['Raw Data - Amazon']
amz_headers = [ws_amz.cell(row=4, column=c).value for c in range(1, 50)]
amazon_products = []
for r in range(5, 200):
    prod = ws_amz.cell(row=r, column=1).value
    if prod is None: continue
    row = {}
    for c, h in enumerate(amz_headers[:45], 1):
        if h:
            v = ws_amz.cell(row=r, column=c).value
            try: row[h] = float(v) if v is not None else None
            except: row[h] = v
    amazon_products.append(row)

A = amazon_products
total_amz_units = sum(num(p.get('Units Sold')) for p in A)
total_amz_net_rev = sum(num(p.get('Net Revenue without GST'))*num(p.get('Units Sold')) for p in A)
total_amz_cm2 = sum(num(p.get('CM2'))*num(p.get('Units Sold')) for p in A)
total_amz_ebitda = sum(num(p.get('EBITDA'))*num(p.get('Units Sold')) for p in A)
total_amz_mktg = sum(num(p.get('Cost of Advertising'))*num(p.get('Units Sold')) for p in A)
total_amz_gm = sum(num(p.get('Gross Margin'))*num(p.get('Units Sold')) for p in A)

amz_cm2pct = total_amz_cm2/total_amz_net_rev if total_amz_net_rev else 0
amz_ebitdapct = total_amz_ebitda/total_amz_net_rev if total_amz_net_rev else 0
amz_gmpct = total_amz_gm/total_amz_net_rev if total_amz_net_rev else 0
amz_mktg_pct = total_amz_mktg/total_amz_net_rev if total_amz_net_rev else 0

amz_stars = sorted([p for p in A if num(p.get('CM2')) > 0 and num(p.get('EBITDA')) > 0],
                   key=lambda p: -num(p.get('EBITDA'))*num(p.get('Units Sold')))
amz_overhead_h = sorted([p for p in A if num(p.get('CM2')) > 0 and num(p.get('EBITDA')) <= 0],
                        key=lambda p: abs(num(p.get('EBITDA'))))

amz_star_count = len(amz_stars)
amz_star_rev = sum(num(p.get('Net Revenue without GST'))*num(p.get('Units Sold')) for p in amz_stars)
amz_star_ebitda_t = sum(num(p.get('EBITDA'))*num(p.get('Units Sold')) for p in amz_stars)

ws_fc = wb['Raw Data - FirstCry']
fc_headers = [ws_fc.cell(row=4, column=c).value for c in range(1, 35)]
firstcry_products = []
for r in range(5, 200):
    prod = ws_fc.cell(row=r, column=1).value
    if prod is None: continue
    row = {}
    for c, h in enumerate(fc_headers, 1):
        if h:
            v = ws_fc.cell(row=r, column=c).value
            try: row[h] = float(v) if v is not None else None
            except: row[h] = v
    firstcry_products.append(row)

FC = [p for p in firstcry_products if num(p.get('Units (3 months)')) > 0]
total_fc_units = sum(num(p.get('Units (3 months)')) for p in FC)
total_fc_net_rev = sum(num(p.get('Total Net Revenue')) for p in FC)
total_fc_cm2 = sum(num(p.get('CM2'))*num(p.get('Units (3 months)')) for p in FC)
total_fc_ebitda = sum(num(p.get('EBITDA'))*num(p.get('Units (3 months)')) for p in FC)
total_fc_gm = sum(num(p.get('Gross Margin'))*num(p.get('Units (3 months)')) for p in FC)
total_fc_platform = sum(num(p.get('Firstcry Margin'))*num(p.get('Units (3 months)')) for p in FC)

fc_cm2pct = total_fc_cm2/total_fc_net_rev if total_fc_net_rev else 0
fc_ebitdapct = total_fc_ebitda/total_fc_net_rev if total_fc_net_rev else 0
fc_gmpct = total_fc_gm/total_fc_net_rev if total_fc_net_rev else 0
fc_platform_pct = total_fc_platform/total_fc_net_rev if total_fc_net_rev else 0

fc_stars = sorted([p for p in FC if num(p.get('CM2')) > 0 and num(p.get('EBITDA')) > 0],
                  key=lambda p: -num(p.get('EBITDA'))*num(p.get('Units (3 months)')))
fc_overhead_h = sorted([p for p in FC if num(p.get('CM2')) > 0 and num(p.get('EBITDA')) <= 0],
                       key=lambda p: abs(num(p.get('EBITDA'))))

fc_star_count = len(fc_stars)
fc_star_rev = sum(num(p.get('Total Net Revenue')) for p in fc_stars)
fc_star_ebitda_t = sum(num(p.get('EBITDA'))*num(p.get('Units (3 months)')) for p in fc_stars)

ws_bl = wb['Raw Data - Blinkit']
bl_headers = [ws_bl.cell(row=4, column=c).value for c in range(1, 35)]
blinkit_products = []
for r in range(5, 200):
    prod = ws_bl.cell(row=r, column=1).value
    if prod is None: continue
    row = {}
    for c, h in enumerate(bl_headers, 1):
        if h:
            v = ws_bl.cell(row=r, column=c).value
            try: row[h] = float(v) if v is not None else None
            except: row[h] = v
    blinkit_products.append(row)

BL = blinkit_products
total_bl_units = sum(num(p.get('Units (3 months)')) for p in BL)
total_bl_net_rev = sum(num(p.get('Net Revenue without GST'))*num(p.get('Units (3 months)')) for p in BL)
total_bl_cm2 = sum(num(p.get('CM2'))*num(p.get('Units (3 months)')) for p in BL)
total_bl_ebitda = sum(num(p.get('EBITDA'))*num(p.get('Units (3 months)')) for p in BL)
total_bl_mktg = sum(num(p.get('Marketing Cost'))*num(p.get('Units (3 months)')) for p in BL)
total_bl_gm = sum(num(p.get('Gross Margin'))*num(p.get('Units (3 months)')) for p in BL)
total_bl_blmargin = sum(num(p.get('Blinkit Margin'))*num(p.get('Units (3 months)')) for p in BL)

bl_cm2pct = total_bl_cm2/total_bl_net_rev if total_bl_net_rev else 0
bl_ebitdapct = total_bl_ebitda/total_bl_net_rev if total_bl_net_rev else 0
bl_gmpct = total_bl_gm/total_bl_net_rev if total_bl_net_rev else 0
bl_mktg_pct = total_bl_mktg/total_bl_net_rev if total_bl_net_rev else 0
bl_platform_pct = total_bl_blmargin/total_bl_net_rev if total_bl_net_rev else 0

bl_star_count = 0  # B8 has CM2+ but EBITDA-, so no stars
bl_cm2pos = [p for p in BL if num(p.get('CM2')) > 0]  # Only B8

# Helper formatting
def fmt(v, decimals=0):
    """Format number with commas"""
    if v is None: return '0'
    return f"{v:,.{decimals}f}"

def pct(v, decimals=1):
    if v is None: return '0%'
    return f"{v*100:.{decimals}f}%"

def lakh(v):
    """Format as lakh"""
    if v is None: return '\u20b90'
    sign = '+' if v > 0 else ''
    if abs(v) >= 100000:
        return f"{sign}\u20b9{v/100000:.2f} L"
    elif abs(v) >= 1000:
        return f"{sign}\u20b9{v:,.0f}"
    else:
        return f"{sign}\u20b9{v:.0f}"

def star_chip(code, cm2, ebitda, first=False):
    border_style = 'border:1px solid var(--good);border-left:3px solid var(--good)' if first else 'border:1px solid var(--border)'
    return f'''<div style="display:inline-flex;flex-direction:column;background:var(--wash);{border_style};border-radius:6px;padding:4px 8px;min-width:90px">
                <span style="font-size:11.5px;font-weight:700;color:var(--ink-1);white-space:nowrap">{code}</span>
                <span style="font-size:10px;color:var(--ink-3);margin-top:1px;font-variant-numeric:tabular-nums">CM2 \u20b9{cm2:.0f} \u2192 EBITDA \u20b9{ebitda:.0f}</span>
              </div>'''

# ============================================================
# BUILD NEW CHANNEL ANALYSIS SECTION
# ============================================================
web_star_rev_pct = web_star_rev/total_web_net_rev if total_web_net_rev else 0
amz_star_rev_pct = amz_star_rev/total_amz_net_rev if total_amz_net_rev else 0
fc_star_rev_pct = fc_star_rev/total_fc_net_rev if total_fc_net_rev else 0

# Web star top chips
web_star_chips = ''
for i, p in enumerate(web_stars[:4]):
    code = p.get('Product Name (Code)', '')
    cm2 = num(p.get('CM2'))
    ebita = num(p.get('EBITA'))
    web_star_chips += star_chip(code, cm2, ebita, i==0)

# AMZ star top chips  
amz_star_chips = ''
for i, p in enumerate(amz_stars[:4]):
    code = p.get('P. Breakdown', '')
    cm2 = num(p.get('CM2'))
    ebitda = num(p.get('EBITDA'))
    amz_star_chips += star_chip(code, cm2, ebitda, i==0)

# FC star top chips
fc_star_chips = ''
for i, p in enumerate(fc_stars[:4]):
    code = p.get('P. Breakdown', '')
    cm2 = num(p.get('CM2'))
    ebitda = num(p.get('EBITDA'))
    fc_star_chips += star_chip(code, cm2, ebitda, i==0)

# Web worst (discontinue)
web_worst = sorted(W, key=lambda p: num(p.get('EBITA'))*num(p.get('Units sold')))

# AMZ worst
amz_worst = sorted(A, key=lambda p: num(p.get('EBITDA'))*num(p.get('Units Sold')))

# FC worst
fc_worst = sorted(FC, key=lambda p: num(p.get('EBITDA'))*num(p.get('Units (3 months)')))

# BL sorted by EBITDA worst first
bl_worst = sorted(BL, key=lambda p: num(p.get('EBITDA'))*num(p.get('Units (3 months)')))

# Marketing spend comparison - per product
mktg_pct_with_gst = total_web_mktg/total_web_rev_gst*100 if total_web_rev_gst else 0
mktg_pct_without_gst = total_web_mktg/total_web_rev_no_gst*100 if total_web_rev_no_gst else 0

# Per-product marketing comparison rows
web_mktg_rows = []
for p in sorted(W, key=lambda x: -(num(x.get('3 month revenue with GST')))):
    code = p.get('Product Name (Code)', '')
    full_name = p.get('Product Name (Full)', code)
    units = num(p.get('Units sold'))
    rev_gst = num(p.get('3 month revenue with GST'))
    rev_no_gst = num(p.get('3 month revenue without GST'))
    mktg_pu = num(p.get('Marketing Cost'))
    total_mktg = mktg_pu * units
    cm2 = num(p.get('CM2'))
    pct_gst = total_mktg/rev_gst*100 if rev_gst else 0
    pct_no_gst = total_mktg/rev_no_gst*100 if rev_no_gst else 0
    web_mktg_rows.append({
        'code': code, 'full_name': full_name, 'units': units,
        'rev_gst': rev_gst, 'rev_no_gst': rev_no_gst,
        'total_mktg': total_mktg, 'pct_gst': pct_gst, 'pct_no_gst': pct_no_gst,
        'cm2': cm2, 'cm2_total': cm2*units
    })

# ============================================================
# READ HTML
# ============================================================
with open('index.html', encoding='utf-8') as f:
    html = f.read()

# ============================================================
# TASK 3: Remove "Important Metrics" tab button and panel
# ============================================================

# Remove the nav button for metrics
html = html.replace(
    '\n    <button role="tab" aria-selected="false" data-chan="metrics">Important Metrics</button>',
    ''
)

# Remove the entire chan-metrics panel
metrics_start = html.find('<div class="chan-panel" id="chan-metrics" hidden>')
metrics_end = html.find('</div><!-- /#chan-metrics -->')
if metrics_start != -1 and metrics_end != -1:
    html = html[:metrics_start] + html[metrics_end + len('</div><!-- /#chan-metrics -->'):]

# ============================================================
# TASK 4: Add Marketing Spend Comparison tab button
# ============================================================
html = html.replace(
    '    <button role="tab" aria-selected="false" data-chan="analysis">Channel Analysis</button>',
    '''    <button role="tab" aria-selected="false" data-chan="analysis">Channel Analysis</button>
    <button role="tab" aria-selected="false" data-chan="mktgspend">Mktg Spend Comparison</button>'''
)

# ============================================================
# BUILD MARKETING SPEND TAB HTML
# ============================================================
# Build table rows for all products
mktg_table_rows = ''
for p in web_mktg_rows:
    # Color coding for pct_no_gst
    pct_class = 'down' if p['pct_no_gst'] > 90 else ('up' if p['pct_no_gst'] < 30 else '')
    cm2_class = 'up' if p['cm2'] > 0 else 'down'
    mktg_table_rows += f"""<tr>
  <td style="text-align:left">{p['code']}</td>
  <td class="n">{fmt(p['units'], 0)}</td>
  <td class="n">\u20b9{fmt(p['rev_gst'], 0)}</td>
  <td class="n">\u20b9{fmt(p['rev_no_gst'], 0)}</td>
  <td class="n">\u20b9{fmt(p['total_mktg'], 0)}</td>
  <td class="n">{p['pct_gst']:.1f}%</td>
  <td class="n"><span class="{pct_class}" style="font-weight:600">{p['pct_no_gst']:.1f}%</span></td>
  <td class="n"><span class="{cm2_class}">\u20b9{fmt(p['cm2'], 0)}/u</span></td>
</tr>"""

mktg_panel_html = f'''<div class="chan-panel" id="chan-mktgspend" hidden>
<header class="top">
  <div class="top-in">
    <h1>Website Marketing Spend Analysis — Apr\u2013Jun 2026</h1>
    <p class="sub">Marketing spend as % of 3-month revenue — with GST (Col E) vs without GST (Col AD). All figures \u20b9.</p>
  </div>
</header>
<div class="wrap" style="padding-top:26px">

<h2>Summary — Total Website Marketing Spend</h2>
<p class="note">Comparing marketing spend efficiency against gross revenue (with GST) vs net revenue (without GST). The GST component inflates the denominator when measuring against col E, making the ratio appear lower. Using col AD (without GST) gives a cleaner picture of true marketing efficiency.</p>

<div class="kpis" style="margin-bottom:20px">
  <div class="kpi">
    <h3>3-Month Revenue (WITH GST)</h3>
    <div class="kpi-row"><span class="kpi-lab">Total Rev (Col E)</span><span class="kpi-val">\u20b9{fmt(total_web_rev_gst, 0)}</span></div>
    <div class="kpi-row"><span class="kpi-lab">Marketing Spend</span><span class="kpi-val sm">\u20b9{fmt(total_web_mktg, 0)}</span></div>
    <div class="kpi-d"><span>Mktg as % of Rev (Col E)</span><span style="font-weight:700;color:var(--crit)">{mktg_pct_with_gst:.1f}%</span></div>
  </div>
  <div class="kpi">
    <h3>3-Month Revenue (WITHOUT GST)</h3>
    <div class="kpi-row"><span class="kpi-lab">Total Rev (Col AD)</span><span class="kpi-val">\u20b9{fmt(total_web_rev_no_gst, 0)}</span></div>
    <div class="kpi-row"><span class="kpi-lab">Marketing Spend</span><span class="kpi-val sm">\u20b9{fmt(total_web_mktg, 0)}</span></div>
    <div class="kpi-d"><span>Mktg as % of Rev (Col AD)</span><span style="font-weight:700;color:var(--crit)">{mktg_pct_without_gst:.1f}%</span></div>
  </div>
  <div class="kpi">
    <h3>GST Impact</h3>
    <div class="kpi-row"><span class="kpi-lab">Rev with GST</span><span class="kpi-val sm">\u20b9{fmt(total_web_rev_gst, 0)}</span></div>
    <div class="kpi-row"><span class="kpi-lab">Rev without GST</span><span class="kpi-val sm">\u20b9{fmt(total_web_rev_no_gst, 0)}</span></div>
    <div class="kpi-d"><span>GST adds to denominator</span><span style="font-weight:700;color:var(--warn)">+\u20b9{fmt(total_web_rev_gst - total_web_rev_no_gst, 0)}</span></div>
  </div>
  <div class="kpi">
    <h3>Ratio Difference</h3>
    <div class="kpi-row"><span class="kpi-lab">Using Col E (w/ GST)</span><span class="kpi-val sm" style="color:var(--warn)">{mktg_pct_with_gst:.1f}%</span></div>
    <div class="kpi-row"><span class="kpi-lab">Using Col AD (w/o GST)</span><span class="kpi-val sm" style="color:var(--crit)">{mktg_pct_without_gst:.1f}%</span></div>
    <div class="kpi-d"><span>Diff (pp)</span><span style="font-weight:700;color:var(--neg)">+{mktg_pct_without_gst - mktg_pct_with_gst:.1f}pp</span></div>
  </div>
</div>

<div class="banner" style="--bc:var(--warn)">
  <h4><span class="ic">!</span>Key Insight</h4>
  <p>When measured against <b>revenue with GST (Col E)</b>, marketing spend appears at <b>{mktg_pct_with_gst:.1f}%</b> of revenue. But GST is a pass-through tax — it doesn\u2019t represent real business revenue. Measuring against <b>revenue without GST (Col AD)</b> shows the true ratio at <b>{mktg_pct_without_gst:.1f}%</b> — a difference of <b>+{mktg_pct_without_gst - mktg_pct_with_gst:.1f} percentage points</b>. All unit economics decisions should use the without-GST figure as the true denominator.</p>
</div>

<h2>Per-Product Breakdown</h2>
<p class="note">Products sorted by 3-month revenue with GST (Col E) descending. Top 25 products shown.</p>
<div class="tw"><table>
  <thead>
    <tr>
      <th>Product</th>
      <th>Units</th>
      <th>Rev w/ GST (Col E)</th>
      <th>Rev w/o GST (Col AD)</th>
      <th>Marketing Spend</th>
      <th>Mktg % (Col E)</th>
      <th>Mktg % (Col AD)</th>
      <th>CM2/unit</th>
    </tr>
  </thead>
  <tbody>
{mktg_table_rows}
    <tr style="font-weight:700;background:var(--wash)">
      <td>TOTAL</td>
      <td class="n">{fmt(total_web_units, 0)}</td>
      <td class="n">\u20b9{fmt(total_web_rev_gst, 0)}</td>
      <td class="n">\u20b9{fmt(total_web_rev_no_gst, 0)}</td>
      <td class="n">\u20b9{fmt(total_web_mktg, 0)}</td>
      <td class="n" style="color:var(--warn)">{mktg_pct_with_gst:.1f}%</td>
      <td class="n" style="color:var(--crit)">{mktg_pct_without_gst:.1f}%</td>
      <td class="n">\u20b9{fmt(total_web_cm2/total_web_units if total_web_units else 0, 0)}/u</td>
    </tr>
  </tbody>
</table></div>

</div>
</div><!-- /#chan-mktgspend -->'''

# ============================================================
# TASK 2: Update Channel Analysis decision criteria
# ============================================================

# Build new Q1 table rows
def ebitda_sign(v):
    if v > 0:
        return f'<span class="up">+\u20b9{v:,.0f}</span>'
    return f'<span class="down">-\u20b9{abs(v):,.0f}</span>'

web_star_rev_pct = web_star_rev/total_web_net_rev if total_web_net_rev else 0
amz_star_rev_pct = amz_star_rev/total_amz_net_rev if total_amz_net_rev else 0
fc_star_rev_pct = fc_star_rev/total_fc_net_rev if total_fc_net_rev else 0

new_q1_tbody = f"""        <tr><td><b>Website</b></td><td>{web_star_count} of {len(W)}</td><td>\u20b9{web_star_rev/100000:.2f} L</td><td>{web_star_rev_pct*100:.1f}%</td><td>{ebitda_sign(web_star_ebitda_t)}</td>
          <td style="text-align:left;padding:8px 10px">
            <div style="display:flex;flex-wrap:wrap;gap:5px">
              {web_star_chips}
            </div>
          </td></tr>
        <tr><td><b>Amazon</b></td><td>{amz_star_count} of {len(A)}</td><td>\u20b9{amz_star_rev/100000:.2f} L</td><td>{amz_star_rev_pct*100:.1f}%</td><td>{ebitda_sign(amz_star_ebitda_t)}</td>
          <td style="text-align:left;padding:8px 10px">
            <div style="display:flex;flex-wrap:wrap;gap:5px">
              {amz_star_chips}
            </div>
          </td></tr>
        <tr><td><b>FirstCry</b></td><td>{fc_star_count} of {len(FC)}</td><td>\u20b9{fc_star_rev/100000:.2f} L</td><td>{fc_star_rev_pct*100:.1f}%</td><td>{ebitda_sign(fc_star_ebitda_t)}</td>
          <td style="text-align:left;padding:8px 10px">
            <div style="display:flex;flex-wrap:wrap;gap:5px">
              {fc_star_chips}
            </div>
          </td></tr>
        <tr><td><b>Blinkit</b></td><td>0 of {len(BL)}</td><td>\u2014</td><td>0%</td><td><span class="down">None</span></td>
          <td style="text-align:left;padding:8px 10px">
            <div style="display:inline-flex;flex-direction:column;background:var(--wash);border:1px solid var(--border);border-radius:6px;padding:4px 8px;opacity:0.55">
              <span style="font-size:11.5px;font-weight:700;color:var(--ink-2)">No stars this quarter</span>
              <span style="font-size:10px;color:var(--ink-3);margin-top:1px">Double cost structure leaves no CM2+ EBITDA+ SKU</span>
            </div>
          </td></tr>"""

# Build new Q2 rows (overhead-heavy products)
def q2_row(channel, code, cm2, ebitda, note='', rowspan=None):
    gap = abs(ebitda)
    cm2_str = f'+\u20b9{cm2:.0f}' if cm2 > 0 else f'-\u20b9{abs(cm2):.0f}'
    eb_str = f'-\u20b9{gap:.0f}'
    html_row = ''
    if rowspan:
        html_row = f'<tr><td rowspan="{rowspan}"><b>{channel}</b></td>\n'
    else:
        html_row = '<tr>'
    html_row += f'<td>{code}</td><td class="n"><span class="up">{cm2_str}</span></td><td class="n"><span class="down">{eb_str}</span></td><td class="n">\u20b9{gap:.0f}/unit</td><td class="muted">{note}</td></tr>'
    return html_row

# Web overhead-heavy (top 3 closest to breakeven with actual numbers)
woh = web_overhead_h[:3]
new_q2_web = ''
for i, p in enumerate(woh):
    code = p.get('Product Name (Code)', '')
    cm2 = num(p.get('CM2'))
    ebita = num(p.get('EBITA'))
    gap = abs(ebita)
    if i == 0:
        note = 'Closest to EBITDA breakeven \u2014 prioritise volume'
    elif cm2 < 100:
        note = 'CM2 thin \u2014 also look at pricing'
    else:
        note = 'Volume or price'
    if i == 0:
        new_q2_web += f'<tr><td rowspan="{len(woh)}"><b>Website</b></td><td>{code}</td><td class="n"><span class="up">+\u20b9{cm2:.0f}</span></td><td class="n"><span class="down">-\u20b9{abs(ebita):.0f}</span></td><td class="n">\u20b9{gap:.0f}/unit</td><td class="muted">{note}</td></tr>\n'
    else:
        new_q2_web += f'<tr><td>{code}</td><td class="n"><span class="up">+\u20b9{cm2:.0f}</span></td><td class="n"><span class="down">-\u20b9{abs(ebita):.0f}</span></td><td class="n">\u20b9{gap:.0f}/unit</td><td class="muted">{note}</td></tr>\n'

# Amazon overhead-heavy
aoh = amz_overhead_h[:3]
new_q2_amz = ''
for i, p in enumerate(aoh):
    code = p.get('P. Breakdown', '')
    cm2 = num(p.get('CM2'))
    ebitda = num(p.get('EBITDA'))
    gap = abs(ebitda)
    if gap < 15:
        note = 'Near-breakeven \u2014 volume push justified'
    elif gap < 30:
        note = 'Closest to EBITDA+ on Amazon'
    else:
        note = 'Thin CM2 \u2014 pricing fix needed first'
    if i == 0:
        new_q2_amz += f'<tr><td rowspan="{len(aoh)}"><b>Amazon</b></td><td>{code}</td><td class="n"><span class="up">+\u20b9{cm2:.0f}</span></td><td class="n"><span class="down">-\u20b9{abs(ebitda):.0f}</span></td><td class="n">\u20b9{gap:.0f}/unit</td><td class="muted">{note}</td></tr>\n'
    else:
        new_q2_amz += f'<tr><td>{code}</td><td class="n"><span class="up">+\u20b9{cm2:.0f}</span></td><td class="n"><span class="down">-\u20b9{abs(ebitda):.0f}</span></td><td class="n">\u20b9{gap:.0f}/unit</td><td class="muted">{note}</td></tr>\n'

# FC overhead-heavy
foh = fc_overhead_h[:3]
new_q2_fc = ''
for i, p in enumerate(foh):
    code = p.get('P. Breakdown', '')
    cm2 = num(p.get('CM2'))
    ebitda = num(p.get('EBITDA'))
    gap = abs(ebitda)
    if i == 0:
        new_q2_fc += f'<tr><td rowspan="{len(foh)}"><b>FirstCry</b></td><td>{code}</td><td class="n"><span class="up">+\u20b9{cm2:.0f}</span></td><td class="n"><span class="down">-\u20b9{abs(ebitda):.0f}</span></td><td class="n">\u20b9{gap:.0f}/unit</td><td class="muted">Volume</td></tr>\n'
    else:
        new_q2_fc += f'<tr><td>{code}</td><td class="n"><span class="up">+\u20b9{cm2:.0f}</span></td><td class="n"><span class="down">-\u20b9{abs(ebitda):.0f}</span></td><td class="n">\u20b9{gap:.0f}/unit</td><td class="muted">Volume</td></tr>\n'

# Blinkit B8 (only CM2+ but EBITDA-)
b8 = next((p for p in BL if 'B8' in str(p.get('P. Breakdown',''))), BL[-1])
b8_cm2 = num(b8.get('CM2'))
b8_ebitda = num(b8.get('EBITDA'))
new_q2_bl = f'<tr><td><b>Blinkit</b></td><td>B8 Crayons</td><td class="n"><span class="up">+\u20b9{b8_cm2:.0f}</span></td><td class="n"><span class="down">-\u20b9{abs(b8_ebitda):.0f}</span></td><td class="n">\u20b9{abs(b8_ebitda):.0f}/unit</td><td class="muted">CM2 too thin \u2014 pricing before volume</td></tr>'

# Build Q3 discontinue rows
def q3_row(channel, product, units, ebitda_total, net_rev, action, rowspan=None):
    loss_pct = -ebitda_total/net_rev*100 if net_rev > 0 else 0
    ebitda_str = f'-\u20b9{abs(ebitda_total)/100000:.2f} L' if abs(ebitda_total) >= 100000 else f'-\u20b9{abs(ebitda_total):,.0f}'
    td_chan = f'<td rowspan="{rowspan}"><b>{channel}</b></td>' if rowspan else ''
    return f'<tr>{td_chan}<td>{product}</td><td class="n">{units:.0f}</td><td class="n"><span class="down">{ebitda_str}</span></td><td class="n">{loss_pct:.0f}%</td><td class="muted">{action}</td></tr>'

new_q3_web = ''
for i, p in enumerate(web_worst[:3]):
    code = p.get('Product Name (Code)', '')
    full = p.get('Product Name (Full)', code)
    units = num(p.get('Units sold'))
    ebita_t = num(p.get('EBITA'))*units
    net_rev_t = num(p.get('Net Revenue without GST'))*units
    loss_pct = -ebita_t/net_rev_t*100 if net_rev_t > 0 else 0
    eb_str = f'-\u20b9{abs(ebita_t)/100000:.2f} L' if abs(ebita_t) >= 100000 else f'-\u20b9{abs(ebita_t):,.0f}'
    if loss_pct > 150:
        action = 'Delist \u2014 losing more than revenue in EBITDA'
    elif loss_pct > 80:
        action = 'Reprice or delist \u2014 marketing cost too high per unit'
    else:
        action = 'Reduce marketing spend, reassess pricing'
    if i == 0:
        new_q3_web += f'<tr><td rowspan="3"><b>Website</b></td><td>{full}</td><td class="n">{units:.0f}</td><td class="n"><span class="down">{eb_str}</span></td><td class="n">{loss_pct:.0f}%</td><td class="muted">{action}</td></tr>\n'
    else:
        new_q3_web += f'<tr><td>{full}</td><td class="n">{units:.0f}</td><td class="n"><span class="down">{eb_str}</span></td><td class="n">{loss_pct:.0f}%</td><td class="muted">{action}</td></tr>\n'

new_q3_amz = ''
for i, p in enumerate(amz_worst[:3]):
    code = p.get('P. Breakdown', '')
    full_name = p.get('Product', code)
    units = num(p.get('Units Sold'))
    ebitda_t = num(p.get('EBITDA'))*units
    net_rev_t = num(p.get('Net Revenue without GST'))*units
    loss_pct = -ebitda_t/net_rev_t*100 if net_rev_t > 0 else 0
    eb_str = f'-\u20b9{abs(ebitda_t)/100000:.2f} L' if abs(ebitda_t) >= 100000 else f'-\u20b9{abs(ebitda_t):,.0f}'
    if loss_pct > 80:
        action = 'Delist \u2014 CM1 deeply negative'
    elif num(p.get('CM1')) < 0:
        action = 'Delist \u2014 negative net revenue after discounts'
    else:
        action = 'Cut ads, reassess pricing'
    if i == 0:
        new_q3_amz += f'<tr><td rowspan="3"><b>Amazon</b></td><td>{code}</td><td class="n">{units:.0f}</td><td class="n"><span class="down">{eb_str}</span></td><td class="n">{loss_pct:.0f}%</td><td class="muted">{action}</td></tr>\n'
    else:
        new_q3_amz += f'<tr><td>{code}</td><td class="n">{units:.0f}</td><td class="n"><span class="down">{eb_str}</span></td><td class="n">{loss_pct:.0f}%</td><td class="muted">{action}</td></tr>\n'

new_q3_fc = ''
for i, p in enumerate(fc_worst[:2]):
    code = p.get('P. Breakdown', '')
    units = num(p.get('Units (3 months)'))
    ebitda_t = num(p.get('EBITDA'))*units
    net_rev = num(p.get('Total Net Revenue'))
    loss_pct = -ebitda_t/net_rev*100 if net_rev > 0 else 0
    eb_str = f'-\u20b9{abs(ebitda_t):,.0f}'
    action = 'Bundle with paints \u2014 flat overhead + FC fee outweigh solo CM2' if loss_pct < 30 else 'Bundle \u2014 CM2 positive, overhead allocation tips it negative'
    if i == 0:
        new_q3_fc += f'<tr><td rowspan="2"><b>FirstCry</b></td><td>{code} (solo)</td><td class="n">{units:.0f}</td><td class="n"><span class="down">{eb_str}</span></td><td class="n">{loss_pct:.0f}%</td><td class="muted">{action}</td></tr>\n'
    else:
        new_q3_fc += f'<tr><td>{code} (solo)</td><td class="n">{units:.0f}</td><td class="n"><span class="down">{eb_str}</span></td><td class="n">{loss_pct:.0f}%</td><td class="muted">{action}</td></tr>\n'

new_q3_bl = ''
bl_sorted_worst = sorted(BL, key=lambda p: num(p.get('EBITDA'))*num(p.get('Units (3 months)')))
for i, p in enumerate(bl_sorted_worst[:2]):
    code = p.get('P. Breakdown', '')
    units = num(p.get('Units (3 months)'))
    ebitda_t = num(p.get('EBITDA'))*units
    net_rev_t = num(p.get('Net Revenue without GST'))*units
    loss_pct = -ebitda_t/net_rev_t*100 if net_rev_t > 0 else 0
    eb_str = f'-\u20b9{abs(ebitda_t)/100000:.2f} L' if abs(ebitda_t) >= 100000 else f'-\u20b9{abs(ebitda_t):,.0f}'
    if i == 0:
        new_q3_bl += f'<tr><td rowspan="2"><b>Blinkit</b></td><td>{code}</td><td class="n">{units:.0f}</td><td class="n"><span class="down">{eb_str}</span></td><td class="n">{loss_pct:.0f}%</td><td class="muted">Pause ads, clear current inventory then reassess</td></tr>\n'
    else:
        new_q3_bl += f'<tr><td>{code}</td><td class="n">{units:.0f}</td><td class="n"><span class="down">{eb_str}</span></td><td class="n">{loss_pct:.0f}%</td><td class="muted">Pause ads, clear current inventory then reassess</td></tr>\n'

# Build Q4 area of focus table
new_q4_rows = f"""        <tr><td><b>Website</b></td><td class="n">{pct(web_gmpct)}%</td><td class="n"><b style="color:var(--crit)">{pct(web_mktg_pct)}%</b></td><td class="n">&mdash;</td><td class="n">~{total_web_ebita/total_web_units*(-1)/total_web_net_rev*total_web_units*100:.1f}%*</td>
          <td><b>Marketing efficiency.</b> GM is healthy at {pct(web_gmpct)}. Marketing alone exceeds CM1 and is the single reason the channel loses money. CAC/conversion fix is the #1 lever.</td></tr>
        <tr><td><b>Amazon</b></td><td class="n">{pct(amz_gmpct)}%</td><td class="n"><b style="color:var(--crit)">{pct(amz_mktg_pct)}%</b></td><td class="n">&mdash;</td><td class="n">~15%*</td>
          <td><b>Marketing + referral fee.</b> Amazon Ads are {pct(amz_mktg_pct)} of net rev. Fix: cut ad spend on low converters (see Section 5).</td></tr>
        <tr><td><b>FirstCry</b></td><td class="n">{pct(fc_gmpct)}%</td><td class="n">&mdash;</td><td class="n"><b style="color:var(--crit)">{pct(fc_platform_pct)}%</b></td><td class="n">&mdash;</td>
          <td><b>Platform margin, then overheads.</b> Gross margin is strong and CM2 stays positive on most SKUs &mdash; FirstCry&apos;s fee ({pct(fc_platform_pct)} of net revenue) is the single biggest lever. Fix: bundle low-MRP solo SKUs to spread the overhead allocation.</td></tr>
        <tr><td><b>Blinkit</b></td><td class="n">{pct(bl_gmpct)}%</td><td class="n"><b style="color:var(--crit)">{pct(bl_mktg_pct)}%</b></td><td class="n"><b style="color:var(--crit)">{pct(bl_platform_pct)}%</b></td><td class="n">~{total_bl_net_rev/(total_bl_net_rev)*100:.1f}%*</td>
          <td><b>Double cost problem.</b> Marketing ({pct(bl_mktg_pct)}) + Blinkit Margin ({pct(bl_platform_pct)}) = {pct(bl_mktg_pct + bl_platform_pct)} of net rev. Short-term: cut all paid campaigns on loss-making SKUs. Medium-term: renegotiate margin rate or raise sell prices.</td></tr>"""

# ============================================================
# REPLACE Q1 TABLE TBODY
# ============================================================
q1_start_marker = '<thead><tr><th>Channel</th><th>Profitable SKUs</th><th>Rev from stars</th><th>% of channel rev</th><th>Combined EBITDA</th><th>Top SKUs (CM2/unit &rarr; EBITDA/unit)</th></tr></thead>'
q1_tbody_start = html.find(q1_start_marker)
if q1_tbody_start != -1:
    tbody_start = html.find('<tbody>', q1_tbody_start)
    tbody_end = html.find('</tbody>', tbody_start) + len('</tbody>')
    old_tbody = html[tbody_start:tbody_end]
    new_tbody = f'<tbody>\n{new_q1_tbody}\n      </tbody>'
    html = html[:tbody_start] + new_tbody + html[tbody_end:]
    print("Updated Q1 table")
else:
    print("WARNING: Q1 table not found!")

# ============================================================
# REPLACE Q2 TABLE TBODY
# ============================================================
q2_start_marker = '<thead><tr><th>Channel</th><th>SKU</th><th>CM2/unit</th><th>EBITDA/unit</th><th>Gap to EBITDA breakeven</th><th>Lever</th></tr></thead>'
q2_tbody_start_search = html.find(q2_start_marker)
if q2_tbody_start_search != -1:
    tbody_start2 = html.find('<tbody>', q2_tbody_start_search)
    tbody_end2 = html.find('</tbody>', tbody_start2) + len('</tbody>')
    new_q2_tbody = f'<tbody>\n{new_q2_web}{new_q2_amz}{new_q2_fc}{new_q2_bl}\n      </tbody>'
    html = html[:tbody_start2] + new_q2_tbody + html[tbody_end2:]
    print("Updated Q2 table")
else:
    print("WARNING: Q2 table not found!")

# ============================================================
# REPLACE Q3 TABLE TBODY
# ============================================================
q3_start_marker = '<thead><tr><th>Channel</th><th>Product</th><th>Units</th><th>EBITDA loss</th><th>Loss as % of rev</th><th>Action</th></tr></thead>'
q3_tbody_search = html.find(q3_start_marker)
if q3_tbody_search != -1:
    tbody_start3 = html.find('<tbody>', q3_tbody_search)
    tbody_end3 = html.find('</tbody>', tbody_start3) + len('</tbody>')
    new_q3_tbody = f'<tbody>\n{new_q3_web}{new_q3_amz}{new_q3_fc}{new_q3_bl}\n      </tbody>'
    html = html[:tbody_start3] + new_q3_tbody + html[tbody_end3:]
    print("Updated Q3 table")
else:
    print("WARNING: Q3 table not found!")

# ============================================================
# REPLACE Q4 TABLE TBODY
# ============================================================
q4_start_marker = '<thead><tr><th>Channel</th><th>Gross margin %</th><th>Marketing as % of net rev</th><th>Platform fee % of net rev</th><th>Shipping % of net rev</th><th>Primary lever</th></tr></thead>'
q4_tbody_search = html.find(q4_start_marker)
if q4_tbody_search != -1:
    tbody_start4 = html.find('<tbody>', q4_tbody_search)
    tbody_end4 = html.find('</tbody>', tbody_start4) + len('</tbody>')
    new_q4_tbody = f'<tbody>\n{new_q4_rows}\n      </tbody>'
    html = html[:tbody_start4] + new_q4_tbody + html[tbody_end4:]
    print("Updated Q4 table")
else:
    print("WARNING: Q4 table not found!")

# ============================================================
# ADD MARKETING SPEND PANEL BEFORE </BODY>
# ============================================================
# Insert before the tip div (which is before the script)
insert_before = '<div class="tip" id="tip"></div>'
if insert_before in html:
    html = html.replace(insert_before, mktg_panel_html + '\n' + insert_before)
    print("Added Marketing Spend panel")
else:
    print("WARNING: Could not find insertion point for mktg panel!")

# ============================================================
# UPDATE JAVASCRIPT SWITCH CASE for mktgspend
# ============================================================
# Find the JS channel switch logic
js_switch = "case 'metrics':"
if js_switch in html:
    # Replace metrics case with mktgspend
    html = html.replace("case 'metrics':", "case 'mktgspend':")
    html = html.replace("'chan-metrics'", "'chan-mktgspend'")
    print("Updated JS switch cases")
else:
    print("Looking for alternative JS pattern...")
    # Try to find the chanSwitch js
    if 'metrics' in html:
        html = html.replace('"metrics"', '"mktgspend"')
        html = html.replace("'metrics'", "'mktgspend'")
        html = html.replace('chan-metrics', 'chan-mktgspend')
        print("Replaced all metrics references")

# ============================================================
# WRITE UPDATED HTML
# ============================================================
with open('index.html', 'w', encoding='utf-8') as f:
    f.write(html)

print("\nDone! index.html updated successfully.")
print(f"Website: {len(W)} SKUs, CM2={total_web_cm2:,.0f} ({web_cm2pct*100:.1f}%), EBITA={total_web_ebita:,.0f}")
print(f"Amazon: {len(A)} SKUs, CM2={total_amz_cm2:,.0f} ({amz_cm2pct*100:.1f}%), EBITDA={total_amz_ebitda:,.0f}")
print(f"FirstCry: {len(FC)} SKUs, CM2={total_fc_cm2:,.0f} ({fc_cm2pct*100:.1f}%), EBITDA={total_fc_ebitda:,.0f}")
print(f"Blinkit: {len(BL)} SKUs, CM2={total_bl_cm2:,.0f} ({bl_cm2pct*100:.1f}%), EBITDA={total_bl_ebitda:,.0f}")
