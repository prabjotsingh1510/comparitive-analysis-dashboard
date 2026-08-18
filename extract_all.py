import openpyxl
import json
import sys

wb = openpyxl.load_workbook(
    r'Channel Economics Dashboard.xlsx',
    data_only=True
)

def cell_val(ws, row, col):
    v = ws.cell(row=row, column=col).value
    return v

def dump_sheet(name, max_row=200, max_col=30):
    ws = wb[name]
    rows = []
    for r in range(1, max_row+1):
        row_data = []
        has_val = False
        for c in range(1, max_col+1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                has_val = True
            row_data.append(v)
        if has_val:
            rows.append({'row': r, 'data': row_data})
    return rows

sheets_to_dump = [
    'Channel Summary',
    'Raw Data - Website',
    'Raw Data - Amazon', 
    'Raw Data - FirstCry',
    'Raw Data - Blinkit',
    'Channel Matrices',
    'Marketing Efficiency',
    'CAC & DRR Analysis',
    'Executive Dashboard',
    'Recommendations',
]

result = {}
for s in sheets_to_dump:
    if s in wb.sheetnames:
        result[s] = dump_sheet(s)
        print(f"Extracted: {s} ({len(result[s])} rows)", flush=True)
    else:
        print(f"MISSING: {s}", flush=True)

with open('extracted_data.json', 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, default=str, indent=2)

print("Done! Saved to extracted_data.json")
