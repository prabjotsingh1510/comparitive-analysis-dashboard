"""
Clean fix: rebuild AMZ revenue map by key-only (no ASIN), handle FBA+FBM by summing,
deduplicate products, patch D_AMZ / D_FC / D_BL correctly.
"""
import openpyxl, warnings, json, re
warnings.filterwarnings('ignore')

def flt(v, d=0.0):
    try: return float(v) if v is not None else d
    except: return d
def sint(v):
    try: return int(float(v)) if v is not None else 0
    except: return 0

# ════ AMAZON: build rev map keyed by P.Breakdown only ═══════════════════════
wb_a = openpyxl.load_workbook('Unit cost economics AMAZON April-June (2).xlsx', data_only=True)
ws_a = wb_a['Final Sheet (Combined)']

amz_rev = {}   # pb_key -> total_nr (sum across FBA+FBM rows)
amz_units = {}
for r in range(2, ws_a.max_row + 1):
    pb    = ws_a.cell(r, 2).value
    units = sint(ws_a.cell(r, 33).value)
    tot_nr= flt(ws_a.cell(r, 35).value)
    if not pb or units == 0: continue
    pb = str(pb).strip()
    amz_rev[pb]   = amz_rev.get(pb, 0)   + tot_nr
    amz_units[pb] = amz_units.get(pb, 0) + units

print('Amazon revenue map:')
for k, v in sorted(amz_rev.items(), key=lambda x: -x[1])[:10]:
    print(f'  {k!r:30s}  rev={v:>12,.0f}  units={amz_units[k]:>5}')
print(f'  ... {len(amz_rev)} total keys')

# ════ FIRSTCRY: SELL PRICE × Units ══════════════════════════════════════════
wb_fc = openpyxl.load_workbook('Firstcry_Unit Cost Economic(Apr-June) (1).xlsx', data_only=True)
ws_fc = wb_fc['SP Vendor Format']
fc_rev = {}
for r in range(2, ws_fc.max_row + 1):
    pb    = ws_fc.cell(r, 4).value
    units = sint(ws_fc.cell(r, 10).value)
    sp    = flt(ws_fc.cell(r, 11).value)
    if not pb or units == 0: continue
    pb = str(pb).strip()
    fc_rev[pb] = round(fc_rev.get(pb, 0) + sp * units, 2)

# ════ BLINKIT: SP × Units ════════════════════════════════════════════════════
wb_bl = openpyxl.load_workbook('Blinkit_Unit Cost Economic(Apr-June) (1).xlsx', data_only=True)
ws_bl = wb_bl['Final(Unit Cost Format)']
bl_rev = {}
for r in range(2, ws_bl.max_row + 1):
    pb    = ws_bl.cell(r, 3).value
    units = sint(ws_bl.cell(r, 5).value)
    sp    = flt(ws_bl.cell(r, 7).value)
    if not pb or units == 0: continue
    pb = str(pb).strip()
    bl_rev[pb] = round(bl_rev.get(pb, 0) + sp * units, 2)

# ════ Load HTML ═══════════════════════════════════════════════════════════════
with open('index.html', encoding='utf-8') as f:
    html = f.read()

def load_var(html, var_name):
    m = re.search('const ' + var_name + r' = (\{.*?\});', html, re.DOTALL)
    return json.loads(m.group(1))

def save_var(html, var_name, obj):
    new_js = json.dumps(obj, ensure_ascii=False, separators=(',', ':'))
    start_tok = 'const ' + var_name + ' = {'
    idx_start = html.find(start_tok)
    depth = 0
    i = idx_start + len('const ' + var_name + ' = ')
    while i < len(html):
        c = html[i]
        if c == '{': depth += 1
        elif c == '}':
            depth -= 1
            if depth == 0:
                idx_end = i + 1
                break
        i += 1
    s = html.find(';', idx_end)
    if s != -1 and s < idx_end + 3:
        idx_end = s + 1
    new = 'const ' + var_name + ' = ' + new_js + ';'
    old = html[idx_start:idx_end]
    print(f'{var_name}: {len(old):,} -> {len(new):,} chars')
    return html[:idx_start] + new + html[idx_end:]

# ════ Patch D_AMZ ═════════════════════════════════════════════════════════════
amz = load_var(html, 'D_AMZ')

# Deduplicate products by key (keep only one record per key, merging totals)
deduped = {}
for p in amz['products']:
    k = p['key']
    if k not in deduped:
        deduped[k] = p.copy()
    else:
        # merge: sum totals, average per-unit weighted by units
        old_u = deduped[k].get('units', 0)
        new_u = p.get('units', 0)
        total_u = old_u + new_u
        if total_u > 0:
            for field in ['netrev_u','gm_u','ship_u','cm1_u','bss_u','evit_u',
                          'mktg_u','cm2_u','oh_u','ebitda_u']:
                if field in deduped[k] and field in p:
                    deduped[k][field] = (deduped[k][field]*old_u + p[field]*new_u) / total_u
        for field in ['netrev_t','gm_t','cm1_t','mktg_t','cm2_t','oh_t','ebitda_t']:
            if field in deduped[k] and field in p:
                deduped[k][field] = deduped[k][field] + p[field]
        deduped[k]['units'] = total_u

amz['products'] = list(deduped.values())

# Now set rev_t from Excel revenue map
unmatched = []
for p in amz['products']:
    k = p['key']
    if k in amz_rev:
        p['rev_t'] = round(amz_rev[k], 2)
        # also update units to match Excel (more reliable)
        p['units'] = amz_units[k]
    else:
        # fallback: MRP * units
        p['rev_t'] = round(flt(p.get('mrp', 0)) * sint(p.get('units', 0)), 2)
        unmatched.append(k)

print(f'\nAMZ: {len(amz["products"])} unique products, unmatched={unmatched}')

# Recalculate totals from products
prods = amz['products']
amz['tot']['rev']    = round(sum(p['rev_t'] for p in prods), 2)
amz['tot']['netrev'] = round(sum(p.get('netrev_t', p['rev_t']) for p in prods), 2)
amz['tot']['units']  = sum(p.get('units', 0) for p in prods)
amz['tot']['cm2']    = round(sum(p.get('cm2_t', 0) for p in prods), 2)
amz['tot']['ebitda'] = round(sum(p.get('ebitda_t', 0) for p in prods), 2)
amz['tot']['mktg']   = round(sum(p.get('mktg_t', 0) for p in prods), 2)
amz['tot']['gm']     = round(sum(p.get('gm_t', 0) for p in prods), 2)
amz['tot']['oh']     = round(sum(p.get('oh_t', 0) for p in prods), 2)
nr = amz['tot']['netrev']
amz['tot']['cm2pct']    = round(amz['tot']['cm2'] / nr * 100, 5) if nr else 0
amz['tot']['ebitdapct'] = round(amz['tot']['ebitda'] / nr * 100, 5) if nr else 0
amz['tot']['gmpct']     = round(amz['tot']['gm'] / nr, 7) if nr else 0
amz['tot']['n']         = len(prods)

# Recalculate quads
for q in ['star', 'overhead', 'loss']:
    qp = [p for p in prods if p.get('quad') == q]
    amz['quad'][q] = {
        'n': len(qp),
        'rev': round(sum(p['rev_t'] for p in qp), 2),
        'units': sum(p.get('units', 0) for p in qp),
        'ebitda': round(sum(p.get('ebitda_t', 0) for p in qp), 2),
    }

# ════ Patch D_FC ═════════════════════════════════════════════════════════════
fc = load_var(html, 'D_FC')
unmatched_fc = []
for p in fc['products']:
    k = p.get('key', '')
    if k in fc_rev:
        p['rev_t'] = fc_rev[k]
    else:
        unmatched_fc.append(k)
fc['tot']['rev'] = round(sum(p.get('rev_t', 0) for p in fc['products']), 2)
print(f'FC: unmatched={unmatched_fc}')

# ════ Patch D_BL ═════════════════════════════════════════════════════════════
bl = load_var(html, 'D_BL')
unmatched_bl = []
for p in bl['products']:
    k = p.get('key', '')
    if k in bl_rev:
        p['rev_t'] = bl_rev[k]
    else:
        unmatched_bl.append(k)
bl['tot']['rev'] = round(sum(p.get('rev_t', 0) for p in bl['products']), 2)
print(f'BL: unmatched={unmatched_bl}')

# ════ Save ═══════════════════════════════════════════════════════════════════
html = save_var(html, 'D_AMZ', amz)
html = save_var(html, 'D_FC',  fc)
html = save_var(html, 'D_BL',  bl)

with open('index.html', 'w', encoding='utf-8') as f:
    f.write(html)
print('\nindex.html written. Size:', len(html))

# ════ Final check ═════════════════════════════════════════════════════════════
print('\n=== FINAL REVENUE CHECK ===')
for var in ['D_AMZ', 'D_FC', 'D_BL']:
    m = re.search('const ' + var + r' = (\{.*?\});', html, re.DOTALL)
    d = json.loads(m.group(1))
    prods = d['products']
    zero = [p['key'] for p in prods if p.get('rev_t', 0) <= 0]
    print(f'\n{var}: {len(prods)} products  tot.rev=₹{d["tot"]["rev"]:,.0f}')
    for p in sorted(prods, key=lambda x: -x.get('rev_t', 0))[:6]:
        print(f'  {p["key"]!r:28s}  rev_t=₹{p["rev_t"]:>12,.0f}  units={p.get("units",0):>5}')
    if zero:
        print(f'  ⚠ ZERO REV: {zero}')

# Also verify website D unchanged
m4 = re.search(r'const D = (\{.*?\});', html, re.DOTALL)
d4 = json.loads(m4.group(1))
print(f'\nWebsite D intact: apr.rev={d4["periods"]["apr"]["tot"]["rev"]:,.2f}')
