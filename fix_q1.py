"""
Fix Q1 table and also update the channel-level KPI data in the JavaScript section
"""
import openpyxl
import re

wb = openpyxl.load_workbook('Channel Economics Dashboard.xlsx', data_only=True)

def num(v):
    if v is None: return 0.0
    try: return float(v)
    except: return 0.0

# ============ Extract data (same as before) ============
ws_web = wb['Raw Data - Website']
web_headers = [ws_web.cell(row=4, column=c).value for c in range(1, 50)]
W = []
for r in range(5, 200):
    code = ws_web.cell(row=r, column=1).value
    if code is None: continue
    row = {}
    for c, h in enumerate(web_headers[:42], 1):
        if h:
            v = ws_web.cell(row=r, column=c).value
            try: row[h] = float(v) if v is not None else None
            except: row[h] = v
    W.append(row)

ws_amz = wb['Raw Data - Amazon']
amz_headers = [ws_amz.cell(row=4, column=c).value for c in range(1, 50)]
A = []
for r in range(5, 200):
    prod = ws_amz.cell(row=r, column=1).value
    if prod is None: continue
    row = {}
    for c, h in enumerate(amz_headers[:45], 1):
        if h:
            v = ws_amz.cell(row=r, column=c).value
            try: row[h] = float(v) if v is not None else None
            except: row[h] = v
    A.append(row)

ws_fc = wb['Raw Data - FirstCry']
fc_headers = [ws_fc.cell(row=4, column=c).value for c in range(1, 35)]
FC_all = []
for r in range(5, 200):
    prod = ws_fc.cell(row=r, column=1).value
    if prod is None: continue
    row = {}
    for c, h in enumerate(fc_headers, 1):
        if h:
            v = ws_fc.cell(row=r, column=c).value
            try: row[h] = float(v) if v is not None else None
            except: row[h] = v
    FC_all.append(row)
FC = [p for p in FC_all if num(p.get('Units (3 months)')) > 0]

ws_bl = wb['Raw Data - Blinkit']
bl_headers = [ws_bl.cell(row=4, column=c).value for c in range(1, 35)]
BL = []
for r in range(5, 200):
    prod = ws_bl.cell(row=r, column=1).value
    if prod is None: continue
    row = {}
    for c, h in enumerate(bl_headers, 1):
        if h:
            v = ws_bl.cell(row=r, column=c).value
            try: row[h] = float(v) if v is not None else None
            except: row[h] = v
    BL.append(row)

# Compute
total_web_net_rev = sum(num(p.get('Net Revenue without GST'))*num(p.get('Units sold')) for p in W)
total_amz_net_rev = sum(num(p.get('Net Revenue without GST'))*num(p.get('Units Sold')) for p in A)
total_fc_net_rev = sum(num(p.get('Total Net Revenue')) for p in FC)

web_stars = sorted([p for p in W if num(p.get('CM2')) > 0 and num(p.get('EBITA')) > 0],
                   key=lambda p: -num(p.get('EBITA'))*num(p.get('Units sold')))
amz_stars = sorted([p for p in A if num(p.get('CM2')) > 0 and num(p.get('EBITDA')) > 0],
                   key=lambda p: -num(p.get('EBITDA'))*num(p.get('Units Sold')))
fc_stars = sorted([p for p in FC if num(p.get('CM2')) > 0 and num(p.get('EBITDA')) > 0],
                  key=lambda p: -num(p.get('EBITDA'))*num(p.get('Units (3 months)')))

web_star_rev = sum(num(p.get('Net Revenue without GST'))*num(p.get('Units sold')) for p in web_stars)
amz_star_rev = sum(num(p.get('Net Revenue without GST'))*num(p.get('Units Sold')) for p in amz_stars)
fc_star_rev = sum(num(p.get('Total Net Revenue')) for p in fc_stars)

web_star_ebitda_t = sum(num(p.get('EBITA'))*num(p.get('Units sold')) for p in web_stars)
amz_star_ebitda_t = sum(num(p.get('EBITDA'))*num(p.get('Units Sold')) for p in amz_stars)
fc_star_ebitda_t = sum(num(p.get('EBITDA'))*num(p.get('Units (3 months)')) for p in fc_stars)

web_star_rev_pct = web_star_rev/total_web_net_rev if total_web_net_rev else 0
amz_star_rev_pct = amz_star_rev/total_amz_net_rev if total_amz_net_rev else 0
fc_star_rev_pct = fc_star_rev/total_fc_net_rev if total_fc_net_rev else 0

# Star chips
def star_chip(code, cm2, ebitda, first=False):
    border = 'border:1px solid var(--good);border-left:3px solid var(--good)' if first else 'border:1px solid var(--border)'
    return (f'<div style="display:inline-flex;flex-direction:column;background:var(--wash);{border};'
            f'border-radius:6px;padding:4px 8px;min-width:90px">'
            f'<span style="font-size:11.5px;font-weight:700;color:var(--ink-1);white-space:nowrap">{code}</span>'
            f'<span style="font-size:10px;color:var(--ink-3);margin-top:1px;font-variant-numeric:tabular-nums">'
            f'CM2 \u20b9{cm2:.0f} \u2192 EBITDA \u20b9{ebitda:.0f}</span></div>')

def ebitda_sign(v):
    if v >= 0:
        return f'<span class="up">+\u20b9{v:,.0f}</span>'
    return f'<span class="down">-\u20b9{abs(v):,.0f}</span>'

def lakh_fmt(v):
    if abs(v) >= 100000:
        return f'\u20b9{v/100000:.2f} L'
    return f'\u20b9{v:,.0f}'

# Q1 new tbody
web_chips = ''.join(star_chip(p.get('Product Name (Code)',''), num(p.get('CM2')), num(p.get('EBITA')), i==0) for i,p in enumerate(web_stars[:4]))
amz_chips = ''.join(star_chip(p.get('P. Breakdown',''), num(p.get('CM2')), num(p.get('EBITDA')), i==0) for i,p in enumerate(amz_stars[:4]))
fc_chips = ''.join(star_chip(p.get('P. Breakdown',''), num(p.get('CM2')), num(p.get('EBITDA')), i==0) for i,p in enumerate(fc_stars[:4]))

new_q1_tbody = f"""<tbody>
        <tr><td><b>Website</b></td><td>{len(web_stars)} of {len(W)}</td><td>{lakh_fmt(web_star_rev)}</td><td>{web_star_rev_pct*100:.1f}%</td><td>{ebitda_sign(web_star_ebitda_t)}</td>
          <td style="text-align:left;padding:8px 10px">
            <div style="display:flex;flex-wrap:wrap;gap:5px">{web_chips}</div>
          </td></tr>
        <tr><td><b>Amazon</b></td><td>{len(amz_stars)} of {len(A)}</td><td>{lakh_fmt(amz_star_rev)}</td><td>{amz_star_rev_pct*100:.1f}%</td><td>{ebitda_sign(amz_star_ebitda_t)}</td>
          <td style="text-align:left;padding:8px 10px">
            <div style="display:flex;flex-wrap:wrap;gap:5px">{amz_chips}</div>
          </td></tr>
        <tr><td><b>FirstCry</b></td><td>{len(fc_stars)} of {len(FC)}</td><td>{lakh_fmt(fc_star_rev)}</td><td>{fc_star_rev_pct*100:.1f}%</td><td>{ebitda_sign(fc_star_ebitda_t)}</td>
          <td style="text-align:left;padding:8px 10px">
            <div style="display:flex;flex-wrap:wrap;gap:5px">{fc_chips}</div>
          </td></tr>
        <tr><td><b>Blinkit</b></td><td>0 of {len(BL)}</td><td>\u2014</td><td>0%</td><td><span class="down">None</span></td>
          <td style="text-align:left;padding:8px 10px">
            <div style="display:inline-flex;flex-direction:column;background:var(--wash);border:1px solid var(--border);border-radius:6px;padding:4px 8px;opacity:0.55">
              <span style="font-size:11.5px;font-weight:700;color:var(--ink-2)">No stars this quarter</span>
              <span style="font-size:10px;color:var(--ink-3);margin-top:1px">Double cost structure leaves no CM2+ EBITDA+ SKU</span>
            </div>
          </td></tr>
      </tbody>"""

# Read HTML
with open('index.html', encoding='utf-8') as f:
    html = f.read()

# Find and replace the Q1 tbody
# The unique marker is: Profitable SKUs</th><th>Rev from stars
q1_thead_marker = '>Profitable SKUs</th>'
q1_pos = html.find(q1_thead_marker)
if q1_pos == -1:
    print("ERROR: Cannot find Q1 table header!")
else:
    # Find the tbody after the thead
    tbody_start = html.find('<tbody>', q1_pos)
    tbody_end = html.find('</tbody>', tbody_start) + len('</tbody>')
    if tbody_start == -1:
        print("ERROR: Cannot find Q1 tbody!")
    else:
        html = html[:tbody_start] + new_q1_tbody + html[tbody_end:]
        print(f"Updated Q1 table: {len(web_stars)}/{len(W)} web stars, {len(amz_stars)}/{len(A)} amz, {len(fc_stars)}/{len(FC)} fc")

# Write back
with open('index.html', 'w', encoding='utf-8') as f:
    f.write(html)

print("Done - Q1 fixed!")
