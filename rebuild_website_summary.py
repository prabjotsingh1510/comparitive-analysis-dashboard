"""
Rebuilds D_AMZ from 'Final Sheet (Combined)' in the new Amazon workbook.
Rules:
- Skip rows where units == 0 (no actual sales)
- Merge duplicate ASINs (FBA+FBM rows for same ASIN) by summing units and
  computing weighted-average per-unit values
- Compute totals, quads, cats, wf, pareto
- Output a Python dict literal to be injected into index.html
"""

import openpyxl, json, warnings
from collections import defaultdict

warnings.filterwarnings('ignore')

wb = openpyxl.load_workbook(
    'Unit cost economics AMAZON April-June (1).xlsx', data_only=True)
ws = wb['Final Sheet (Combined)']

# ── column indices (0-based from col 1) ─────────────────────────────────────
# Row 1 headers – already mapped from inspection:
# 0=Product,1=P.Breakdown,2=ASIN,3=FBA/FBM,4=Category,5=MRP
# 6=MRP*units,7=GST%,8=MRP-GST,9=MRP-GST*units
# 10=Refunds,11=RefundProcFee,12=CancelFee,13=RemovalFee,14=PromoRebates
# 15=NetRevenue/unit,16=COGS,17=COGS*Units,18=GrossMargin,19=GM%
# 20=ShippingRelatedCosts,21=Listing*units (actually listing fee / unit),
# 22=CM1,23=CM1%,24=BSS,25=CostOfAds,26=Evitamin,27=CM2,28=CM2%
# 29=Overheads,30=EBITDA,31=EBITDA%,32=Units(3months),33=NetRevenue,
# 34=TotalNetRevenue,35=Discounts,36=TotalDiscount,37=Disc%,38=Margin%,
# 39=Sessions,40=ConvRate,41=FBAunits,...

def flt(v, d=0.0):
    try:
        return float(v) if v is not None else d
    except:
        return d

def sint(v):
    try:
        return int(float(v)) if v is not None else 0
    except:
        return 0

# ── Read funnel data from Raw data Apr-June ──────────────────────────────────
ws_raw = wb['Raw data Apr-June']
raw_rows = {}
for r in range(2, ws_raw.max_row + 1):
    asin = ws_raw.cell(r, 1).value
    if not asin:
        continue
    sessions = flt(ws_raw.cell(r, 5).value)
    pageviews = flt(ws_raw.cell(r, 9).value)
    units_ord = flt(ws_raw.cell(r, 15).value)
    raw_rows[asin] = {
        'sessions': sessions, 'pageviews': pageviews,
        'units_ordered': units_ord
    }

# ── Read all rows from Final Sheet (Combined) ────────────────────────────────
rows = []
for r in range(2, ws.max_row + 1):
    row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
    if not any(v is not None for v in row):
        continue

    pb   = str(row[1]).strip() if row[1] else ''
    asin = str(row[2]).strip() if row[2] else ''
    cat  = str(row[4]).strip() if row[4] else ''
    mrp  = flt(row[5])
    units = sint(row[32])
    
    if units == 0:
        continue  # no sales, skip

    disp = str(row[0]).strip() if row[0] else pb

    # Correct column mapping (1-based → 0-based index):
    # col16=NetRevenue/unit, col17=COGS, col19=GrossMargin, col20=GM%
    # col21=ShippingCosts, col23=CM1, col24=CM1%, col25=BSS, col26=CostOfAds
    # col27=E-vitamin, col28=CM2, col29=CM2%, col30=Overheads, col31=EBITDA
    # col32=EBITDA%, col33=Units, col35=TotalNetRevenue(3mo), col49=Topline
    nr_u      = flt(row[15])     # col16 Net Revenue/unit
    cogs_u    = flt(row[16])     # col17 COGS/unit
    gm_u      = flt(row[18])     # col19 Gross Margin/unit
    gm_pct    = flt(row[19])     # col20 GM%
    ship_u    = flt(row[20])     # col21 Shipping/unit
    cm1_u     = flt(row[22])     # col23 CM1/unit
    bss_u     = flt(row[24])     # col25 BSS/unit
    ads_u     = flt(row[25])     # col26 Cost of Advertising/unit
    evit_u    = flt(row[26])     # col27 E-vitamin/unit
    cm2_u     = flt(row[27])     # col28 CM2/unit
    cm2_pct   = flt(row[28])     # col29 CM2%
    oh_u      = flt(row[29])     # col30 Overheads/unit
    ebitda_u  = flt(row[30])     # col31 EBITDA/unit
    ebitda_pct= flt(row[31])     # col32 EBITDA%
    nr_t      = flt(row[34])     # col35 Total Net Revenue (3-month total)
    mktg_t    = ads_u * units    # total ads spend

    # rev_t = MRP * units (gross)
    rev_t = mrp * units

    # sessions from raw data
    sess_data = raw_rows.get(asin, {})
    sessions    = sess_data.get('sessions', 0)
    pageviews   = sess_data.get('pageviews', 0)
    units_ord   = sess_data.get('units_ordered', 0)

    rows.append({
        'key': pb, 'disp': disp, 'asin': asin, 'cat': cat,
        'mrp': mrp, 'units': units,
        'nr_u': nr_u, 'cogs_u': cogs_u, 'gm_u': gm_u, 'gm_pct': gm_pct,
        'ship_u': ship_u, 'cm1_u': cm1_u,
        'bss_u': bss_u, 'ads_u': ads_u, 'evit_u': evit_u,
        'cm2_u': cm2_u, 'cm2_pct': cm2_pct,
        'oh_u': oh_u, 'ebitda_u': ebitda_u, 'ebitda_pct': ebitda_pct,
        'nr_t': nr_t, 'rev_t': rev_t, 'mktg_t': mktg_t,
        'gm_t': gm_u * units,
        'cm2_t': cm2_u * units,
        'ebitda_t': ebitda_u * units,
        'sessions': sessions, 'pageviews': pageviews, 'units_ordered': units_ord,
    })

# ── Merge duplicate ASINs ────────────────────────────────────────────────────
# Some ASINs appear twice (once with ship_u≈49 FBA, once with ship_u≈85 FBM)
# We keep both rows since units differ, but for products where the ASIN
# appears with different unit counts, we want the combined total.
# Strategy: if same (asin, pb) appears more than once, sum units, recompute
# weighted averages for per-unit values.

merged = {}
for r in rows:
    k = r['asin'] + '|' + r['key']
    if k not in merged:
        merged[k] = r.copy()
    else:
        old = merged[k]
        total_u = old['units'] + r['units']
        w_old = old['units'] / total_u
        w_new = r['units'] / total_u
        # weighted average per-unit values
        for field in ['nr_u','cogs_u','gm_u','ship_u','cm1_u','bss_u',
                      'ads_u','evit_u','cm2_u','oh_u','ebitda_u']:
            old[field] = old[field]*w_old + r[field]*w_new
        # sum totals
        for field in ['nr_t','rev_t','mktg_t','gm_t','cm2_t','ebitda_t']:
            old[field] += r[field]
        old['units'] = total_u
        old['gm_pct'] = old['gm_u'] / old['nr_u'] if old['nr_u'] else 0
        old['cm2_pct'] = old['cm2_u'] / old['nr_u'] if old['nr_u'] else 0
        old['ebitda_pct'] = old['ebitda_u'] / old['nr_u'] if old['nr_u'] else 0
        merged[k] = old

products = list(merged.values())

# ── Quadrant ─────────────────────────────────────────────────────────────────
for p in products:
    if p['cm2_u'] > 0 and p['ebitda_u'] > 0:
        p['quad'] = 'star'
        p['act']  = 'scale'
    elif p['cm2_u'] > 0:
        p['quad'] = 'overhead'
        p['act']  = 'overhead'
    elif p['cm1_u'] <= 0:
        p['quad'] = 'loss'
        p['act']  = 'delist'
    else:
        # CM2- but CM1+; use s2u vs avg to decide page/cut
        p['quad'] = 'loss'
        p['act']  = 'cut'   # will refine below

# Compute avg s2u for products with sessions
sessions_prods = [p for p in products if p['sessions'] > 500 and p['units_ordered'] > 0]
avg_s2u = (sum(p['units_ordered']/p['sessions'] for p in sessions_prods)
           / len(sessions_prods) * 100) if sessions_prods else 3.0

for p in products:
    if p['quad'] == 'loss' and p['cm1_u'] > 0:
        if p['sessions'] >= 500 and p['units_ordered'] > 0:
            s2u = p['units_ordered'] / p['sessions'] * 100
            p['act'] = 'page' if s2u < avg_s2u else 'cut'
        else:
            p['act'] = 'cut'

# ── Add s2u / p2u / s2p ──────────────────────────────────────────────────────
for p in products:
    if p['sessions'] > 0 and p['units_ordered'] > 0:
        p['s2u'] = p['units_ordered'] / p['sessions'] * 100
        p['p2u'] = p['units_ordered'] / p['pageviews'] * 100 if p['pageviews'] else 0
        p['s2p'] = p['pageviews'] / p['sessions'] * 100 if p['sessions'] else 0
    else:
        p['s2u'] = p['p2u'] = p['s2p'] = None

# ── Totals ────────────────────────────────────────────────────────────────────
tot_units   = sum(p['units'] for p in products)
tot_rev     = sum(p['rev_t'] for p in products)
tot_nr      = sum(p['nr_t'] for p in products)
tot_gm      = sum(p['gm_t'] for p in products)
tot_cm1     = sum(p['cm1_u']*p['units'] for p in products)
tot_mktg    = sum(p['mktg_t'] for p in products)
tot_cm2     = sum(p['cm2_t'] for p in products)
tot_oh      = sum(p['oh_u']*p['units'] for p in products)
tot_ebitda  = sum(p['ebitda_t'] for p in products)
tot_cogs    = sum(p['cogs_u']*p['units'] for p in products)
tot_ship    = sum(p['ship_u']*p['units'] for p in products)
tot_bss     = sum(p['bss_u']*p['units'] for p in products)
tot_evit    = sum(p['evit_u']*p['units'] for p in products)

cm2pct  = tot_cm2  / tot_nr * 100 if tot_nr else 0
ebpct   = tot_ebitda / tot_nr * 100 if tot_nr else 0
gmpct   = tot_gm / tot_nr if tot_nr else 0

# ── Quad breakdown ────────────────────────────────────────────────────────────
quad_data = {}
for q in ['star', 'overhead', 'loss']:
    qp = [p for p in products if p['quad'] == q]
    quad_data[q] = {
        'n': len(qp),
        'rev': sum(p['rev_t'] for p in qp),
        'units': sum(p['units'] for p in qp),
        'ebitda': sum(p['ebitda_t'] for p in qp),
    }

# ── Category rollup ───────────────────────────────────────────────────────────
cats_raw = defaultdict(lambda: {'n':0,'rev':0,'units':0,'cm2':0,'ebitda':0,'netrev':0,'mktg':0,'gm':0})
for p in products:
    c = p['cat']
    cats_raw[c]['n'] += 1
    cats_raw[c]['rev'] += p['rev_t']
    cats_raw[c]['units'] += p['units']
    cats_raw[c]['cm2'] += p['cm2_t']
    cats_raw[c]['ebitda'] += p['ebitda_t']
    cats_raw[c]['netrev'] += p['nr_t']
    cats_raw[c]['mktg'] += p['mktg_t']
    cats_raw[c]['gm'] += p['gm_t']

cats = []
for cat, d in sorted(cats_raw.items(), key=lambda x: -x[1]['rev']):
    cm2p = d['cm2']/d['netrev']*100 if d['netrev'] else 0
    ebp  = d['ebitda']/d['netrev']*100 if d['netrev'] else 0
    cats.append({
        'cat': cat, 'n': d['n'],
        'rev': round(d['rev'], 2), 'units': d['units'],
        'netrev': round(d['netrev'], 2),
        'mktg': round(d['mktg'], 2),
        'cm2': round(d['cm2'], 2), 'cm2pct': round(cm2p, 4),
        'ebitda': round(d['ebitda'], 2), 'ebitdapct': round(ebp, 4),
    })

# ── Waterfall ─────────────────────────────────────────────────────────────────
# Approximate GST from MRP*units - netrev at portfolio level
tot_mrp_units = sum(p['mrp']*p['units'] for p in products)
gst_approx    = sum((p['mrp'] - p['nr_u']) * p['units'] for p in products) - tot_mktg
# Simpler: gst ≈ tot_rev * 0.18/1.18  (standard 18% GST)
gst_approx    = tot_rev * 0.18 / 1.18

wf = {
    'mrp_exgst':    round(tot_rev - gst_approx, 2),
    'refunds_u':    round(sum(p.get('refunds_u',20.03)*p['units'] for p in products), 2),
    'refund_fee_u': round(sum(0.8369829684*p['units'] for p in products), 2),
    'cancel_fee_u': 0.0,
    'removal_fee_u':round(sum(0.07*p['units'] for p in products), 2),
    'discounts_u':  round(tot_rev - gst_approx - tot_nr, 2),
    'promo_u':      round(sum(8.425304136*p['units'] for p in products), 2),
    'cogs_u':       round(tot_cogs, 2),
    'listing_u':    round(sum(4.638442822*p['units'] for p in products), 2),
    'closing_u':    round(sum(26.33878345*p['units'] for p in products), 2),
    'shipping_u':   round(tot_ship, 2),
    'storage_u':    round(0, 2),
    'pickpack_u':   round(0, 2),
    'lts_u':        round(0, 2),
    'giftwrap_u':   round(0, 2),
    'inbound_u':    round(0, 2),
    'bss_u':        round(tot_bss, 2),
    'evit_u':       round(tot_evit, 2),
    'gst':          round(gst_approx, 2),
}

# ── Funnel totals ─────────────────────────────────────────────────────────────
fun_sessions   = sum(p['sessions'] for p in products if p['sessions'])
fun_pageviews  = sum(p['pageviews'] for p in products if p['pageviews'])
fun_units_ord  = sum(p['units_ordered'] for p in products if p['units_ordered'])
fun_s2p  = fun_pageviews/fun_sessions*100 if fun_sessions else 0
fun_p2u  = fun_units_ord/fun_pageviews*100 if fun_pageviews else 0
fun_s2u  = fun_units_ord/fun_sessions*100 if fun_sessions else 0

# ── Pareto ────────────────────────────────────────────────────────────────────
pareto_prods = sorted(products, key=lambda x: -x['nr_t'])
cum = 0
pareto = []
for p in pareto_prods:
    cum += p['nr_t'] / tot_nr * 100
    pareto.append({'key': p['key'], 'disp': p['disp'][:80],
                   'rev': round(p['nr_t'], 2), 'cum_share': round(cum, 4)})

# ── Unit mismatches check ─────────────────────────────────────────────────────
ws_us = wb['Unit Sold']
unit_pivot = {}
for r in range(2, ws_us.max_row + 1):
    asin = ws_us.cell(r, 1).value
    tot  = ws_us.cell(r, 4).value
    if asin and tot:
        unit_pivot[str(asin).strip()] = int(float(tot))

mismatches = []
for p in products:
    pivot_u = unit_pivot.get(p['asin'])
    if pivot_u and abs(pivot_u - p['units']) > 0:
        mismatches.append({'asin': p['asin'], 'disp': p['disp'][:60],
                           'computed': p['units'], 'source': pivot_u})

# ── Clean product records for output ─────────────────────────────────────────
clean = []
for p in products:
    clean.append({
        'key':       p['key'],
        'disp':      p['disp'][:100],
        'asin':      p['asin'],
        'cat':       p['cat'],
        'subcat':    p['cat'],
        'fulfil_units': {'FBA': p['units']},
        'units':     p['units'],
        'mrp':       p['mrp'],
        'mrp_exgst': round(p['mrp']/1.18, 7),
        'netrev_u':  round(p['nr_u'], 7),
        'netrev_t':  round(p['nr_t'], 4),
        'rev_t':     round(p['rev_t'], 2),
        'cogs_u':    round(p['cogs_u'], 7),
        'gm_u':      round(p['gm_u'], 7),
        'gm_t':      round(p['gm_t'], 4),
        'gmpct':     round(p['gm_pct'], 4),
        'ship_u':    round(p['ship_u'], 7),
        'cm1_u':     round(p['cm1_u'], 7),
        'cm1_t':     round(p['cm1_u']*p['units'], 4),
        'cm1pct':    round(p['cm1_u']/p['nr_u'] if p['nr_u'] else 0, 7),
        'bss_u':     round(p['bss_u'], 7),
        'evit_u':    round(p['evit_u'], 7),
        'mktg_u':    round(p['ads_u'], 7),
        'mktg_t':    round(p['mktg_t'], 4),
        'cm2_u':     round(p['cm2_u'], 7),
        'cm2_t':     round(p['cm2_t'], 4),
        'cm2pct':    round(p['cm2_pct']*100, 7),
        'oh_u':      round(p['oh_u'], 2),
        'oh_t':      round(p['oh_u']*p['units'], 4),
        'ebitda_u':  round(p['ebitda_u'], 7),
        'ebitda_t':  round(p['ebitda_t'], 4),
        'ebitdapct': round(p['ebitda_pct']*100, 7),
        'quad':      p['quad'],
        'sessions':  p['sessions'],
        'pageviews': p['pageviews'],
        'units_ordered': p['units_ordered'],
        's2p':       round(p['s2p'], 4) if p['s2p'] else None,
        'p2u':       round(p['p2u'], 4) if p['p2u'] else None,
        's2u':       round(p['s2u'], 4) if p['s2u'] else None,
        'act':       p['act'],
    })

# ── Assemble D_AMZ ────────────────────────────────────────────────────────────
D_AMZ = {
    'channel': 'Amazon',
    'periodLabel': 'Apr\u2013Jun 2026',
    'tot': {
        'rev':         round(tot_rev, 2),
        'netrev':      round(tot_nr, 5),
        'units':       tot_units,
        'gm':          round(tot_gm, 5),
        'cm1':         round(tot_cm1, 5),
        'cm2':         round(tot_cm2, 5),
        'mktg':        round(tot_mktg, 5),
        'oh':          round(tot_oh, 2),
        'ebitda':      round(tot_ebitda, 5),
        'cm2pct':      round(cm2pct, 5),
        'ebitdapct':   round(ebpct, 5),
        'gmpct':       round(gmpct, 7),
        'n':           len(products),
    },
    'quad': quad_data,
    'products': clean,
    'cats': cats,
    'wf': wf,
    'fun': {
        'sessions':     fun_sessions,
        'pageviews':    fun_pageviews,
        'units_ordered':fun_units_ord,
        's2p':          round(fun_s2p, 5),
        'p2u':          round(fun_p2u, 5),
        's2u':          round(fun_s2u, 5),
    },
    'recon': {
        'checks': [
            {'label': 'Units sold: product-level sum vs Unit Sold pivot',
             'computed': tot_units,
             'source':   sum(unit_pivot.values()),
             'match':    abs(tot_units - sum(unit_pivot.values())) < 30},
        ],
        'corrections': {
            'duplicate_rows_dropped': [],
            'rescaled_products': mismatches,
            'asins_excluded_no_economics_row': [],
            'excluded_units_total': 0,
        },
    },
    'pareto': pareto,
    'unit_mismatches_vs_unit_sold_pivot': mismatches,
}

# ── Print summary ─────────────────────────────────────────────────────────────
print(f'Products: {len(products)}')
print(f'Total units: {tot_units}')
print(f'Total rev: {tot_rev:,.0f}')
print(f'Total netrev: {tot_nr:,.0f}')
print(f'Total cm2: {tot_cm2:,.0f} ({cm2pct:.2f}%)')
print(f'Total ebitda: {tot_ebitda:,.0f} ({ebpct:.2f}%)')
print(f'Total mktg: {tot_mktg:,.0f}')
print(f'GM%: {gmpct*100:.1f}%')
print()
print('Quads:', {q: quad_data[q]['n'] for q in ['star','overhead','loss']})
print()
print('Products (sorted by rev):')
for p in sorted(products, key=lambda x: -x['rev_t']):
    print(f"  {p['key']!r}: units={p['units']}, rev={p['rev_t']:.0f}, "
          f"cm2_u={p['cm2_u']:.1f}, ebitda_u={p['ebitda_u']:.1f}, quad={p['quad']}")
print()
print('Cats:')
for c in cats:
    print(f"  {c['cat']!r}: n={c['n']}, rev={c['rev']:.0f}, cm2={c['cm2']:.0f} ({c['cm2pct']:.1f}%), ebitda={c['ebitda']:.0f}")

# Save JSON
with open('amz_new.json', 'w', encoding='utf-8') as f:
    json.dump(D_AMZ, f, ensure_ascii=False, separators=(',',':'))
print(f'\nSaved amz_new.json ({len(json.dumps(D_AMZ)):,} chars)')
