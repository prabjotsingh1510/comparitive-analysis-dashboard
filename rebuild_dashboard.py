"""
Comprehensive rebuild script for the Dabble Dashboard.
Reads Channel Economics Dashboard.xlsx and updates index.html with correct data.
Changes:
1. Update all channel data (Website, Amazon, FirstCry, Blinkit)
2. Update Channel Analysis decision criteria 
3. Remove "Important Metrics" tab
4. Add new "Marketing Spend Comparison" tab
"""
import openpyxl
import json
import re

wb = openpyxl.load_workbook('Channel Economics Dashboard.xlsx', data_only=True)

def safe(v):
    if v is None:
        return None
    try:
        f = float(v)
        return f
    except (TypeError, ValueError):
        return str(v) if v else None

def pct(v, decimals=1):
    if v is None: return 'N/A'
    return f"{v*100:.{decimals}f}%"

def inr(v, decimals=0):
    if v is None: return '—'
    if abs(v) >= 100000:
        return f"₹{v/100000:.2f} L"
    elif abs(v) >= 1000:
        return f"₹{v:,.0f}"
    else:
        return f"₹{v:.{decimals}f}"

def inr_simple(v):
    if v is None: return '—'
    return f"₹{v:,.0f}"

# ============================================================
# EXTRACT ALL DATA
# ============================================================

# --- WEBSITE ---
ws_web = wb['Raw Data - Website']
web_headers = [ws_web.cell(row=4, column=c).value for c in range(1, 50)]

website_products = []
for r in range(5, 200):
    code = ws_web.cell(row=r, column=1).value
    if code is None: continue
    row = {}
    for c, h in enumerate(web_headers[:42], 1):
        if h:
            row[h] = safe(ws_web.cell(row=r, column=c).value)
    website_products.append(row)

# Compute totals for website
# Key fields: Units sold, Net Revenue without GST, CM2, CM2%, EBITA, EBITA%, Marketing Cost, 3 month revenue with GST, 3 month revenue without GST
total_web_units = sum(p.get('Units sold') or 0 for p in website_products)
total_web_rev_gst = sum((p.get('3 month revenue with GST') or 0) for p in website_products)
total_web_rev_no_gst = sum((p.get('3 month revenue without GST') or 0) for p in website_products)
total_web_net_rev = sum((p.get('Net Revenue without GST') or 0) * (p.get('Units sold') or 0) for p in website_products)
total_web_cm2 = sum((p.get('CM2') or 0) * (p.get('Units sold') or 0) for p in website_products)
total_web_ebita = sum((p.get('EBITA') or 0) * (p.get('Units sold') or 0) for p in website_products)
total_web_mktg = sum((p.get('Marketing Cost') or 0) * (p.get('Units sold') or 0) for p in website_products)
total_web_gm = sum((p.get('Gross margin') or 0) * (p.get('Units sold') or 0) for p in website_products)

web_cm2pct = total_web_cm2 / total_web_net_rev if total_web_net_rev else 0
web_ebitapct = total_web_ebita / total_web_net_rev if total_web_net_rev else 0
web_gmpct = total_web_gm / total_web_net_rev if total_web_net_rev else 0
web_mktg_pct_of_net = total_web_mktg / total_web_net_rev if total_web_net_rev else 0

# Website profitable products (CM2>0 AND EBITDA>0)
web_stars = [(p.get('Product Name (Code)',''), p.get('CM2',0), p.get('EBITA',0)) 
             for p in website_products 
             if (p.get('CM2') or 0) > 0 and (p.get('EBITA') or 0) > 0]
web_star_count = len(web_stars)
web_star_rev = sum((p.get('Net Revenue without GST') or 0) * (p.get('Units sold') or 0) 
                   for p in website_products 
                   if (p.get('CM2') or 0) > 0 and (p.get('EBITA') or 0) > 0)
web_star_ebitda = sum((p.get('EBITA') or 0) * (p.get('Units sold') or 0)
                      for p in website_products 
                      if (p.get('CM2') or 0) > 0 and (p.get('EBITA') or 0) > 0)
total_web_skus = len(website_products)

print(f"Website: {total_web_skus} SKUs, {total_web_units} units")
print(f"  Rev w/GST: {total_web_rev_gst:,.0f}")
print(f"  Rev no/GST: {total_web_rev_no_gst:,.0f}")
print(f"  Net Rev: {total_web_net_rev:,.0f}")
print(f"  CM2: {total_web_cm2:,.0f} ({web_cm2pct*100:.1f}%)")
print(f"  EBITA: {total_web_ebita:,.0f} ({web_ebitapct*100:.1f}%)")
print(f"  Marketing: {total_web_mktg:,.0f} ({web_mktg_pct_of_net*100:.1f}%)")
print(f"  Stars (CM2+ EBITDA+): {web_star_count} of {total_web_skus}")
print(f"  Top stars: {sorted(web_stars, key=lambda x: -x[2])[:5]}")

# --- AMAZON ---
ws_amz = wb['Raw Data - Amazon']
amz_headers = [ws_amz.cell(row=4, column=c).value for c in range(1, 50)]
amz_headers_full = [ws_amz.cell(row=4, column=c).value for c in range(1, 50)]

amazon_products = []
for r in range(5, 200):
    prod = ws_amz.cell(row=r, column=1).value
    if prod is None: continue
    row = {}
    for c, h in enumerate(amz_headers[:45], 1):
        if h:
            row[h] = safe(ws_amz.cell(row=r, column=c).value)
    amazon_products.append(row)

# Get Amazon extended headers for CM2, EBITDA columns
# Headers go: Product, P.Breakdown, ASIN, Category, Sub-Cat, FBA/FBM, Units, MRP, GST%, MRP-GST, 
# Refunds, Refund Proc Fee, Cancel Fee, Removal Fee, Discounts, Promo Rebates, Net Rev, COGS,
# Gross Margin, GM%, Listing Fee, Closing Fee, Shipping, Storage, Pick&Pack, LTSFBA, Gift, Inbound, CM1, CM1%,
# Marketing Cost (col 31?), CM2(col32?), CM2%(col33?), Overhead(col34?), EBITDA(col35?), EBITDA%(col36?)

# Let's check
for c in range(28, 50):
    v = ws_amz.cell(row=4, column=c).value
    print(f"  AMZ col {c}: {v}")

print()

# Re-extract amazon with extended cols
amazon_products2 = []
for r in range(5, 200):
    prod = ws_amz.cell(row=r, column=1).value
    if prod is None: continue
    row = {}
    for c in range(1, 50):
        h = amz_headers_full[c-1] if c-1 < len(amz_headers_full) else f"col{c}"
        row[h or f"col{c}"] = safe(ws_amz.cell(row=r, column=c).value)
    amazon_products2.append(row)

# Print first product keys
print("Amazon product keys:", list(amazon_products2[0].keys())[:45])
print("Amazon first product sample:", {k: v for k, v in amazon_products2[0].items() if v is not None and k})
