"""
Pull gross revenue (MRP x units) for every product from each channel's source Excel.
Amazon  : 'Unit cost economics AMAZON April-June (2).xlsx'  -> Final Sheet (Combined): col7=MRP*units, col33=Units, col6=MRP
FirstCry: 'Firstcry_Unit Cost Economic(Apr-June) (1).xlsx'  -> SP Vendor Format sheet: sell_price * units
Blinkit : 'Blinkit_Unit Cost Economic(Apr-June) (1).xlsx'   -> Final sheet: sp * units
"""
import openpyxl, warnings, json, re
warnings.filterwarnings('ignore')

# ── AMAZON ─────────────────────────────────────────────────────────────────
wb_a = openpyxl.load_workbook('Unit cost economics AMAZON April-June (2).xlsx', data_only=True)
ws_a = wb_a['Final Sheet (Combined)']

print('=== AMAZON: Final Sheet (Combined) ===')
print('Headers (cols 1-35):')
hdrs = [ws_a.cell(1,c).value for c in range(1,36)]
for i,h in enumerate(hdrs,1):
    if h: print(f'  col{i}: {h!r}')
print()

# col6=MRP, col7=MRP*units (but is None for many), col33=Units, col35=Total Net Revenue
# Check first few data rows
print('First 5 data rows (key, mrp, mrp*units, units, total_nr):')
for r in range(2,7):
    pb     = ws_a.cell(r,2).value
    mrp    = ws_a.cell(r,6).value
    mrp_t  = ws_a.cell(r,7).value
    units  = ws_a.cell(r,33).value
    tot_nr = ws_a.cell(r,35).value
    if pb:
        gross = mrp*units if mrp and units else None
        print(f'  {pb!r:20s} mrp={mrp} mrp*units(col7)={mrp_t} units={units} total_nr(col35)={tot_nr} calc_gross={gross}')

# ── FIRSTCRY ───────────────────────────────────────────────────────────────
print()
wb_fc = openpyxl.load_workbook('Firstcry_Unit Cost Economic(Apr-June) (1).xlsx', data_only=True)
print('FirstCry sheets:', wb_fc.sheetnames)
ws_fc = wb_fc[wb_fc.sheetnames[0]]
print('FC Sheet headers (cols 1-20):')
for c in range(1,21):
    h = ws_fc.cell(1,c).value
    if h: print(f'  col{c}: {h!r}')
print()
print('First 5 FC rows:')
for r in range(2,7):
    row = [ws_fc.cell(r,c).value for c in range(1,20)]
    if any(v is not None for v in row):
        print(f'  {row[:8]}')

# ── BLINKIT ───────────────────────────────────────────────────────────────
print()
wb_bl = openpyxl.load_workbook('Blinkit_Unit Cost Economic(Apr-June) (1).xlsx', data_only=True)
print('Blinkit sheets:', wb_bl.sheetnames)
ws_bl = wb_bl[wb_bl.sheetnames[0]]
print('BL Sheet headers (cols 1-20):')
for c in range(1,21):
    h = ws_bl.cell(1,c).value
    if h: print(f'  col{c}: {h!r}')
print()
print('First 5 BL rows:')
for r in range(2,7):
    row = [ws_bl.cell(r,c).value for c in range(1,20)]
    if any(v is not None for v in row):
        print(f'  {row[:8]}')
