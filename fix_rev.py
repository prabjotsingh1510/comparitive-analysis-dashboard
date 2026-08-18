"""
Fix rev_t for every product in D_AMZ, D_FC, D_BL using source Excel files.
Revenue definition per channel:
  Amazon   -> col35 'Total Net Revenue' (the workbook's own 3-month total)
  FirstCry -> col11 (SELL PRICE) × col10 (Units)
  Blinkit  -> col7  (SP)         × col5  (Units)
"""
import openpyxl, warnings, json, re
warnings.filterwarnings('ignore')

def flt(v, d=0.0):
    try: return float(v) if v is not None else d
    except: return d
def sint(v):
    try: return int(float(v)) if v is not None else 0
    except: return 0

# ════════════════════════════════════════════════════════════════
# 1. AMAZON  — key by P.Breakdown + ASIN  (some keys appear >1 row: FBA+FBM)
#    Revenue = col35 Total Net Revenue (already the 3-mo sum)
# ════════════════════════════════════════════════════════════════
wb_a = openpyxl.load_workbook('Unit cost economics AMAZON April-June (2).xlsx', data_only=True)
ws_a = wb_a['Final Sheet (Combined)']

amz_rev = {}   # key -> (total_nr, units, mrp)
for r in range(2, ws_a.max_row + 1):
    pb    = ws_a.cell(r, 2).value
    asin  = ws_a.cell(r, 3).value
    mrp   = flt(ws_a.cell(r, 6).value)
    units = sint(ws_a.cell(r, 33).value)
    tot_nr= flt(ws_a.cell(r, 35).value)   # col35 = Total Net Revenue

    if not pb or units == 0:
        continue

    pb = str(pb).strip()
    asin = str(asin).strip() if asin else ''
    k = pb + '|' + asin

    if k not in amz_rev:
        amz_rev[k] = {'tot_nr': 0.0, 'units': 0, 'mrp': mrp}
    amz_rev[k]['tot_nr'] += tot_nr
    amz_rev[k]['units']  += units
    if mrp: amz_rev[k]['mrp'] = mrp

print('Amazon keys:', len(amz_rev))
for k, v in sorted(amz_rev.items(), key=lambda x: -x[1]['tot_nr'])[:6]:
    print(f'  {k!r}: tot_nr={v["tot_nr"]:,.0f}  units={v["units"]}  mrp={v["mrp"]}')

# ════════════════════════════════════════════════════════════════
# 2. FIRSTCRY — key by P.Breakdown
#    Revenue = SELL PRICE (col11) × Units (col10)   — gross revenue
# ════════════════════════════════════════════════════════════════
wb_fc = openpyxl.load_workbook('Firstcry_Unit Cost Economic(Apr-June) (1).xlsx', data_only=True)
ws_fc = wb_fc['SP Vendor Format']

fc_rev = {}
for r in range(2, ws_fc.max_row + 1):
    pb    = ws_fc.cell(r, 4).value    # col4  P.Breakdown
    units = sint(ws_fc.cell(r, 10).value)  # col10 Units
    sp    = flt(ws_fc.cell(r, 11).value)   # col11 SELL PRICE
    mrp   = flt(ws_fc.cell(r, 7).value)    # col7  MRP

    if not pb or units == 0:
        continue
    pb = str(pb).strip()
    gross = sp * units

    if pb not in fc_rev:
        fc_rev[pb] = {'gross': 0.0, 'units': 0, 'mrp': mrp, 'sp': sp}
    fc_rev[pb]['gross'] += gross
    fc_rev[pb]['units'] += units

print('\nFirstCry keys:', len(fc_rev))
for k, v in sorted(fc_rev.items(), key=lambda x: -x[1]['gross'])[:6]:
    print(f'  {k!r}: gross={v["gross"]:,.0f}  units={v["units"]}  mrp={v["mrp"]}  sp={v["sp"]}')

# ════════════════════════════════════════════════════════════════
# 3. BLINKIT — key by P.Breakdown (col3)
#    Revenue = SP (col7) × Units (col5)  — actual sell price x units
# ════════════════════════════════════════════════════════════════
wb_bl = openpyxl.load_workbook('Blinkit_Unit Cost Economic(Apr-June) (1).xlsx', data_only=True)
ws_bl = wb_bl['Final(Unit Cost Format)']

bl_rev = {}
for r in range(2, ws_bl.max_row + 1):
    pb    = ws_bl.cell(r, 3).value    # col3  P.Breakdown
    units = sint(ws_bl.cell(r, 5).value)   # col5  Units
    sp    = flt(ws_bl.cell(r, 7).value)    # col7  SP
    mrp   = flt(ws_bl.cell(r, 6).value)    # col6  MRP

    if not pb or units == 0:
        continue
    pb = str(pb).strip()
    gross = sp * units

    if pb not in bl_rev:
        bl_rev[pb] = {'gross': 0.0, 'units': 0, 'mrp': mrp, 'sp': sp}
    bl_rev[pb]['gross'] += gross
    bl_rev[pb]['units'] += units

print('\nBlinkit keys:', len(bl_rev))
for k, v in sorted(bl_rev.items(), key=lambda x: -x[1]['gross'])[:8]:
    print(f'  {k!r}: gross={v["gross"]:,.0f}  units={v["units"]}  mrp={v["mrp"]}  sp={v["sp"]}')

# ════════════════════════════════════════════════════════════════
# 4. Patch index.html
# ════════════════════════════════════════════════════════════════
with open('index.html', encoding='utf-8') as f:
    html = f.read()

def load_var(html, var_name):
    m = re.search('const ' + var_name + r' = (\{.*?\});', html, re.DOTALL)
    return json.loads(m.group(1))

def save_var(html, var_name, obj):
    new_js = json.dumps(obj, ensure_ascii=False, separators=(',', ':'))
    start_tok = 'const ' + var_name + ' = {'
    idx_start = html.find(start_tok)
    depth = 0
    i = idx_start + len('const ' + var_name + ' = ')
    while i < len(html):
        c = html[i]
        if c == '{': depth += 1
        elif c == '}':
            depth -= 1
            if depth == 0:
                idx_end = i + 1
                break
        i += 1
    s = html.find(';', idx_end)
    if s != -1 and s < idx_end + 3:
        idx_end = s + 1
    old = html[idx_start:idx_end]
    new = 'const ' + var_name + ' = ' + new_js + ';'
    print(f'{var_name}: {len(old):,} -> {len(new):,} chars')
    return html[:idx_start] + new + html[idx_end:]

# ── Patch D_AMZ ──────────────────────────────────────────────────
amz = load_var(html, 'D_AMZ')
matched_amz = 0
unmatched_amz = []
for p in amz['products']:
    key_asin = p['key'] + '|' + p.get('asin', '')
    if key_asin in amz_rev:
        p['rev_t'] = round(amz_rev[key_asin]['tot_nr'], 2)
        p['units'] = amz_rev[key_asin]['units']
        matched_amz += 1
    elif p['key'] in {k.split('|')[0]: True for k in amz_rev}:
        # try key-only match (some ASINs differ)
        candidates = [(k, v) for k, v in amz_rev.items() if k.startswith(p['key'] + '|')]
        if candidates:
            best = max(candidates, key=lambda x: x[1]['tot_nr'])
            p['rev_t'] = round(best[1]['tot_nr'], 2)
            p['units'] = best[1]['units']
            matched_amz += 1
        else:
            unmatched_amz.append(p['key'])
    else:
        # fallback: MRP * units already stored
        p['rev_t'] = round(p.get('mrp', 0) * p.get('units', 0), 2)
        unmatched_amz.append(p['key'])

print(f'\nAMZ matched: {matched_amz}/{len(amz["products"])}, unmatched: {unmatched_amz}')

# Recompute AMZ tot.rev as sum of rev_t
amz['tot']['rev']    = round(sum(p['rev_t'] for p in amz['products']), 2)
amz['tot']['netrev'] = round(sum(p.get('netrev_t', p['rev_t']) for p in amz['products']), 2)

# ── Patch D_FC ───────────────────────────────────────────────────
fc = load_var(html, 'D_FC')
matched_fc = 0
unmatched_fc = []
for p in fc['products']:
    key = p.get('key', p.get('sku', ''))
    if key in fc_rev:
        p['rev_t'] = round(fc_rev[key]['gross'], 2)
        matched_fc += 1
    else:
        unmatched_fc.append(key)

print(f'FC matched: {matched_fc}/{len(fc["products"])}, unmatched: {unmatched_fc}')

fc['tot']['rev'] = round(sum(p.get('rev_t', 0) for p in fc['products']), 2)

# ── Patch D_BL ───────────────────────────────────────────────────
bl = load_var(html, 'D_BL')
matched_bl = 0
unmatched_bl = []
for p in bl['products']:
    key = p.get('key', '')
    if key in bl_rev:
        p['rev_t'] = round(bl_rev[key]['gross'], 2)
        matched_bl += 1
    else:
        unmatched_bl.append(key)

print(f'BL matched: {matched_bl}/{len(bl["products"])}, unmatched: {unmatched_bl}')

bl['tot']['rev'] = round(sum(p.get('rev_t', 0) for p in bl['products']), 2)

# ── Write back ───────────────────────────────────────────────────
html = save_var(html, 'D_AMZ', amz)
html = save_var(html, 'D_FC',  fc)
html = save_var(html, 'D_BL',  bl)

with open('index.html', 'w', encoding='utf-8') as f:
    f.write(html)
print('\nindex.html written. Size:', len(html))

# ── Final verification ───────────────────────────────────────────
print('\n=== VERIFICATION ===')
for var in ['D_AMZ','D_FC','D_BL']:
    m = re.search('const '+var+r' = (\{.*?\});', html, re.DOTALL)
    d = json.loads(m.group(1))
    prods = d['products']
    zero = [p['key'] for p in prods if p.get('rev_t', 0) <= 0]
    print(f'{var}: tot.rev={d["tot"]["rev"]:,.0f}  products={len(prods)}  zero_rev_t={len(zero)}')
    for p in sorted(prods, key=lambda x: -x.get('rev_t', 0))[:4]:
        print(f'  {p["key"]!r:25s}  rev_t={p["rev_t"]:>12,.0f}  units={p["units"]:>5}')
    if zero: print(f'  ZERO REV: {zero}')
    print()
