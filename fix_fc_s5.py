"""
Fix D_FC.wf and D_FC.fun so FirstCry Money & Traffic Funnel (f5) opens.

firstcryWfStages needs:  w.gst, w.retcancel_u, w.discounts_u, w.cogs_u
renderChanS5 needs:      D.fun  (funnelMode='none' so no funnel chart,
                                 but the field must exist or it crashes)
D_FC.tot also needs:     fcm  (total FirstCry Margin) for the waterfall
"""
import openpyxl, warnings, json, re
warnings.filterwarnings('ignore')

def flt(v, d=0.0):
    try: return float(v) if v is not None else d
    except: return d
def sint(v):
    try: return int(float(v)) if v is not None else 0
    except: return 0

with open('index.html', encoding='utf-8') as f:
    html = f.read()

def load_var(var):
    m = re.search('const ' + var + r' = (\{.*?\});', html, re.DOTALL)
    return json.loads(m.group(1))

def save_var(html, var, obj):
    js = json.dumps(obj, ensure_ascii=False, separators=(',', ':'))
    tok = 'const ' + var + ' = {'
    idx_s = html.find(tok)
    depth, i = 0, idx_s + len('const ' + var + ' = ')
    while i < len(html):
        c = html[i]
        if c == '{': depth += 1
        elif c == '}':
            depth -= 1
            if depth == 0: idx_e = i + 1; break
        i += 1
    s = html.find(';', idx_e)
    if s != -1 and s < idx_e + 3: idx_e = s + 1
    new = 'const ' + var + ' = ' + js + ';'
    print(f'  {var}: {len(html[idx_s:idx_e]):,} -> {len(new):,}')
    return html[:idx_s] + new + html[idx_e:]

# ── Load source Excel ─────────────────────────────────────────────────────
wb = openpyxl.load_workbook(
    'Firstcry_Unit Cost Economic(Apr-June) (1).xlsx', data_only=True)
ws = wb['SP Vendor Format']

# Columns (1-based):
# col4=P.Breakdown, col7=MRP, col10=Units, col11=SELL PRICE,
# col12=MRP-GST (per unit), col13=Discounts (per unit),
# col14=Return/Cancellation (per unit),
# col15=Net Revenue without GST (per unit),
# col16=COGS, col17=Gross Margin, col18=GM%,
# col19=Firstcry Margin (per unit),
# col20=CM2, col21=CM2%, col22=Overheads, col23=EBITDA, col24=EBITDA%

print('Reading FirstCry Excel...')
# Verify headers
hdrs = {c: ws.cell(1, c).value for c in range(1, 30)}
for c, h in hdrs.items():
    if h: print(f'  col{c}: {h!r}')

print()
tot_gm = tot_cogs = tot_fcm = tot_ret = tot_disc = tot_nr = 0.0
tot_units = 0

for r in range(2, ws.max_row + 1):
    pb    = ws.cell(r, 4).value
    units = sint(ws.cell(r, 10).value)
    if not pb or units == 0:
        continue

    disc_u   = flt(ws.cell(r, 13).value)   # discounts per unit
    ret_u    = flt(ws.cell(r, 14).value)   # return/cancellation per unit
    nr_u     = flt(ws.cell(r, 15).value)   # net rev without GST per unit
    cogs_u   = flt(ws.cell(r, 16).value)   # COGS per unit
    gm_u     = flt(ws.cell(r, 17).value)   # gross margin per unit
    fcm_u    = flt(ws.cell(r, 19).value)   # FirstCry margin per unit

    tot_cogs  += cogs_u  * units
    tot_fcm   += fcm_u   * units
    tot_ret   += ret_u   * units
    tot_disc  += disc_u  * units
    tot_nr    += nr_u    * units
    tot_gm    += gm_u    * units
    tot_units += units

print(f'Totals from Excel: units={tot_units} nr={tot_nr:,.0f} cogs={tot_cogs:,.0f} fcm={tot_fcm:,.0f} ret={tot_ret:,.0f} disc={tot_disc:,.0f} gm={tot_gm:,.0f}')

# ── Patch D_FC ────────────────────────────────────────────────────────────
fc = load_var('D_FC')
T  = fc['tot']

# Gross revenue already in T.rev
gross = T['rev']

# GST: products are GST-inclusive MRP; net rev is ex-GST
# GST = gross - discounts - returns - net_rev(ex-gst)  approximately
# Simpler: GST rate is 18% for most products
# gross_ex_gst ≈ gross / 1.18 * 0.18  (but mix of 0% and 18%)
# Use: gst = gross * 0.18/1.18  as approximation
gst_approx = round(gross * 0.18 / 1.18, 2)

fc['wf'] = {
    'mrp_exgst':  round(gross - gst_approx, 2),
    'gst':        gst_approx,
    'discounts_u':round(tot_disc, 2),
    'retcancel_u':round(tot_ret, 2),
    'cogs_u':     round(tot_cogs, 2),
    'gst':        gst_approx,
}

# Ensure T.fcm exists (FirstCry margin total) — needed by firstcryWfStages T.fcm
if 'fcm' not in T or not T.get('fcm'):
    T['fcm'] = round(tot_fcm, 2)
    fc['tot'] = T
    print(f'Added T.fcm = {T["fcm"]:,.0f}')
else:
    print(f'T.fcm already present: {T["fcm"]:,.0f}')

# D_FC.fun must exist (even if empty) — funnelMode='none' so only existence matters
fc['fun'] = {
    'note': 'No traffic/funnel data available for FirstCry channel.',
}

print(f'\nD_FC.wf built: {list(fc["wf"].keys())}')
print(f'D_FC.fun: {fc["fun"]}')

# ── Save ──────────────────────────────────────────────────────────────────
print('\nSaving...')
html = save_var(html, 'D_FC', fc)

with open('index.html', 'w', encoding='utf-8') as f:
    f.write(html)
print('Written. Size:', len(html))

# ── Verify ────────────────────────────────────────────────────────────────
print('\n=== VERIFICATION ===')
with open('index.html', encoding='utf-8') as f:
    html2 = f.read()

m = re.search(r'const D_FC = (\{.*?\});', html2, re.DOTALL)
fc2 = json.loads(m.group(1))

wf_needed  = ['gst', 'retcancel_u', 'discounts_u', 'cogs_u']
fun_needed = ['note']
tot_needed = ['fcm', 'rev', 'netrev', 'gm', 'cm2', 'oh', 'ebitda']

ok = True
for f in wf_needed:
    if f not in fc2.get('wf', {}):
        print(f'  D_FC.wf.{f}: MISSING'); ok = False
    else:
        print(f'  D_FC.wf.{f}: {fc2["wf"][f]:,.0f}')

for f in fun_needed:
    if f not in fc2.get('fun', {}):
        print(f'  D_FC.fun.{f}: MISSING'); ok = False
    else:
        print(f'  D_FC.fun.{f}: OK')

for f in tot_needed:
    if f not in fc2.get('tot', {}):
        print(f'  D_FC.tot.{f}: MISSING'); ok = False

if ok:
    print('\n  D_FC: ALL CHECKS PASSED')

# Full parse check
for var in ['D', 'D_AMZ', 'D_FC', 'D_BL']:
    m2 = re.search('const ' + var + r' = (\{.*?\});', html2, re.DOTALL)
    try:
        json.loads(m2.group(1))
        print(f'  {var}: parse OK')
    except Exception as e:
        print(f'  {var}: PARSE ERROR {e}')
