import openpyxl, json

wb = openpyxl.load_workbook('Website_Unit_Cost_Economics_Summary (1) (1).xlsx', data_only=True)
ws = wb['April-June Unit Cost Ecomomic']

rows = []
for r in range(2, ws.max_row+1):
    row = [ws.cell(r,c).value for c in range(1, ws.max_column+1)]
    if any(v is not None for v in row) and row[0] not in (None, 'GRAND TOTAL', '—'):
        rows.append(row)

def flt(v, default=0.0):
    try: return float(v) if v is not None else default
    except: return default

products = []
SKIP_REV = -50

for row in rows:
    key = str(row[0]).strip()
    disp = str(row[1]).strip() if row[1] else key
    cat = str(row[2]).strip() if row[2] else ''
    rev = flt(row[3])
    mrp = flt(row[4])
    mrp_exgst = flt(row[5])
    returns_u = flt(row[6])
    disc_u = flt(row[7])
    netrev_u = flt(row[8])
    cogs_u = flt(row[9])
    gm_u = flt(row[10])
    pack_u = flt(row[11])
    rzp_u = flt(row[12])
    ship_u = flt(row[13])
    cod_u = flt(row[14])
    cm1_u = flt(row[15])
    mktg_u = flt(row[16])
    units = int(flt(row[17]))
    cm2_u = flt(row[18])
    cm2pct_raw = flt(row[19])
    oh_u = flt(row[20])
    ebitda_u = flt(row[21])
    ebitdapct_raw = flt(row[22])
    gmpct = flt(row[23])

    if rev < SKIP_REV or units <= 0:
        print(f'SKIPPING: {key!r} rev={rev} units={units}')
        continue

    # cm2pct and ebitdapct in the Excel are stored as decimals (-0.4858)
    # The dashboard stores them as percentages (-48.58)
    cm2pct = cm2pct_raw * 100
    ebitdapct = ebitdapct_raw * 100

    k = key + '||' + disp
    cm2_t = cm2_u * units
    ebitda_t = ebitda_u * units
    gm_t = gm_u * units
    netrev_t = netrev_u * units

    if cm2_u > 0 and ebitda_u > 0:
        quad = 'star'
    elif cm2_u > 0 and ebitda_u <= 0:
        quad = 'overhead'
    else:
        quad = 'loss'

    products.append({
        'k': k, 'key': key, 'disp': disp, 'cat': cat,
        'rev': rev, 'units': units, 'mrp': mrp,
        'netrev_u': round(netrev_u, 8),
        'gm_u': round(gm_u, 8),
        'ship_u': round(ship_u, 8),
        'mktg_u': round(mktg_u, 8),
        'cm1_u': round(cm1_u, 8),
        'cm2_u': round(cm2_u, 8),
        'oh_u': round(oh_u, 2),
        'ebitda_u': round(ebitda_u, 8),
        'cm2_t': round(cm2_t, 8),
        'ebitda_t': round(ebitda_t, 8),
        'cm2pct': round(cm2pct, 5),
        'ebitdapct': round(ebitdapct, 5),
        'quad': quad,
        '_gmpct': gmpct,
        '_gm_t': round(gm_t, 5),
        '_netrev_t': round(netrev_t, 5),
        '_cogs_u': cogs_u,
        '_pack_u': pack_u,
        '_rzp_u': rzp_u,
        '_cod_u': cod_u,
        '_returns_u': returns_u,
        '_disc_u': disc_u,
    })

print(f'Total valid products: {len(products)}')
print()

# Compute totals
tot_units = sum(p['units'] for p in products)
tot_rev = sum(p['rev'] for p in products)
tot_gm = sum(p['_gm_t'] for p in products)
tot_netrev = sum(p['_netrev_t'] for p in products)
tot_cm2 = sum(p['cm2_t'] for p in products)
tot_ebitda = sum(p['ebitda_t'] for p in products)
tot_mktg = sum(p['mktg_u'] * p['units'] for p in products)
tot_oh = sum(p['oh_u'] * p['units'] for p in products)

cm2pct_nr = tot_cm2 / tot_netrev * 100 if tot_netrev else 0
ebitdapct_nr = tot_ebitda / tot_netrev * 100 if tot_netrev else 0

print(f'Totals:')
print(f'  units={tot_units}, rev={tot_rev:.2f}, netrev={tot_netrev:.2f}')
print(f'  gm={tot_gm:.2f}, cm2={tot_cm2:.2f}, ebitda={tot_ebitda:.2f}')
print(f'  mktg={tot_mktg:.2f}, oh={tot_oh:.2f}')
print(f'  cm2pct={cm2pct_nr:.5f}, ebitdapct={ebitdapct_nr:.5f}')
print()

# Quad breakdown
for q in ['star','overhead','loss']:
    qp = [p for p in products if p['quad']==q]
    print(f"  {q}: n={len(qp)}, rev={sum(p['rev'] for p in qp):.2f}, units={sum(p['units'] for p in qp)}, ebitda={sum(p['ebitda_t'] for p in qp):.2f}, cm2={sum(p['cm2_t'] for p in qp):.2f}")

print()
# Category rollup
from collections import defaultdict
cats_data = defaultdict(lambda: {'n':0,'rev':0,'units':0,'cm2':0,'ebitda':0,'netrev':0,'mktg':0,'gm':0})
for p in products:
    c = p['cat']
    cats_data[c]['n'] += 1
    cats_data[c]['rev'] += p['rev']
    cats_data[c]['units'] += p['units']
    cats_data[c]['cm2'] += p['cm2_t']
    cats_data[c]['ebitda'] += p['ebitda_t']
    cats_data[c]['netrev'] += p['_netrev_t']
    cats_data[c]['mktg'] += p['mktg_u'] * p['units']
    cats_data[c]['gm'] += p['_gm_t']

for cat, d in sorted(cats_data.items(), key=lambda x: x[1]['rev'], reverse=True):
    cm2p = d['cm2']/d['netrev']*100 if d['netrev'] else 0
    ebp = d['ebitda']/d['netrev']*100 if d['netrev'] else 0
    print(f"  {cat!r}: n={d['n']}, rev={d['rev']:.2f}, units={d['units']}, cm2={d['cm2']:.2f}({cm2p:.2f}%), ebitda={d['ebitda']:.2f}({ebp:.2f}%), mktg={d['mktg']:.2f}")

print()
# Waterfall: sum of each cost type across all products
tot_pack = sum(p['_pack_u'] * p['units'] for p in products)
tot_rzp = sum(p['_rzp_u'] * p['units'] for p in products)
tot_ship = sum(p['ship_u'] * p['units'] for p in products)
tot_cod = sum(p['_cod_u'] * p['units'] for p in products)
tot_cogs = sum(p['_cogs_u'] * p['units'] for p in products)
tot_ret = sum(p['_returns_u'] * p['units'] for p in products)
tot_disc = sum(p['_disc_u'] * p['units'] for p in products)
gross_rev = sum(p['mrp'] * p['units'] for p in products)
tot_cm1 = sum(p['cm1_u'] * p['units'] for p in products)

print('Waterfall components:')
print(f'  Gross (MRP x units): {gross_rev:.2f}')
print(f'  Returns: {tot_ret:.2f}')
print(f'  Discounts: {tot_disc:.2f}')
print(f'  NetRev (from sheet sum): {tot_netrev:.2f}')
print(f'  COGS: {tot_cogs:.2f}')
print(f'  GM: {tot_gm:.2f}')
print(f'  Pack: {tot_pack:.2f}')
print(f'  Rzp: {tot_rzp:.2f}')
print(f'  Ship: {tot_ship:.2f}')
print(f'  COD: {tot_cod:.2f}')
print(f'  CM1: {tot_cm1:.2f}')
print(f'  Mktg: {tot_mktg:.2f}')
print(f'  CM2: {tot_cm2:.2f}')
print(f'  OH: {tot_oh:.2f}')
print(f'  EBITDA: {tot_ebitda:.2f}')
print(f'  mktg_pct_nr: {tot_mktg/tot_netrev*100:.5f}')
print(f'  mktg_pct_gross: {tot_mktg/tot_rev*100:.5f}')

# Per-unit averages for bridge
print()
print('Per-unit averages:')
print(f'  netrev_u: {tot_netrev/tot_units:.5f}')
print(f'  gm_u: {tot_gm/tot_units:.5f}')
print(f'  ship_u: {tot_ship/tot_units:.5f}')
print(f'  mktg_u: {tot_mktg/tot_units:.5f}')
print(f'  oh_u: {tot_oh/tot_units:.5f}')
print(f'  cm2_u: {tot_cm2/tot_units:.5f}')
print(f'  ebitda_u: {tot_ebitda/tot_units:.5f}')
gm_u_avg = tot_gm/tot_units
rzp_u_avg = tot_rzp/tot_units
pack_u_avg = tot_pack/tot_units
cod_u_avg = tot_cod/tot_units
cogs_u_avg = tot_cogs/tot_units
disc_u_avg = tot_disc/tot_units
ret_u_avg = tot_ret/tot_units
print(f'  cogs_u: {cogs_u_avg:.5f}')
print(f'  pack_u: {pack_u_avg:.5f}')
print(f'  rzp_u: {rzp_u_avg:.5f}')
print(f'  cod_u: {cod_u_avg:.5f}')

# Grand (new summary format)
gm_pct_avg = tot_gm/tot_netrev if tot_netrev else 0
profit_u_avg = tot_ebitda/tot_units
print()
print('Grand:')
print(f'  units={tot_units}, rev={tot_rev:.2f}, gm_u={gm_u_avg:.0f}, gmpct={gm_pct_avg:.2f}, ship_u={tot_ship/tot_units:.0f}, profit_u={profit_u_avg:.0f}, profit={tot_ebitda:.0f}')

# CAC from Summary sheet
ws_sum = wb['Apr-Jun Summary']
cac = ws_sum.cell(2,4).value
print(f'CAC from Summary sheet: {cac}')
