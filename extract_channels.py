import openpyxl
import json

wb = openpyxl.load_workbook('Channel Economics Dashboard.xlsx', data_only=True)

def safe(v):
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return v

# ============================================================
# WEBSITE DATA
# ============================================================
ws_web = wb['Raw Data - Website']
website_headers = [ws_web.cell(row=4, column=c).value for c in range(1, 50)]
print("Website headers:", website_headers[:35])

website_rows = []
for r in range(5, 200):
    code = ws_web.cell(row=r, column=1).value
    if code is None:
        continue
    row = {}
    for c, h in enumerate(website_headers[:40], 1):
        if h:
            row[h] = safe(ws_web.cell(row=r, column=c).value)
    website_rows.append(row)

print(f"\nWebsite rows: {len(website_rows)}")
for r in website_rows:
    print(f"  {r.get('Product Name (Code)')} | Units={r.get('Units sold')} | RevGST={r.get('3 month revenue with GST')} | RevNoGST={r.get('3 month revenue without GST')} | CM2={r.get('CM2')} | CM2%={r.get('CM2%')} | EBITA={r.get('EBITA')} | EBITA%={r.get('EBITA%')} | Marketing={r.get('Marketing Cost')} | NetRev={r.get('Net Revenue without GST')}")

# ============================================================
# AMAZON DATA
# ============================================================
ws_amz = wb['Raw Data - Amazon']
amz_headers = [ws_amz.cell(row=4, column=c).value for c in range(1, 50)]
print("\n\nAmazon headers:", amz_headers[:35])

amz_rows = []
for r in range(5, 200):
    prod = ws_amz.cell(row=r, column=1).value
    if prod is None:
        continue
    row = {}
    for c, h in enumerate(amz_headers[:45], 1):
        if h:
            row[h] = safe(ws_amz.cell(row=r, column=c).value)
    amz_rows.append(row)

print(f"\nAmazon rows: {len(amz_rows)}")
for r in amz_rows:
    print(f"  {r.get('P. Breakdown')} | Units={r.get('Units Sold')} | CM2={r.get('CM2')} | CM2%={r.get('CM2%')} | EBITDA={r.get('EBITDA')} | EBITDA%={r.get('EBITDA%')} | MktCost={r.get('Marketing Cost')} | NetRev={r.get('Net Revenue without GST')}")

# ============================================================
# FIRSTCRY DATA
# ============================================================
ws_fc = wb['Raw Data - FirstCry']
fc_headers = [ws_fc.cell(row=4, column=c).value for c in range(1, 35)]
print("\n\nFirstCry headers:", fc_headers)

fc_rows = []
for r in range(5, 200):
    prod = ws_fc.cell(row=r, column=1).value
    if prod is None:
        continue
    row = {}
    for c, h in enumerate(fc_headers, 1):
        if h:
            row[h] = safe(ws_fc.cell(row=r, column=c).value)
    fc_rows.append(row)

print(f"\nFirstCry rows: {len(fc_rows)}")
for r in fc_rows:
    print(f"  {r.get('P. Breakdown')} | Units={r.get('Units (3 months)')} | CM2={r.get('CM2')} | CM2%={r.get('CM2%')} | EBITDA={r.get('EBITDA')} | EBITDA%={r.get('EBITDA%')} | TotRev={r.get('Total Net Revenue')} | DRR={r.get('Daily Run Rate(DRR)')}")

# ============================================================
# BLINKIT DATA
# ============================================================
ws_bl = wb['Raw Data - Blinkit']
bl_headers = [ws_bl.cell(row=4, column=c).value for c in range(1, 35)]
print("\n\nBlinkit headers:", bl_headers)

bl_rows = []
for r in range(5, 200):
    prod = ws_bl.cell(row=r, column=1).value
    if prod is None:
        continue
    row = {}
    for c, h in enumerate(bl_headers, 1):
        if h:
            row[h] = safe(ws_bl.cell(row=r, column=c).value)
    bl_rows.append(row)

print(f"\nBlinkit rows: {len(bl_rows)}")
for r in bl_rows:
    print(f"  {r.get('P. Breakdown')} | Units={r.get('Units (3 months)')} | CM2={r.get('CM2')} | CM2%={r.get('CM2%')} | EBITDA={r.get('EBITDA')} | EBITDA%={r.get('EBITDA%')} | NetRev={r.get('Net Revenue without GST')} | MktCost={r.get('Marketing Cost')}")

# Save to json for later
all_data = {
    'website': website_rows,
    'amazon': amz_rows,
    'firstcry': fc_rows,
    'blinkit': bl_rows
}

with open('all_channel_data.json', 'w', encoding='utf-8') as f:
    json.dump(all_data, f, ensure_ascii=False, default=str, indent=2)
print("\nSaved to all_channel_data.json")
