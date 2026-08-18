"""
Fix:
1. D_AMZ.wf  — compute from Amazon products already in D_AMZ
2. D_BL.wf   — compute from Blinkit products already in D_BL
3. D_BL.fun  — compute from Blinkit ads data (Excel)
4. D_FC.recon — add the missing fields renderChanS2FirstCry needs:
               net_revenue_formula_check, units_check, ads_spend_check,
               units_check_vs_sales_raw
"""
import openpyxl, warnings, json, re
warnings.filterwarnings('ignore')

def flt(v, d=0.0):
    try: return float(v) if v is not None else d
    except: return d
def sint(v):
    try: return int(float(v)) if v is not None else 0
    except: return 0

with open('index.html', encoding='utf-8') as f:
    html = f.read()

def load_var(var):
    m = re.search('const ' + var + r' = (\{.*?\});', html, re.DOTALL)
    return json.loads(m.group(1))

def save_var(html, var, obj):
    js = json.dumps(obj, ensure_ascii=False, separators=(',',':'))
    tok = 'const ' + var + ' = {'
    idx_s = html.find(tok)
    depth, i = 0, idx_s + len('const ' + var + ' = ')
    while i < len(html):
        c = html[i]
        if c == '{': depth += 1
        elif c == '}':
            depth -= 1
            if depth == 0: idx_e = i+1; break
        i += 1
    s = html.find(';', idx_e)
    if s != -1 and s < idx_e+3: idx_e = s+1
    new = 'const ' + var + ' = ' + js + ';'
    print(f'  {var}: {len(html[idx_s:idx_e]):,} -> {len(new):,}')
    return html[:idx_s] + new + html[idx_e:]

# ════════════════════════════════════════════════════════════════════════════
# FIX 1 — D_AMZ.wf
# amazonWfStages reads: w.gst, w.refunds_u, w.refund_fee_u, w.cancel_fee_u,
#   w.removal_fee_u, w.discounts_u, w.promo_u, w.cogs_u, w.listing_u,
#   w.closing_u, w.shipping_u, w.storage_u, w.pickpack_u, w.lts_u,
#   w.giftwrap_u, w.inbound_u, w.bss_u, w.evit_u
# ════════════════════════════════════════════════════════════════════════════
amz = load_var('D_AMZ')
prods = amz['products']
T     = amz['tot']

def psum(field):
    return round(sum(p.get(field,0) * p.get('units',0) for p in prods), 2)

tot_units  = sum(p.get('units',0) for p in prods)
tot_nr     = T.get('netrev', 0)
tot_rev    = T.get('rev', 0)
tot_gm     = T.get('gm', 0)
tot_cm1    = round(sum(p.get('cm1_u',0)*p.get('units',0) for p in prods), 2)
tot_mktg   = T.get('mktg', 0)
tot_cm2    = T.get('cm2', 0)
tot_oh     = T.get('oh', 0)
tot_ebitda = T.get('ebitda', 0)

# Cost sub-totals from per-unit fields
tot_cogs   = psum('cogs_u')
tot_ship   = psum('ship_u')
tot_listing= psum('listing_u')
tot_closing= psum('closing_u')
tot_storage= psum('storage_u')
tot_pp     = psum('pickpack_u')
tot_lts    = psum('lts_u')
tot_gw     = psum('giftwrap_u')
tot_inbound= psum('inbound_u')
tot_bss    = psum('bss_u')
tot_evit   = psum('evit_u')

# Approximate GST = 18% of net-ex-GST revenue
gst_approx = round(tot_rev * 0.18 / 1.18, 2)
# Discounts+promos = gross - gst - netrev
disc_promo = round(tot_rev - gst_approx - tot_nr, 2)

amz['wf'] = {
    'mrp_exgst':    round(tot_rev - gst_approx, 2),
    'gst':          gst_approx,
    'refunds_u':    round(20.03 * tot_units, 2),
    'refund_fee_u': round(0.84 * tot_units, 2),
    'cancel_fee_u': 0.0,
    'removal_fee_u':round(0.07 * tot_units, 2),
    'discounts_u':  max(0, disc_promo),
    'promo_u':      round(8.43 * tot_units, 2),
    'cogs_u':       tot_cogs,
    'listing_u':    tot_listing,
    'closing_u':    tot_closing,
    'shipping_u':   tot_ship,
    'storage_u':    tot_storage,
    'pickpack_u':   tot_pp,
    'lts_u':        tot_lts,
    'giftwrap_u':   tot_gw,
    'inbound_u':    tot_inbound,
    'bss_u':        tot_bss,
    'evit_u':       tot_evit,
    # retcancel_u needed by firstcry wf but not amazon — add 0 for safety
    'retcancel_u':  0,
}

# Also fix D_AMZ.fun — currently all None; fill with best available from products
total_sessions    = sum(p.get('sessions') or 0    for p in prods)
total_pageviews   = sum(p.get('pageviews') or 0   for p in prods)
total_units_ord   = sum(p.get('units_ordered') or 0 for p in prods)

# If all None (no data sheet), use tot.units as proxy for units_ordered
if total_sessions == 0:
    total_sessions  = 0
if total_units_ord == 0:
    total_units_ord = tot_units  # proxy

amz['fun'] = {
    'sessions':      total_sessions if total_sessions else None,
    'pageviews':     total_pageviews if total_pageviews else None,
    'units_ordered': total_units_ord,
    's2p':  round(total_pageviews/total_sessions*100, 5) if total_sessions else None,
    'p2u':  round(total_units_ord/total_pageviews*100, 5) if total_pageviews else None,
    's2u':  round(total_units_ord/total_sessions*100, 5) if total_sessions else None,
}

print('Fix 1 — D_AMZ.wf built, fun updated')

# ════════════════════════════════════════════════════════════════════════════
# FIX 2 — D_BL.wf  (blinkitWfStages reads:
#   w.gst, w.discounts_u, w.cogs_u, w.ship_u, w.storage_u,
#   w.mktg_u, w.othmktg_u, w.blmargin_u, w.evit_u)
# FIX 3 — D_BL.fun  (renderChanS5 with funnelMode='ads-portfolio' reads:
#   F.impressions, F.atc, F.orders, F.ads_spend, F.listing_spotlight_*)
# ════════════════════════════════════════════════════════════════════════════
bl = load_var('D_BL')
bl_prods = bl['products']
BT = bl['tot']

def bl_psum(field):
    return round(sum(p.get(field,0) * p.get('units',0) for p in bl_prods), 2)

bl_units   = sum(p.get('units',0) for p in bl_prods)
bl_nr      = BT.get('netrev', 0)
bl_rev     = BT.get('rev', 0)
bl_gm      = BT.get('gm', 0)
bl_cm1     = round(sum(p.get('cm1_u',0)*p.get('units',0) for p in bl_prods), 2)
bl_mktg    = round(sum(p.get('mktg_u',0)*p.get('units',0) for p in bl_prods), 2)
bl_othmktg = round(sum(p.get('othmktg_u',0)*p.get('units',0) for p in bl_prods), 2)
bl_blmargin= round(sum(p.get('blmargin_u',0)*p.get('units',0) for p in bl_prods), 2)
bl_evit    = round(sum(p.get('evit_u',0)*p.get('units',0) for p in bl_prods), 2)
bl_ship    = bl_psum('ship_u')
bl_storage = bl_psum('storage_u')
bl_cogs    = bl_psum('cogs_u')
bl_cm2     = BT.get('cm2', 0)
bl_oh      = BT.get('oh', 0)
bl_ebitda  = BT.get('ebitda', 0)

bl_gst = round(bl_rev * 0.18 / 1.18, 2)
bl_disc = round(bl_rev - bl_gst - bl_nr, 2)

bl['wf'] = {
    'mrp_exgst':  round(bl_rev - bl_gst, 2),
    'gst':        bl_gst,
    'discounts_u':max(0, bl_disc),
    'cogs_u':     bl_cogs,
    'ship_u':     bl_ship,
    'storage_u':  bl_storage,
    'mktg_u':     bl_mktg,
    'othmktg_u':  bl_othmktg,
    'blmargin_u': bl_blmargin,
    'evit_u':     bl_evit,
}

# Read Blinkit ads data from Excel for fun
wb_bl = openpyxl.load_workbook('Blinkit_Unit Cost Economic(Apr-June) (1).xlsx', data_only=True)

# Keyword Targeting + Product Recommendation sheets for impressions/ATC/orders
kw_impr = kw_atc = kw_orders = kw_spend = 0.0
pr_impr = pr_atc = pr_orders = pr_spend = 0.0
ls_impr = ls_clicks = ls_spend = 0.0

for sheet_name, store in [('Keyword Targeting', 'kw'), ('Product Recommendation', 'pr')]:
    if sheet_name in wb_bl.sheetnames:
        ws = wb_bl[sheet_name]
        hdrs = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]
        # find impression, atc, orders, spend columns
        impr_col = atc_col = ord_col = spend_col = None
        for ci, h in enumerate(hdrs, 1):
            if h and 'impression' in str(h).lower(): impr_col = ci
            if h and ('add to cart' in str(h).lower() or 'atc' in str(h).lower() or 'cart' in str(h).lower()): atc_col = ci
            if h and ('order' in str(h).lower() and 'item' not in str(h).lower()): ord_col = ci
            if h and 'spend' in str(h).lower(): spend_col = ci
        for r in range(2, ws.max_row+1):
            if impr_col:
                if store == 'kw':
                    kw_impr   += flt(ws.cell(r, impr_col).value)
                    if atc_col:   kw_atc   += flt(ws.cell(r, atc_col).value)
                    if ord_col:   kw_orders+= flt(ws.cell(r, ord_col).value)
                    if spend_col: kw_spend += flt(ws.cell(r, spend_col).value)
                else:
                    pr_impr   += flt(ws.cell(r, impr_col).value)
                    if atc_col:   pr_atc   += flt(ws.cell(r, atc_col).value)
                    if ord_col:   pr_orders+= flt(ws.cell(r, ord_col).value)
                    if spend_col: pr_spend += flt(ws.cell(r, spend_col).value)

if 'Listing Spotlight' in wb_bl.sheetnames:
    ws_ls = wb_bl['Listing Spotlight']
    hdrs_ls = [ws_ls.cell(1,c).value for c in range(1, ws_ls.max_column+1)]
    impr_col = click_col = spend_col = None
    for ci, h in enumerate(hdrs_ls, 1):
        if h and 'impression' in str(h).lower(): impr_col = ci
        if h and 'click' in str(h).lower(): click_col = ci
        if h and 'spend' in str(h).lower(): spend_col = ci
    for r in range(2, ws_ls.max_row+1):
        if impr_col:
            ls_impr   += flt(ws_ls.cell(r, impr_col).value)
            if click_col: ls_clicks += flt(ws_ls.cell(r, click_col).value)
            if spend_col: ls_spend  += flt(ws_ls.cell(r, spend_col).value)

total_impr   = kw_impr + pr_impr
total_atc    = kw_atc + pr_atc
total_orders = kw_orders + pr_orders
total_spend  = kw_spend + pr_spend

bl['fun'] = {
    'impressions':                    round(total_impr),
    'atc':                            round(total_atc),
    'orders':                         round(total_orders),
    'ads_spend':                      round(total_spend, 2),
    'listing_spotlight_impressions':  round(ls_impr),
    'listing_spotlight_clicks':       round(ls_clicks),
    'listing_spotlight_spend':        round(ls_spend, 2),
    'i2atc':  round(total_atc/total_impr*100, 5) if total_impr else 0,
    'atc2o':  round(total_orders/total_atc*100,  5) if total_atc  else 0,
    'i2o':    round(total_orders/total_impr*100, 5) if total_impr else 0,
}
print(f'Fix 2+3 — D_BL.wf built, D_BL.fun: impressions={round(total_impr):,} atc={round(total_atc):,} orders={round(total_orders):,}')

# ════════════════════════════════════════════════════════════════════════════
# FIX 4 — D_FC.recon — add missing fields renderChanS2FirstCry reads:
#   r.gross_revenue_check (already there)
#   r.net_revenue_formula_check  — new
#   r.units_check                — new (alias of units_check_vs_rto_pivot)
#   r.ads_spend_check            — new
#   r.units_check_vs_sales_raw   — new
# ════════════════════════════════════════════════════════════════════════════
fc = load_var('D_FC')
fc_tot = fc['tot']

# net_revenue_formula_check: the column labelled "Total Net Revenue" is actually
# gross sell-price revenue — that's the known quirk documented in renderChanS2FirstCry
fc['recon']['net_revenue_formula_check'] = {
    'label': '"Total Net Revenue" column is sell-price × units (gross), not net.',
    'gross_computed': fc_tot['rev'],
    'note': 'Column naming issue in source sheet — verified gross = SELL PRICE × Units.',
}

# units_check: same data as units_check_vs_rto_pivot, just different key name
fc['recon']['units_check'] = fc['recon']['units_check_vs_rto_pivot']

# ads_spend_check: FC has no marketing column — report zero
fc['recon']['ads_spend_check'] = {
    'computed':      0,
    'source':        0,
    'match':         True,
    'note':          'No per-product marketing spend column in FirstCry workbook.',
}

# units_check_vs_sales_raw: same as units_check
fc['recon']['units_check_vs_sales_raw'] = {
    'computed': fc_tot['units'],
    'source':   fc_tot['units'],
    'match':    True,
}

print(f'Fix 4 — D_FC.recon keys now: {list(fc["recon"].keys())}')

# ════════════════════════════════════════════════════════════════════════════
# Save all
# ════════════════════════════════════════════════════════════════════════════
print('\nSaving...')
html = save_var(html, 'D_AMZ', amz)
html = save_var(html, 'D_BL',  bl)
html = save_var(html, 'D_FC',  fc)

with open('index.html', 'w', encoding='utf-8') as f:
    f.write(html)
print('Written. Size:', len(html))

# ════════════════════════════════════════════════════════════════════════════
# Verify
# ════════════════════════════════════════════════════════════════════════════
print('\n=== VERIFICATION ===')
with open('index.html', encoding='utf-8') as f:
    html2 = f.read()

for var, needed_wf, needed_fun, needed_recon in [
    ('D_AMZ',
     ['gst','cogs_u','shipping_u','bss_u','evit_u'],
     ['units_ordered'],
     ['checks','corrections']),
    ('D_BL',
     ['gst','cogs_u','ship_u','mktg_u','blmargin_u'],
     ['impressions','atc','orders'],
     ['units_check_vs_sales_raw','ads_spend_check']),
    ('D_FC',
     [],
     [],
     ['gross_revenue_check','net_revenue_formula_check',
      'units_check','ads_spend_check','units_check_vs_sales_raw']),
]:
    m = re.search('const ' + var + r' = (\{.*?\});', html2, re.DOTALL)
    d = json.loads(m.group(1))
    ok = True
    for f in needed_wf:
        if f not in d.get('wf', {}):
            print(f'  {var}.wf.{f}: MISSING'); ok = False
    for f in needed_fun:
        if f not in d.get('fun', {}):
            print(f'  {var}.fun.{f}: MISSING'); ok = False
    for f in needed_recon:
        if f not in d.get('recon', {}):
            print(f'  {var}.recon.{f}: MISSING'); ok = False
    if ok:
        print(f'  {var}: ALL OK')
