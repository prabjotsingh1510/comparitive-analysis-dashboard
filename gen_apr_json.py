import openpyxl, json, re

wb = openpyxl.load_workbook('Website_Unit_Cost_Economics_Summary (1) (1).xlsx', data_only=True)
ws = wb['April-June Unit Cost Ecomomic']

rows = []
for r in range(2, ws.max_row+1):
    row = [ws.cell(r,c).value for c in range(1, ws.max_column+1)]
    if any(v is not None for v in row) and row[0] not in (None, 'GRAND TOTAL', '—'):
        rows.append(row)

def flt(v, default=0.0):
    try: return float(v) if v is not None else default
    except: return default

products = []

for row in rows:
    key = str(row[0]).strip()
    disp = str(row[1]).strip() if row[1] else key
    cat = str(row[2]).strip() if row[2] else ''
    rev = flt(row[3])
    mrp = flt(row[4])
    mrp_exgst = flt(row[5])
    returns_u = flt(row[6])
    disc_u = flt(row[7])
    netrev_u = flt(row[8])
    cogs_u = flt(row[9])
    gm_u = flt(row[10])
    pack_u = flt(row[11])
    rzp_u = flt(row[12])
    ship_u = flt(row[13])
    cod_u = flt(row[14])
    cm1_u = flt(row[15])
    mktg_u = flt(row[16])
    units = int(flt(row[17]))
    cm2_u = flt(row[18])
    cm2pct_raw = flt(row[19])
    oh_u = flt(row[20])
    ebitda_u = flt(row[21])
    ebitdapct_raw = flt(row[22])
    gmpct = flt(row[23])

    # Skip error row
    if rev < -50 or units <= 0:
        continue

    # Percentages stored as decimals in Excel, convert to pct
    cm2pct = cm2pct_raw * 100
    ebitdapct = ebitdapct_raw * 100

    k = key + '||' + disp
    cm2_t = cm2_u * units
    ebitda_t = ebitda_u * units
    gm_t = gm_u * units
    netrev_t = netrev_u * units

    if cm2_u > 0 and ebitda_u > 0:
        quad = 'star'
    elif cm2_u > 0 and ebitda_u <= 0:
        quad = 'overhead'
    else:
        quad = 'loss'

    # new_profit_u from the Summary sheet (CAC=700)
    # = gm_u - ship_u - CAC (simplified)
    # We'll compute from the per-unit data: new_profit_u = cm1_u - 700
    cac = 700
    new_profit_u = round(cm1_u - cac)
    new_profit_t = new_profit_u * units

    products.append({
        'k': k, 'key': key, 'disp': disp, 'cat': cat,
        'rev': rev, 'units': units, 'mrp': mrp,
        'netrev_u': round(netrev_u, 10),
        'gm_u': round(gm_u, 10),
        'ship_u': round(ship_u, 10),
        'mktg_u': round(mktg_u, 10),
        'cm1_u': round(cm1_u, 10),
        'cm2_u': round(cm2_u, 10),
        'oh_u': round(oh_u, 2),
        'ebitda_u': round(ebitda_u, 10),
        'cm2_t': round(cm2_t, 10),
        'ebitda_t': round(ebitda_t, 10),
        'cm2pct': round(cm2pct, 10),
        'ebitdapct': round(ebitdapct, 10),
        'quad': quad,
        'new_profit_u': new_profit_u,
        'new_profit_t': new_profit_t,
        # private fields for totals
        '_gm_t': gm_t,
        '_netrev_t': netrev_t,
        '_cogs_u': cogs_u,
        '_pack_u': pack_u,
        '_rzp_u': rzp_u,
        '_cod_u': cod_u,
        '_returns_u': returns_u,
        '_disc_u': disc_u,
        '_gmpct': gmpct,
    })

# Aggregate totals
tot_units = sum(p['units'] for p in products)
tot_rev = sum(p['rev'] for p in products)
tot_gm = sum(p['_gm_t'] for p in products)
tot_netrev = sum(p['_netrev_t'] for p in products)
tot_cm2 = sum(p['cm2_t'] for p in products)
tot_ebitda = sum(p['ebitda_t'] for p in products)
tot_mktg = sum(p['mktg_u'] * p['units'] for p in products)
tot_oh = sum(p['oh_u'] * p['units'] for p in products)
tot_pack = sum(p['_pack_u'] * p['units'] for p in products)
tot_rzp = sum(p['_rzp_u'] * p['units'] for p in products)
tot_ship = sum(p['ship_u'] * p['units'] for p in products)
tot_cod = sum(p['_cod_u'] * p['units'] for p in products)
tot_cogs = sum(p['_cogs_u'] * p['units'] for p in products)
tot_cm1 = sum(p['cm1_u'] * p['units'] for p in products)

cm2pct_nr = tot_cm2 / tot_netrev * 100
ebitdapct_nr = tot_ebitda / tot_netrev * 100
cm2pct_rev = tot_cm2 / tot_rev * 100
ebitdapct_rev = tot_ebitda / tot_rev * 100

n = len(products)

# Quad counts
quad_data = {}
for q in ['star', 'overhead', 'loss']:
    qp = [p for p in products if p['quad'] == q]
    quad_data[q] = {
        'n': len(qp),
        'rev': sum(p['rev'] for p in qp),
        'ebitda': sum(p['ebitda_t'] for p in qp),
        'cm2': sum(p['cm2_t'] for p in qp),
        'units': sum(p['units'] for p in qp),
    }

# Category rollup
from collections import defaultdict
cats_raw = defaultdict(lambda: {'n':0,'rev':0,'units':0,'cm2':0,'ebitda':0,'netrev':0,'mktg':0,'gm':0})
for p in products:
    c = p['cat']
    cats_raw[c]['n'] += 1
    cats_raw[c]['rev'] += p['rev']
    cats_raw[c]['units'] += p['units']
    cats_raw[c]['cm2'] += p['cm2_t']
    cats_raw[c]['ebitda'] += p['ebitda_t']
    cats_raw[c]['netrev'] += p['_netrev_t']
    cats_raw[c]['mktg'] += p['mktg_u'] * p['units']
    cats_raw[c]['gm'] += p['_gm_t']

cats = []
for cat, d in sorted(cats_raw.items(), key=lambda x: x[1]['rev'], reverse=True):
    cm2p = d['cm2']/d['netrev']*100 if d['netrev'] else 0
    ebp = d['ebitda']/d['netrev']*100 if d['netrev'] else 0
    cats.append({
        'cat': cat,
        'rev': round(d['rev'], 10),
        'units': d['units'],
        'n': d['n'],
        'netrev': round(d['netrev'], 10),
        'mktg': round(d['mktg'], 10),
        'cm2': round(d['cm2'], 10),
        'ebitda': round(d['ebitda'], 10),
        'cm2pct': round(cm2p, 10),
        'ebitdapct': round(ebp, 10),
    })

# Grand/summary
gm_u_avg = tot_gm / tot_units
cac = 700

grand = {
    'units': tot_units,
    'rev': round(tot_rev, 2),
    'gm_u': round(gm_u_avg, 0),
    'gmpct': round(tot_gm / tot_netrev, 2),
    'ship_u': round(tot_ship / tot_units, 0),
    'profit_u': round(tot_ebitda / tot_units, 0),
    'profit': round(tot_ebitda, 0),
}

# Waterfall for s5
wf = {
    'gross': round(tot_rev, 2),
    'gst': round(sum(p['mrp'] * p['units'] - p['mrp'] / (1 + 0.18) * p['units'] for p in products), 2),  # approx
    'ret': round(sum(p['_returns_u'] * p['units'] for p in products), 2),
    'disc': round(tot_pack + tot_rzp + tot_cod, 2),  # placeholder
    'nr': round(tot_netrev, 2),
    'com': round(tot_cogs, 2),
    'gm': round(tot_gm, 2),
    'pack': round(tot_pack, 2),
    'rzp': round(tot_rzp, 2),
    'ship': round(tot_ship, 2),
    'cod': round(tot_cod, 2),
    'opsbundle': round(tot_pack + tot_rzp + tot_ship + tot_cod, 2),
    'cm1': round(tot_cm1, 2),
    'mkt': round(tot_mktg, 2),
    'cm2': round(tot_cm2, 2),
    'oh': round(tot_oh, 2),
    'eb': round(tot_ebitda, 2),
    'mkt_pct_nr': round(tot_mktg / tot_netrev * 100, 5),
    'mkt_pct_gross': round(tot_mktg / tot_rev * 100, 5),
}

# Now build the swings list (new_profit vs cm2, top 60 by abs delta)
swings = []
for p in products:
    delta_u = p['new_profit_u'] - p['cm2_u']
    delta_t = delta_u * p['units']
    swings.append({
        'key': p['key'],
        'disp': p['disp'],
        'units': p['units'],
        'old_cm2_u': p['cm2_u'],
        'new_profit_u': p['new_profit_u'],
        'delta_u': round(delta_u, 5),
        'delta_t': round(delta_t, 5),
        'old_mktg_u': p['mktg_u'],
        'cac': cac,
    })
swings.sort(key=lambda x: abs(x['delta_t']), reverse=True)
swings = swings[:60]

# Clean products for output (remove private fields)
clean_products = []
for p in products:
    cp = {k: v for k, v in p.items() if not k.startswith('_')}
    clean_products.append(cp)

# Build the apr object
apr_obj = {
    'label': 'Apr\u2013Jun 2026',
    'tot': {
        'rev': round(tot_rev, 2),
        'netrev': round(tot_netrev, 10),
        'units': tot_units,
        'cm2': round(tot_cm2, 10),
        'ebitda': round(tot_ebitda, 10),
        'cm2pct': round(cm2pct_nr, 10),
        'ebitdapct': round(ebitdapct_nr, 10),
        'cm2pct_rev': round(cm2pct_rev, 10),
        'ebitdapct_rev': round(ebitdapct_rev, 10),
        'mktg': round(tot_mktg, 10),
        'gm': round(tot_gm, 10),
        'oh': round(tot_oh, 10),
        'n': n,
    },
    'quad': quad_data,
    'cats': cats,
    'products': clean_products,
    'cac': cac,
    'grand': grand,
    'new_rev': round(tot_rev, 2),
    'new_units': tot_units,
    'swings': swings,
}

# Output as JSON
json_str = json.dumps(apr_obj, ensure_ascii=False)
print(f'JSON length: {len(json_str)}')
with open('apr_data.json', 'w', encoding='utf-8') as f:
    f.write(json_str)
print('Written to apr_data.json')

# Also print key fields for verification
print(f'\nVerification:')
print(f'  products: {len(clean_products)}')
print(f'  units: {tot_units}')
print(f'  rev: {tot_rev:.2f}')
print(f'  cm2: {tot_cm2:.2f} ({cm2pct_nr:.2f}%)')
print(f'  ebitda: {tot_ebitda:.2f} ({ebitdapct_nr:.2f}%)')
print(f'  mktg: {tot_mktg:.2f}')
print(f'  oh: {tot_oh:.2f}')
print(f'  cac: {cac}')
print(f'  stars: {quad_data["star"]["n"]}, overhead: {quad_data["overhead"]["n"]}, loss: {quad_data["loss"]["n"]}')
