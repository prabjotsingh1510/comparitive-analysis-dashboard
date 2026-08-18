"""
Final dashboard rebuild - extracts all data from Excel and generates updated index.html
"""
import openpyxl
import json
import re

wb = openpyxl.load_workbook('Channel Economics Dashboard.xlsx', data_only=True)

def safe(v):
    if v is None: return None
    try:
        return float(v)
    except: return str(v) if v else None

def fmt_inr(v, lakh_threshold=100000):
    if v is None: return '—'
    sign = '+' if v > 0 else ''
    if abs(v) >= lakh_threshold:
        return f"{sign}₹{v/100000:.2f} L"
    return f"{sign}₹{v:,.0f}"

def fmt_pct(v, decimals=1):
    if v is None: return '—'
    return f"{v*100:.{decimals}f}%"

# ============================================================
# WEBSITE
# ============================================================
ws_web = wb['Raw Data - Website']
web_headers = [ws_web.cell(row=4, column=c).value for c in range(1, 50)]

website_products = []
for r in range(5, 200):
    code = ws_web.cell(row=r, column=1).value
    if code is None: continue
    row = {}
    for c, h in enumerate(web_headers[:42], 1):
        if h:
            row[h] = safe(ws_web.cell(row=r, column=c).value)
    website_products.append(row)

# Aggregate
def web_total(field, multiplier_field='Units sold'):
    return sum((p.get(field) or 0) * (p.get(multiplier_field) or 1) for p in website_products)

total_web_units = sum(p.get('Units sold') or 0 for p in website_products)
total_web_rev_gst = sum(p.get('3 month revenue with GST') or 0 for p in website_products)
total_web_rev_no_gst = sum(p.get('3 month revenue without GST') or 0 for p in website_products)
total_web_net_rev = web_total('Net Revenue without GST')
total_web_cm2 = web_total('CM2')
total_web_ebita = web_total('EBITA')
total_web_mktg = web_total('Marketing Cost')
total_web_gm = web_total('Gross margin')

web_cm2pct = total_web_cm2 / total_web_net_rev if total_web_net_rev else 0
web_ebitapct = total_web_ebita / total_web_net_rev if total_web_net_rev else 0
web_gmpct = total_web_gm / total_web_net_rev if total_web_net_rev else 0
web_mktg_pct = total_web_mktg / total_web_net_rev if total_web_net_rev else 0

# Website profitable SKUs
web_stars = [p for p in website_products if (p.get('CM2') or 0) > 0 and (p.get('EBITA') or 0) > 0]
web_overhead = [p for p in website_products if (p.get('CM2') or 0) > 0 and (p.get('EBITA') or 0) <= 0]
web_loss = [p for p in website_products if (p.get('CM2') or 0) <= 0]

web_star_count = len(web_stars)
web_star_rev = sum((p.get('Net Revenue without GST') or 0) * (p.get('Units sold') or 0) for p in web_stars)
web_star_ebitda_total = sum((p.get('EBITA') or 0) * (p.get('Units sold') or 0) for p in web_stars)
total_web_skus = len(website_products)

# ============================================================
# AMAZON
# ============================================================
ws_amz = wb['Raw Data - Amazon']
amz_headers = [ws_amz.cell(row=4, column=c).value for c in range(1, 50)]

amazon_products = []
for r in range(5, 200):
    prod = ws_amz.cell(row=r, column=1).value
    if prod is None: continue
    row = {}
    for c, h in enumerate(amz_headers[:45], 1):
        if h:
            row[h] = safe(ws_amz.cell(row=r, column=c).value)
    amazon_products.append(row)

total_amz_units = sum(p.get('Units Sold') or 0 for p in amazon_products)
total_amz_net_rev = sum((p.get('Net Revenue without GST') or 0) * (p.get('Units Sold') or 0) for p in amazon_products)
total_amz_cm2 = sum((p.get('CM2') or 0) * (p.get('Units Sold') or 0) for p in amazon_products)
total_amz_ebitda = sum((p.get('EBITDA') or 0) * (p.get('Units Sold') or 0) for p in amazon_products)
total_amz_mktg = sum((p.get('Cost of Advertising') or 0) * (p.get('Units Sold') or 0) for p in amazon_products)
total_amz_gm = sum((p.get('Gross Margin') or 0) * (p.get('Units Sold') or 0) for p in amazon_products)

amz_cm2pct = total_amz_cm2 / total_amz_net_rev if total_amz_net_rev else 0
amz_ebitdapct = total_amz_ebitda / total_amz_net_rev if total_amz_net_rev else 0
amz_gmpct = total_amz_gm / total_amz_net_rev if total_amz_net_rev else 0
amz_mktg_pct = total_amz_mktg / total_amz_net_rev if total_amz_net_rev else 0

amz_stars = [p for p in amazon_products if (p.get('CM2') or 0) > 0 and (p.get('EBITDA') or 0) > 0]
amz_overhead = [p for p in amazon_products if (p.get('CM2') or 0) > 0 and (p.get('EBITDA') or 0) <= 0]
amz_loss = [p for p in amazon_products if (p.get('CM2') or 0) <= 0]

amz_star_count = len(amz_stars)
amz_star_rev = sum((p.get('Net Revenue without GST') or 0) * (p.get('Units Sold') or 0) for p in amz_stars)
amz_star_ebitda_total = sum((p.get('EBITDA') or 0) * (p.get('Units Sold') or 0) for p in amz_stars)
total_amz_skus = len(amazon_products)

# ============================================================
# FIRSTCRY
# ============================================================
ws_fc = wb['Raw Data - FirstCry']
fc_headers = [ws_fc.cell(row=4, column=c).value for c in range(1, 35)]

firstcry_products = []
for r in range(5, 200):
    prod = ws_fc.cell(row=r, column=1).value
    if prod is None: continue
    row = {}
    for c, h in enumerate(fc_headers, 1):
        if h:
            row[h] = safe(ws_fc.cell(row=r, column=c).value)
    firstcry_products.append(row)

# Only active (Units > 0)
fc_active = [p for p in firstcry_products if (p.get('Units (3 months)') or 0) > 0]

total_fc_units = sum(p.get('Units (3 months)') or 0 for p in fc_active)
total_fc_net_rev = sum((p.get('Total Net Revenue') or 0) for p in fc_active)
total_fc_cm2 = sum((p.get('CM2') or 0) * (p.get('Units (3 months)') or 0) for p in fc_active)
total_fc_ebitda = sum((p.get('EBITDA') or 0) * (p.get('Units (3 months)') or 0) for p in fc_active)
total_fc_gm = sum((p.get('Gross Margin') or 0) * (p.get('Units (3 months)') or 0) for p in fc_active)

fc_cm2pct = total_fc_cm2 / total_fc_net_rev if total_fc_net_rev else 0
fc_ebitdapct = total_fc_ebitda / total_fc_net_rev if total_fc_net_rev else 0
fc_gmpct = total_fc_gm / total_fc_net_rev if total_fc_net_rev else 0

fc_stars = [p for p in fc_active if (p.get('CM2') or 0) > 0 and (p.get('EBITDA') or 0) > 0]
fc_overhead = [p for p in fc_active if (p.get('CM2') or 0) > 0 and (p.get('EBITDA') or 0) <= 0]

fc_star_count = len(fc_stars)
fc_star_rev = sum(p.get('Total Net Revenue') or 0 for p in fc_stars)
fc_star_ebitda_total = sum((p.get('EBITDA') or 0) * (p.get('Units (3 months)') or 0) for p in fc_stars)
total_fc_skus = len(fc_active)

fc_platform_fee = sum((p.get('Firstcry Margin') or 0) * (p.get('Units (3 months)') or 0) for p in fc_active)
fc_platform_pct = fc_platform_fee / total_fc_net_rev if total_fc_net_rev else 0

# ============================================================
# BLINKIT
# ============================================================
ws_bl = wb['Raw Data - Blinkit']
bl_headers = [ws_bl.cell(row=4, column=c).value for c in range(1, 35)]

blinkit_products = []
for r in range(5, 200):
    prod = ws_bl.cell(row=r, column=1).value
    if prod is None: continue
    row = {}
    for c, h in enumerate(bl_headers, 1):
        if h:
            row[h] = safe(ws_bl.cell(row=r, column=c).value)
    blinkit_products.append(row)

total_bl_units = sum(p.get('Units (3 months)') or 0 for p in blinkit_products)
total_bl_net_rev = sum((p.get('Net Revenue without GST') or 0) * (p.get('Units (3 months)') or 0) for p in blinkit_products)
total_bl_cm2 = sum((p.get('CM2') or 0) * (p.get('Units (3 months)') or 0) for p in blinkit_products)
total_bl_ebitda = sum((p.get('EBITDA') or 0) * (p.get('Units (3 months)') or 0) for p in blinkit_products)
total_bl_mktg = sum((p.get('Marketing Cost') or 0) * (p.get('Units (3 months)') or 0) for p in blinkit_products)
total_bl_gm = sum((p.get('Gross Margin') or 0) * (p.get('Units (3 months)') or 0) for p in blinkit_products)
total_bl_blmargin = sum((p.get('Blinkit Margin') or 0) * (p.get('Units (3 months)') or 0) for p in blinkit_products)

bl_cm2pct = total_bl_cm2 / total_bl_net_rev if total_bl_net_rev else 0
bl_ebitdapct = total_bl_ebitda / total_bl_net_rev if total_bl_net_rev else 0
bl_gmpct = total_bl_gm / total_bl_net_rev if total_bl_net_rev else 0
bl_mktg_pct = total_bl_mktg / total_bl_net_rev if total_bl_net_rev else 0
bl_platform_pct = total_bl_blmargin / total_bl_net_rev if total_bl_net_rev else 0

bl_stars = [p for p in blinkit_products if (p.get('CM2') or 0) > 0 and (p.get('EBITDA') or 0) > 0]
bl_overhead = [p for p in blinkit_products if (p.get('CM2') or 0) > 0 and (p.get('EBITDA') or 0) <= 0]

bl_star_count = len(bl_stars)
total_bl_skus = len(blinkit_products)

print("=== SUMMARY ===")
print(f"Website: {total_web_skus} SKUs, {int(total_web_units)} units, Net Rev {total_web_net_rev:,.0f}, CM2 {total_web_cm2:,.0f} ({web_cm2pct*100:.1f}%), EBITA {total_web_ebita:,.0f} ({web_ebitapct*100:.1f}%)")
print(f"Amazon: {total_amz_skus} SKUs, {int(total_amz_units)} units, Net Rev {total_amz_net_rev:,.0f}, CM2 {total_amz_cm2:,.0f} ({amz_cm2pct*100:.1f}%), EBITDA {total_amz_ebitda:,.0f} ({amz_ebitdapct*100:.1f}%)")
print(f"FirstCry: {total_fc_skus} active SKUs, {int(total_fc_units)} units, Net Rev {total_fc_net_rev:,.0f}, CM2 {total_fc_cm2:,.0f} ({fc_cm2pct*100:.1f}%), EBITDA {total_fc_ebitda:,.0f} ({fc_ebitdapct*100:.1f}%)")
print(f"Blinkit: {total_bl_skus} SKUs, {int(total_bl_units)} units, Net Rev {total_bl_net_rev:,.0f}, CM2 {total_bl_cm2:,.0f} ({bl_cm2pct*100:.1f}%), EBITDA {total_bl_ebitda:,.0f} ({bl_ebitdapct*100:.1f}%)")
print()
print(f"Website Stars: {web_star_count}/{total_web_skus}, rev {web_star_rev:,.0f}, ebitda {web_star_ebitda_total:,.0f}")
print(f"Amazon Stars: {amz_star_count}/{total_amz_skus}, rev {amz_star_rev:,.0f}, ebitda {amz_star_ebitda_total:,.0f}")
print(f"FirstCry Stars: {fc_star_count}/{total_fc_skus}, rev {fc_star_rev:,.0f}, ebitda {fc_star_ebitda_total:,.0f}")
print(f"Blinkit Stars: {bl_star_count}/{total_bl_skus}")

# Print detailed product lists for analysis
print("\n=== WEB STARS ===")
web_stars_sorted = sorted(web_stars, key=lambda p: -(p.get('EBITA',0)*(p.get('Units sold',0) or 1)))
for p in web_stars_sorted:
    code = p.get('Product Name (Code)','')
    units = p.get('Units sold') or 0
    cm2 = p.get('CM2') or 0
    ebita = p.get('EBITA') or 0
    print(f"  {code}: CM2/u={cm2:.0f} EBITA/u={ebita:.0f} units={units:.0f}")

print("\n=== AMZ STARS ===")
amz_stars_sorted = sorted(amz_stars, key=lambda p: -(p.get('EBITDA',0)*(p.get('Units Sold',0) or 1)))
for p in amz_stars_sorted:
    code = p.get('P. Breakdown','')
    units = p.get('Units Sold') or 0
    cm2 = p.get('CM2') or 0
    ebitda = p.get('EBITDA') or 0
    print(f"  {code}: CM2/u={cm2:.0f} EBITDA/u={ebitda:.0f} units={units:.0f}")

print("\n=== FC STARS ===")
fc_stars_sorted = sorted(fc_stars, key=lambda p: -(p.get('EBITDA',0)*(p.get('Units (3 months)',0) or 1)))
for p in fc_stars_sorted:
    code = p.get('P. Breakdown','')
    units = p.get('Units (3 months)') or 0
    cm2 = p.get('CM2') or 0
    ebitda = p.get('EBITDA') or 0
    print(f"  {code}: CM2/u={cm2:.0f} EBITDA/u={ebitda:.0f} units={units:.0f}")

print("\n=== WEB OVERHEAD-HEAVY (CM2+ EBITDA-) ===")
web_oh_sorted = sorted(web_overhead, key=lambda p: p.get('EBITA',0)*(p.get('Units sold',0) or 1))
for p in web_oh_sorted[:5]:
    code = p.get('Product Name (Code)','')
    cm2 = p.get('CM2') or 0
    ebita = p.get('EBITA') or 0
    gap = -ebita
    print(f"  {code}: CM2/u={cm2:.0f} EBITA/u={ebita:.0f} gap={gap:.0f}")

print("\n=== AMZ OVERHEAD-HEAVY ===")
amz_oh_sorted = sorted(amz_overhead, key=lambda p: p.get('EBITDA',0)*(p.get('Units Sold',0) or 1))
for p in amz_oh_sorted[:5]:
    code = p.get('P. Breakdown','')
    cm2 = p.get('CM2') or 0
    ebitda = p.get('EBITDA') or 0
    gap = -ebitda
    print(f"  {code}: CM2/u={cm2:.0f} EBITDA/u={ebitda:.0f} gap={gap:.0f}")

print("\n=== FC OVERHEAD-HEAVY ===")
fc_oh_sorted = sorted(fc_overhead, key=lambda p: p.get('EBITDA',0)*(p.get('Units (3 months)',0) or 1))
for p in fc_oh_sorted[:5]:
    code = p.get('P. Breakdown','')
    cm2 = p.get('CM2') or 0
    ebitda = p.get('EBITDA') or 0
    gap = -ebitda
    print(f"  {code}: CM2/u={cm2:.0f} EBITDA/u={ebitda:.0f} gap={gap:.0f}")

print("\n=== BLINKIT ALL ===")
for p in sorted(blinkit_products, key=lambda p: p.get('CM2',0)):
    code = p.get('P. Breakdown','')
    units = p.get('Units (3 months)') or 0
    cm2 = p.get('CM2') or 0
    ebitda = p.get('EBITDA') or 0
    mktg = p.get('Marketing Cost') or 0
    bl_margin = p.get('Blinkit Margin') or 0
    net_rev = p.get('Net Revenue without GST') or 0
    mktg_pct = mktg/net_rev if net_rev else 0
    bl_pct = bl_margin/net_rev if net_rev else 0
    print(f"  {code}: units={units:.0f} CM2/u={cm2:.0f} EBITDA/u={ebitda:.0f} mktg/u={mktg:.0f} ({mktg_pct*100:.0f}%) BL_margin/u={bl_margin:.0f} ({bl_pct*100:.0f}%)")

# ============================================================
# MARKETING SPEND TAB - Website with vs without GST
# ============================================================
print("\n=== MARKETING SPEND COMPARISON (Website) ===")
print(f"3-month Net Rev WITH GST (col E): {total_web_rev_gst:,.0f}")
print(f"3-month Net Rev WITHOUT GST (col AD): {total_web_rev_no_gst:,.0f}")
print(f"Marketing Spend: {total_web_mktg:,.0f}")
print(f"Mktg as % of Rev WITH GST: {total_web_mktg/total_web_rev_gst*100:.1f}%")
print(f"Mktg as % of Rev WITHOUT GST: {total_web_mktg/total_web_rev_no_gst*100:.1f}%")

# Per-product for the marketing tab
print("\nPer-product marketing comparison:")
for p in sorted(website_products, key=lambda x: -(x.get('3 month revenue with GST') or 0))[:15]:
    code = p.get('Product Name (Code)','')
    units = p.get('Units sold') or 0
    rev_gst = p.get('3 month revenue with GST') or 0
    rev_no_gst = p.get('3 month revenue without GST') or 0
    mktg_pu = p.get('Marketing Cost') or 0
    total_mktg = mktg_pu * units
    pct_gst = total_mktg/rev_gst*100 if rev_gst else 0
    pct_no_gst = total_mktg/rev_no_gst*100 if rev_no_gst else 0
    print(f"  {code}: Rev(GST)={rev_gst:,.0f} Rev(NoGST)={rev_no_gst:,.0f} Mktg={total_mktg:,.0f} %GST={pct_gst:.1f}% %NoGST={pct_no_gst:.1f}%")
