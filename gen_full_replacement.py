"""
Generates the full replacement for D.periods.apr in index.html,
plus recomputes bridge, flips, recon, and relevant facts using the new Excel data.
Outputs a JSON file with the complete replacement object.
"""
import openpyxl, json, re
from collections import defaultdict

# ── helpers ─────────────────────────────────────────────────────────────────
def flt(v, default=0.0):
    try: return float(v) if v is not None else default
    except: return default

# ── read sheets ──────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook('Website_Unit_Cost_Economics_Summary (1) (1).xlsx', data_only=True)

# === APR-JUN (new verified data) ===
ws_a = wb['April-June Unit Cost Ecomomic']
raw_a = []
for r in range(2, ws_a.max_row+1):
    row = [ws_a.cell(r,c).value for c in range(1, ws_a.max_column+1)]
    if any(v is not None for v in row) and row[0] not in (None, 'GRAND TOTAL', '—'):
        raw_a.append(row)

# === JAN-MAR (already correct, read for bridge/flip comparison) ===
ws_j = wb['Jan-Mar Unit cost Economic']
raw_j = []
for r in range(2, ws_j.max_row+1):
    row = [ws_j.cell(r,c).value for c in range(1, ws_j.max_column+1)]
    if any(v is not None for v in row) and row[0] not in (None, 'GRAND TOTAL', '—'):
        raw_j.append(row)

CAC_APR = 700
CAC_JAN = 650  # unchanged from original dashboard

def parse_product(row, cac, oh_default=229.33):
    key = str(row[0]).strip()
    disp = str(row[1]).strip() if row[1] else key
    cat = str(row[2]).strip() if row[2] else ''
    rev = flt(row[3])
    mrp = flt(row[4])
    netrev_u = flt(row[8])
    cogs_u = flt(row[9])
    gm_u = flt(row[10])
    pack_u = flt(row[11])
    rzp_u = flt(row[12])
    ship_u = flt(row[13])
    cod_u = flt(row[14])
    cm1_u = flt(row[15])
    mktg_u = flt(row[16])
    units = int(flt(row[17]))
    cm2_u = flt(row[18])
    cm2pct_raw = flt(row[19])
    oh_u = flt(row[20]) if flt(row[20]) != 0 else oh_default
    ebitda_u = flt(row[21])
    ebitdapct_raw = flt(row[22])
    gmpct = flt(row[23])

    # Excel stores pct as decimals
    cm2pct = cm2pct_raw * 100
    ebitdapct = ebitdapct_raw * 100

    k = key + '||' + disp
    cm2_t = cm2_u * units
    ebitda_t = ebitda_u * units
    gm_t = gm_u * units
    netrev_t = netrev_u * units
    mktg_t = mktg_u * units

    if cm2_u > 0 and ebitda_u > 0:
        quad = 'star'
    elif cm2_u > 0:
        quad = 'overhead'
    else:
        quad = 'loss'

    new_profit_u = round(cm1_u - cac)
    new_profit_t = new_profit_u * units

    return {
        'k': k, 'key': key, 'disp': disp, 'cat': cat,
        'rev': rev, 'units': units, 'mrp': mrp,
        'netrev_u': round(netrev_u, 10),
        'gm_u': round(gm_u, 10),
        'ship_u': round(ship_u, 10),
        'mktg_u': round(mktg_u, 10),
        'cm1_u': round(cm1_u, 10),
        'cm2_u': round(cm2_u, 10),
        'oh_u': round(oh_u, 2),
        'ebitda_u': round(ebitda_u, 10),
        'cm2_t': round(cm2_t, 10),
        'ebitda_t': round(ebitda_t, 10),
        'cm2pct': round(cm2pct, 10),
        'ebitdapct': round(ebitdapct, 10),
        'quad': quad,
        'new_profit_u': new_profit_u,
        'new_profit_t': new_profit_t,
        # private
        '_gm_t': gm_t, '_netrev_t': netrev_t, '_mktg_t': mktg_t,
        '_cogs': cogs_u*units, '_pack': pack_u*units, '_rzp': rzp_u*units,
        '_ship': ship_u*units, '_cod': cod_u*units,
        '_gmpct': gmpct,
    }

# Parse APR products (skip error rows)
apr_prods = []
for row in raw_a:
    rev = flt(row[3])
    units = int(flt(row[17]))
    if rev < -50 or units <= 0:
        continue
    p = parse_product(row, CAC_APR)
    apr_prods.append(p)

# Parse JAN products
jan_prods = []
for row in raw_j:
    rev = flt(row[3])
    units = int(flt(row[17]))
    if rev < -50 or units <= 0:
        continue
    p = parse_product(row, CAC_JAN, oh_default=270.0)
    jan_prods.append(p)

# ── APR totals ───────────────────────────────────────────────────────────────
def totals(prods):
    tot = {}
    tot['rev'] = sum(p['rev'] for p in prods)
    tot['netrev'] = sum(p['_netrev_t'] for p in prods)
    tot['units'] = sum(p['units'] for p in prods)
    tot['cm2'] = sum(p['cm2_t'] for p in prods)
    tot['ebitda'] = sum(p['ebitda_t'] for p in prods)
    tot['mktg'] = sum(p['_mktg_t'] for p in prods)
    tot['gm'] = sum(p['_gm_t'] for p in prods)
    tot['oh'] = sum(p['oh_u'] * p['units'] for p in prods)
    tot['cm2pct'] = tot['cm2'] / tot['netrev'] * 100 if tot['netrev'] else 0
    tot['ebitdapct'] = tot['ebitda'] / tot['netrev'] * 100 if tot['netrev'] else 0
    tot['cm2pct_rev'] = tot['cm2'] / tot['rev'] * 100 if tot['rev'] else 0
    tot['ebitdapct_rev'] = tot['ebitda'] / tot['rev'] * 100 if tot['rev'] else 0
    tot['n'] = len(prods)
    return tot

def quad_breakdown(prods):
    out = {}
    for q in ['star', 'overhead', 'loss']:
        qp = [p for p in prods if p['quad'] == q]
        out[q] = {
            'n': len(qp),
            'rev': sum(p['rev'] for p in qp),
            'ebitda': sum(p['ebitda_t'] for p in qp),
            'cm2': sum(p['cm2_t'] for p in qp),
            'units': sum(p['units'] for p in qp),
        }
    return out

def cats_rollup(prods):
    cd = defaultdict(lambda: {'n':0,'rev':0,'units':0,'cm2':0,'ebitda':0,'netrev':0,'mktg':0,'gm':0})
    for p in prods:
        c = p['cat']
        cd[c]['n'] += 1
        cd[c]['rev'] += p['rev']
        cd[c]['units'] += p['units']
        cd[c]['cm2'] += p['cm2_t']
        cd[c]['ebitda'] += p['ebitda_t']
        cd[c]['netrev'] += p['_netrev_t']
        cd[c]['mktg'] += p['_mktg_t']
        cd[c]['gm'] += p['_gm_t']
    cats = []
    for cat, d in sorted(cd.items(), key=lambda x: -x[1]['rev']):
        cm2p = d['cm2']/d['netrev']*100 if d['netrev'] else 0
        ebp  = d['ebitda']/d['netrev']*100 if d['netrev'] else 0
        cats.append({
            'cat': cat, 'rev': round(d['rev'],10), 'units': d['units'], 'n': d['n'],
            'netrev': round(d['netrev'],10), 'mktg': round(d['mktg'],10),
            'cm2': round(d['cm2'],10), 'ebitda': round(d['ebitda'],10),
            'cm2pct': round(cm2p,10), 'ebitdapct': round(ebp,10),
        })
    return cats

def wf_build(prods, label='apr'):
    tot = totals(prods)
    cogs = sum(p['_cogs'] for p in prods)
    pack = sum(p['_pack'] for p in prods)
    rzp  = sum(p['_rzp']  for p in prods)
    ship = sum(p['_ship'] for p in prods)
    cod  = sum(p['_cod']  for p in prods)
    gm   = tot['gm']
    cm1  = sum(p['cm1_u']*p['units'] for p in prods)
    return {
        'gross': round(tot['rev'], 2),
        'gst': round(sum((p['mrp'] - p['mrp']/(1.18))*p['units'] for p in prods), 2),
        'ret': round(sum(flt(0)*p['units'] for p in prods), 2),  # returns in per-unit from netrev formula
        'disc': round(tot['rev'] - tot['netrev'] - cogs + gm - pack - rzp - ship - cod, 2),  # residual
        'nr': round(tot['netrev'], 2),
        'com': round(cogs, 2),
        'gm': round(gm, 2),
        'pack': round(pack, 2),
        'rzp': round(rzp, 2),
        'ship': round(ship, 2),
        'cod': round(cod, 2),
        'opsbundle': round(pack+rzp+ship+cod, 2),
        'cm1': round(cm1, 2),
        'mkt': round(tot['mktg'], 2),
        'cm2': round(tot['cm2'], 2),
        'oh': round(tot['oh'], 2),
        'eb': round(tot['ebitda'], 2),
        'mkt_pct_nr': round(tot['mktg']/tot['netrev']*100 if tot['netrev'] else 0, 5),
        'mkt_pct_gross': round(tot['mktg']/tot['rev']*100 if tot['rev'] else 0, 5),
    }

def swings_build(prods, cac):
    sw = []
    for p in prods:
        du = p['new_profit_u'] - p['cm2_u']
        dt = du * p['units']
        sw.append({
            'key': p['key'], 'disp': p['disp'],
            'units': p['units'],
            'old_cm2_u': p['cm2_u'],
            'new_profit_u': p['new_profit_u'],
            'delta_u': round(du, 5),
            'delta_t': round(dt, 5),
            'old_mktg_u': p['mktg_u'],
            'cac': cac,
        })
    sw.sort(key=lambda x: -abs(x['delta_t']))
    return sw[:60]

def grand_build(prods):
    t = totals(prods)
    return {
        'units': t['units'],
        'rev': round(t['rev'], 2),
        'gm_u': round(t['gm']/t['units']) if t['units'] else 0,
        'gmpct': round(t['gm']/t['netrev'], 2) if t['netrev'] else 0,
        'ship_u': round(sum(p['_ship'] for p in prods)/t['units']) if t['units'] else 0,
        'profit_u': round(t['ebitda']/t['units']) if t['units'] else 0,
        'profit': round(t['ebitda']),
    }

def clean(p):
    return {k: v for k, v in p.items() if not k.startswith('_')}

# ── BUILD APR PERIOD ─────────────────────────────────────────────────────────
apr_tot = totals(apr_prods)
apr_period = {
    'label': 'Apr\u2013Jun 2026',
    'tot': {k: round(v, 10) if isinstance(v, float) else v for k, v in apr_tot.items()},
    'quad': quad_breakdown(apr_prods),
    'cats': cats_rollup(apr_prods),
    'products': [clean(p) for p in apr_prods],
    'cac': CAC_APR,
    'grand': grand_build(apr_prods),
    'new_rev': round(apr_tot['rev'], 2),
    'new_units': apr_tot['units'],
    'swings': swings_build(apr_prods, CAC_APR),
}

# ── BRIDGE (jan vs apr per-unit) ─────────────────────────────────────────────
jan_tot = totals(jan_prods)
jan_u = jan_tot['units']
apr_u = apr_tot['units']

def pu(key, prods, tot_units):
    return sum(p.get('_'+key, p.get(key, 0)) * p['units'] for p in prods) / tot_units if tot_units else 0

# Compute per-unit averages for each cost line
def avg(field, prods, units):
    return sum(p[field] * p['units'] for p in prods) / units if units else 0

jan_nr_u  = jan_tot['netrev'] / jan_u
apr_nr_u  = apr_tot['netrev'] / apr_u
jan_gm_u  = jan_tot['gm'] / jan_u
apr_gm_u  = apr_tot['gm'] / apr_u
jan_cogs_u= sum(p['_cogs'] for p in jan_prods)/jan_u
apr_cogs_u= sum(p['_cogs'] for p in apr_prods)/apr_u
jan_pack_u= sum(p['_pack'] for p in jan_prods)/jan_u
apr_pack_u= sum(p['_pack'] for p in apr_prods)/apr_u
jan_rzp_u = sum(p['_rzp']  for p in jan_prods)/jan_u
apr_rzp_u = sum(p['_rzp']  for p in apr_prods)/apr_u
jan_ship_u= sum(p['_ship'] for p in jan_prods)/jan_u
apr_ship_u= sum(p['_ship'] for p in apr_prods)/apr_u
jan_cod_u = sum(p['_cod']  for p in jan_prods)/jan_u
apr_cod_u = sum(p['_cod']  for p in apr_prods)/apr_u
jan_mkt_u = jan_tot['mktg'] / jan_u
apr_mkt_u = apr_tot['mktg'] / apr_u
jan_oh_u  = jan_tot['oh'] / jan_u
apr_oh_u  = apr_tot['oh'] / apr_u

def bridge_row(label, j, a):
    d = a - j
    # positive impact = improvement (for cost lines, negative delta = better = positive impact)
    return {'label': label, 'jan': round(j,2), 'apr': round(a,2), 'delta': round(d,2), 'impact': round(d,2)}

bridge = [
    bridge_row('Net Revenue/unit', jan_nr_u, apr_nr_u),
    # costs: impact is negative of delta (cost increase = negative impact)
    {'label':'Cost of Material/unit','jan':round(jan_cogs_u,2),'apr':round(apr_cogs_u,2),'delta':round(apr_cogs_u-jan_cogs_u,2),'impact':round(-(apr_cogs_u-jan_cogs_u),2)},
    {'label':'Packaging/unit','jan':round(jan_pack_u,2),'apr':round(apr_pack_u,2),'delta':round(apr_pack_u-jan_pack_u,2),'impact':round(-(apr_pack_u-jan_pack_u),2)},
    {'label':'Razorpay/unit','jan':round(jan_rzp_u,2),'apr':round(apr_rzp_u,2),'delta':round(apr_rzp_u-jan_rzp_u,2),'impact':round(-(apr_rzp_u-jan_rzp_u),2)},
    {'label':'Shipping/unit','jan':round(jan_ship_u,2),'apr':round(apr_ship_u,2),'delta':round(apr_ship_u-jan_ship_u,2),'impact':round(-(apr_ship_u-jan_ship_u),2)},
    {'label':'COD/unit','jan':round(jan_cod_u,2),'apr':round(apr_cod_u,2),'delta':round(apr_cod_u-jan_cod_u,2),'impact':round(-(apr_cod_u-jan_cod_u),2)},
    {'label':'Marketing/unit','jan':round(jan_mkt_u,2),'apr':round(apr_mkt_u,2),'delta':round(apr_mkt_u-jan_mkt_u,2),'impact':round(-(apr_mkt_u-jan_mkt_u),2)},
    {'label':'Overheads/unit','jan':round(jan_oh_u,2),'apr':round(apr_oh_u,2),'delta':round(apr_oh_u-jan_oh_u,2),'impact':round(-(apr_oh_u-jan_oh_u),2)},
]

# ── FLIPS (products that changed quadrant) ───────────────────────────────────
jan_map = {p['k']: p for p in jan_prods}
apr_map = {p['k']: p for p in apr_prods}
both_keys = [k for k in jan_map if k in apr_map]

flips = []
for k in both_keys:
    j = jan_map[k]
    a = apr_map[k]
    if j['quad'] != a['quad']:
        flips.append({
            'key': j['key'], 'disp': j['disp'], 'cat': j['cat'],
            'jq': j['quad'], 'aq': a['quad'],
            'jrev': j['rev'], 'arev': a['rev'],
            'jcm2p': round(j['cm2pct'],2), 'acm2p': round(a['cm2pct'],2),
            'jebp': round(j['ebitdapct'],2), 'aebp': round(a['ebitdapct'],2),
        })

# ── RECON ────────────────────────────────────────────────────────────────────
# Old recon.apr was using different source rev (2354727.31) — new is 2420404.36
# Jan recon stays same
jan_sum_rev = sum(p['rev'] for p in jan_prods)
jan_sum_units = sum(p['units'] for p in jan_prods)
jan_sum_profit = sum(p['ebitda_t'] for p in jan_prods)

apr_sum_rev = sum(p['rev'] for p in apr_prods)
apr_sum_units = sum(p['units'] for p in apr_prods)
apr_sum_profit = sum(p['ebitda_t'] for p in apr_prods)

recon = {
    'jan': {
        'old_rev': 2614540.52,
        'new_rev': round(jan_sum_rev, 2),
        'old_units': 3206,
        'new_units': jan_sum_units,
        'sum_rev': round(jan_sum_rev, 2),
        'sum_units': jan_sum_units,
        'sum_profit': round(jan_sum_profit, 2),
        'missing': [],
    },
    'apr': {
        'old_rev': 2420404.36,
        'new_rev': round(apr_sum_rev, 2),
        'old_units': apr_sum_units,
        'new_units': apr_sum_units,
        'sum_rev': round(apr_sum_rev, 2),
        'sum_units': apr_sum_units,
        'sum_profit': round(apr_sum_profit, 2),
        'missing': [],
    },
}

# ── S5 FUNNEL — APR (traffic data not refreshed — keep existing structure) ──
# We only update waterfall; the fun/prods for s5 apr stays from existing
# but we need to output the wf

apr_wf = {
    'gross': round(apr_tot['rev'], 2),
    'variance': 0,
    'gst': 0,  # not in this sheet; approximate from mrp difference
    'ret': 0,
    'disc': round(apr_tot['rev'] - apr_tot['netrev'], 2),  # combined returns+disc+gst
    'nr': round(apr_tot['netrev'], 2),
    'com': round(sum(p['_cogs'] for p in apr_prods), 2),
    'gm': round(apr_tot['gm'], 2),
    'pack': round(sum(p['_pack'] for p in apr_prods), 2),
    'rzp': round(sum(p['_rzp'] for p in apr_prods), 2),
    'ship': round(sum(p['_ship'] for p in apr_prods), 2),
    'cod': round(sum(p['_cod'] for p in apr_prods), 2),
    'opsbundle': round(sum(p['_pack']+p['_rzp']+p['_ship']+p['_cod'] for p in apr_prods), 2),
    'cm1': round(sum(p['cm1_u']*p['units'] for p in apr_prods), 2),
    'mkt': round(apr_tot['mktg'], 2),
    'cm2': round(apr_tot['cm2'], 2),
    'oh': round(apr_tot['oh'], 2),
    'eb': round(apr_tot['ebitda'], 2),
    'mkt_pct_nr': round(apr_tot['mktg']/apr_tot['netrev']*100, 5),
    'mkt_pct_gross': round(apr_tot['mktg']/apr_tot['rev']*100, 5),
}

# ── FACTS (persistent loss / star both) — recompute from combined ─────────────
# Persistent loss = products that are loss in BOTH jan and apr
persistent_loss = []
for k in both_keys:
    j = jan_map[k]
    a = apr_map[k]
    if j['quad'] == 'loss' and a['quad'] == 'loss':
        combined_rev = j['rev'] + a['rev']
        combined_ebitda = j['ebitda_t'] + a['ebitda_t']
        persistent_loss.append({
            'key': j['key'],
            'rev': round(combined_rev, 2),
            'ebitda': round(combined_ebitda, 2),
            'jcm2p': round(j['cm2pct'], 2),
            'acm2p': round(a['cm2pct'], 2),
        })
persistent_loss.sort(key=lambda x: x['ebitda'])

# Star in both
star_both = []
for k in both_keys:
    j = jan_map[k]
    a = apr_map[k]
    if j['quad'] == 'star' and a['quad'] == 'star':
        star_both.append({
            'key': j['key'],
            'rev': round(j['rev']+a['rev'], 2),
            'ebitda': round(j['ebitda_t']+a['ebitda_t'], 2),
        })
star_both.sort(key=lambda x: -x['rev'])

# Output summary
print("=== BRIDGE ===")
for b in bridge:
    print(f"  {b['label']}: jan={b['jan']}, apr={b['apr']}, impact={b['impact']}")

print(f"\n=== FLIPS: {len(flips)} ===")
for f in flips:
    print(f"  {f['key']}: {f['jq']} -> {f['aq']}")

print(f"\n=== STAR BOTH: {len(star_both)} ===")
for s in star_both:
    print(f"  {s['key']}: rev={s['rev']}, ebitda={s['ebitda']}")

print(f"\n=== PERSISTENT LOSS (top 5): ===")
for p in persistent_loss[:5]:
    print(f"  {p['key']}: rev={p['rev']}, ebitda={p['ebitda']}")

print(f"\n=== APR PERIOD TOTALS ===")
t = apr_period['tot']
print(f"  units={t['units']}, rev={t['rev']:.2f}, cm2={t['cm2']:.2f}({t['cm2pct']:.2f}%), ebitda={t['ebitda']:.2f}({t['ebitdapct']:.2f}%)")

# ── WRITE FULL REPLACEMENT JSON ───────────────────────────────────────────────
output = {
    'apr_period': apr_period,
    'bridge': bridge,
    'flips': flips,
    'recon': recon,
    'apr_wf': apr_wf,
    'star_both': star_both,
    'persistent_loss': persistent_loss[:10],
}

with open('full_replacement.json', 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False)

print("\nWritten to full_replacement.json")
