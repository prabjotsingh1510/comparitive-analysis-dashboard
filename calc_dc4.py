import openpyxl

wb = openpyxl.load_workbook('Channel Economics Dashboard.xlsx', data_only=True)

def num(v):
    if v is None: return 0.0
    try: return float(v)
    except: return 0.0

# WEBSITE
ws_web = wb['Raw Data - Website']
web_headers = [ws_web.cell(row=4, column=c).value for c in range(1, 50)]
print("Web headers:", [(i+1, h) for i, h in enumerate(web_headers[:42]) if h])

W = []
for r in range(5, 200):
    code = ws_web.cell(row=r, column=1).value
    if code is None: continue
    row = {}
    for c, h in enumerate(web_headers[:42], 1):
        if h:
            v = ws_web.cell(row=r, column=c).value
            try: row[h] = float(v) if v is not None else None
            except: row[h] = v
    W.append(row)

total_marketing = sum(num(p.get('Marketing Cost')) * num(p.get('Units sold')) for p in W)
total_net_rev = sum(num(p.get('Net Revenue without GST')) * num(p.get('Units sold')) for p in W)
total_rev_gst = sum(num(p.get('3 month revenue with GST')) for p in W)
total_rev_no_gst = sum(num(p.get('3 month revenue without GST')) for p in W)
total_gross_margin = sum(num(p.get('Gross margin')) * num(p.get('Units sold')) for p in W)
total_shipping = sum(num(p.get('Shipping to Customer')) * num(p.get('Units sold')) for p in W)
total_units = sum(num(p.get('Units sold')) for p in W)

print(f"\nWebsite Totals:")
print(f"  Total Units: {total_units:,.0f}")
print(f"  Rev WITH GST (3mo column): {total_rev_gst:,.0f}")
print(f"  Rev WITHOUT GST (3mo column): {total_rev_no_gst:,.0f}")
print(f"  Net Rev without GST (per-unit × units): {total_net_rev:,.0f}")
print(f"  Gross Margin (per-unit × units): {total_gross_margin:,.0f}")
print(f"  Total Marketing: {total_marketing:,.0f}")
print(f"  Total Shipping: {total_shipping:,.0f}")
print()
print(f"  GM% of net rev (per-unit): {total_gross_margin/total_net_rev*100:.1f}%")
print(f"  Marketing% of net rev (per-unit): {total_marketing/total_net_rev*100:.1f}%")
print(f"  Marketing% of rev WITH GST: {total_marketing/total_rev_gst*100:.1f}%")
print(f"  Marketing% of rev WITHOUT GST (3mo col): {total_marketing/total_rev_no_gst*100:.1f}%")
print(f"  Shipping% of net rev: {total_shipping/total_net_rev*100:.1f}%")

# Per-product breakdown for marketing
print(f"\nTop marketing spenders:")
prods_by_mktg = sorted(W, key=lambda p: -num(p.get('Marketing Cost', 0)) * num(p.get('Units sold', 0)))
for p in prods_by_mktg[:10]:
    code = p.get('Product Name (Code)', '')
    mktg_u = num(p.get('Marketing Cost'))
    units = num(p.get('Units sold'))
    netrev_u = num(p.get('Net Revenue without GST'))
    mktg_pct = mktg_u/netrev_u*100 if netrev_u else 0
    print(f"  {code}: mktg/unit={mktg_u:.0f}, units={units:.0f}, mktg_total={mktg_u*units:.0f}, mktg%={mktg_pct:.1f}%")

# AMAZON
ws_amz = wb['Raw Data - Amazon']
amz_headers = [ws_amz.cell(row=4, column=c).value for c in range(1, 50)]
A = []
for r in range(5, 200):
    prod = ws_amz.cell(row=r, column=1).value
    if prod is None: continue
    row = {}
    for c, h in enumerate(amz_headers[:45], 1):
        if h:
            v = ws_amz.cell(row=r, column=c).value
            try: row[h] = float(v) if v is not None else None
            except: row[h] = v
    A.append(row)

amz_mktg = sum(num(p.get('Cost of Advertising')) * num(p.get('Units Sold')) for p in A)
amz_net_rev = sum(num(p.get('Net Revenue without GST')) * num(p.get('Units Sold')) for p in A)
amz_gm = sum(num(p.get('Gross Margin')) * num(p.get('Units Sold')) for p in A)
amz_shipping_est = sum((num(p.get('Shipping charge', 0)) + num(p.get('Pick and pack Fee (FBA)', 0)) + num(p.get('Inbound Transportation Fee (FBA)', 0)) + num(p.get('Storage Fee (FBA)', 0))) * num(p.get('Units Sold')) for p in A)
print(f"\nAmazon Totals:")
print(f"  Net Rev (per-unit × units): {amz_net_rev:,.0f}")
print(f"  Gross Margin: {amz_gm:,.0f} ({amz_gm/amz_net_rev*100:.1f}%)")
print(f"  Marketing (Ads): {amz_mktg:,.0f} ({amz_mktg/amz_net_rev*100:.1f}% of net rev)")
print(f"  Shipping-related fees: {amz_shipping_est:,.0f} ({amz_shipping_est/amz_net_rev*100:.1f}% of net rev)")

# FIRSTCRY
ws_fc = wb['Raw Data - FirstCry']
fc_headers = [ws_fc.cell(row=4, column=c).value for c in range(1, 35)]
FC = []
for r in range(5, 200):
    prod = ws_fc.cell(row=r, column=1).value
    if prod is None: continue
    row = {}
    for c, h in enumerate(fc_headers, 1):
        if h:
            v = ws_fc.cell(row=r, column=c).value
            try: row[h] = float(v) if v is not None else None
            except: row[h] = v
    if num(row.get('Units (3 months)')) > 0:
        FC.append(row)

fc_net_rev = sum(num(p.get('Total Net Revenue', 0)) for p in FC)
fc_gm = sum(num(p.get('Gross Margin', 0)) * num(p.get('Units (3 months)')) for p in FC)
fc_fcm = sum(num(p.get('Firstcry Margin', 0)) * num(p.get('Units (3 months)')) for p in FC)
print(f"\nFirstCry Totals:")
print(f"  Net Rev: {fc_net_rev:,.0f}")
print(f"  Gross Margin: {fc_gm:,.0f} ({fc_gm/fc_net_rev*100:.1f}%)")
print(f"  Firstcry Margin (platform fee): {fc_fcm:,.0f} ({fc_fcm/fc_net_rev*100:.1f}%)")

# BLINKIT
ws_bl = wb['Raw Data - Blinkit']
bl_headers = [ws_bl.cell(row=4, column=c).value for c in range(1, 35)]
BL = []
for r in range(5, 200):
    prod = ws_bl.cell(row=r, column=1).value
    if prod is None: continue
    row = {}
    for c, h in enumerate(bl_headers, 1):
        if h:
            v = ws_bl.cell(row=r, column=c).value
            try: row[h] = float(v) if v is not None else None
            except: row[h] = v
    BL.append(row)

bl_net_rev = sum(num(p.get('Net Revenue without GST', 0)) * num(p.get('Units (3 months)', 0)) for p in BL)
bl_gm = sum(num(p.get('Gross Margin', 0)) * num(p.get('Units (3 months)', 0)) for p in BL)
bl_mktg = sum(num(p.get('Marketing Cost', 0)) * num(p.get('Units (3 months)', 0)) for p in BL)
bl_blmargin = sum(num(p.get('Blinkit Margin', 0)) * num(p.get('Units (3 months)', 0)) for p in BL)
bl_shipping = sum(num(p.get('Shipping Cost', 0)) * num(p.get('Units (3 months)', 0)) for p in BL)
print(f"\nBlinkit Totals:")
print(f"  Net Rev: {bl_net_rev:,.0f}")
print(f"  Gross Margin: {bl_gm:,.0f} ({bl_gm/bl_net_rev*100:.1f}%)")
print(f"  Marketing: {bl_mktg:,.0f} ({bl_mktg/bl_net_rev*100:.1f}%)")
print(f"  Blinkit Margin (platform): {bl_blmargin:,.0f} ({bl_blmargin/bl_net_rev*100:.1f}%)")
print(f"  Shipping: {bl_shipping:,.0f} ({bl_shipping/bl_net_rev*100:.1f}%)")
