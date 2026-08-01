import json
import openpyxl

def num(v, default=None):
    if v is None or v == '':
        return default
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in ('NA', '#N/A', '#VALUE!', '#DIV/0!', ''):
        return default
    try:
        return float(s)
    except ValueError:
        return default

def rows_of(wb, sheet):
    rows = list(wb[sheet].iter_rows(values_only=True))
    header = rows[0]
    return header, rows[1:]

with open('website_D_full.json', 'r', encoding='utf-8') as f:
    D = json.load(f)

SUMMARY_PATH = "Website_Unit_Cost_Economics_Summary (1) (1).xlsx"
wb = openpyxl.load_workbook(SUMMARY_PATH, data_only=True, read_only=True)

def load_crosscheck(sheet):
    """Unit cost Economic sheets: reconciliation cross-check, one row per product, keyed by short 'Product Name' (col0)."""
    h, rows = rows_of(wb, sheet)
    out = {}
    for r in rows:
        if not r or r[0] in (None, ''):
            continue
        key = str(r[0]).strip()
        rev = num(r[h.index('3 month revenue')], 0)
        units = num(r[h.index('Units sold')], 0)
        out[key] = {'rev': rev, 'units': units, 'disp': r[1]}
    return out

def load_grand(sheet):
    """Summary sheets: fixed-CAC methodology. Returns (cac, per-product dict keyed by display name, grand-total dict)."""
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    cac = num(rows[1][3])
    header_row_idx = None
    for i, r in enumerate(rows):
        if r[0] == 'Product':
            header_row_idx = i
            break
    header = rows[header_row_idx]
    products = {}
    grand = None
    for r in rows[header_row_idx+1:]:
        if not r or r[0] in (None, ''):
            continue
        if r[0] == 'GRAND TOTAL':
            grand = {
                'units': num(r[header.index('Units (3mo)')]),
                'rev': num(r[header.index('Revenue (₹)')]),
                'gm_u': num(r[header.index('GM/unit (₹)')]),
                'gmpct': num(r[header.index('GM %')]),
                'ship_u': num(r[header.index('Shipping/unit (₹)')]),
                'profit_u': num(r[header.index('Profit/unit after ads+ship (₹)')]),
                'profit': num(r[header.index('3-mo profit (₹)')]),
            }
            continue
        disp = str(r[0]).strip()
        products[disp] = {
            'units': num(r[header.index('Units (3mo)')]),
            'rev': num(r[header.index('Revenue (₹)')]),
            'profit_u': num(r[header.index('Profit/unit after ads+ship (₹)')]),
            'profit': num(r[header.index('3-mo profit (₹)')]),
        }
    return cac, products, grand

def process(pk, crosscheck_sheet, grand_sheet):
    old_products = D['periods'][pk]['products']
    old_rev = D['periods'][pk]['tot']['rev']
    old_units = D['periods'][pk]['tot']['units']

    cc = load_crosscheck(crosscheck_sheet)
    cac, grand_products, grand = load_grand(grand_sheet)

    new_rev = sum(v['rev'] for v in cc.values())
    new_units = sum(v['units'] for v in cc.values())

    missing = []
    for p in old_products:
        if p['key'] not in cc:
            missing.append({'key': p['key'], 'disp': p['disp'], 'rev': p['rev'], 'units': p['units']})

    # swings: join old-format product (by disp, matching how Summary sheet keys by display name) with grand_products
    swings = []
    matched_new_profit = {}
    unmatched_old = []
    for p in old_products:
        gp = grand_products.get(p['disp'])
        if gp is None:
            unmatched_old.append(p['disp'])
            continue
        new_profit_u = gp['profit_u']
        old_cm2_u = p['cm2_u']
        delta_u = new_profit_u - old_cm2_u
        delta_t = delta_u * p['units']
        swings.append({
            'key': p['key'], 'disp': p['disp'], 'units': p['units'],
            'old_cm2_u': old_cm2_u, 'new_profit_u': new_profit_u,
            'delta_u': round(delta_u, 2), 'delta_t': round(delta_t, 2),
            'old_mktg_u': p['mktg_u'], 'cac': cac,
        })
        matched_new_profit[p['key']] = (new_profit_u, new_profit_u * p['units'])
    swings.sort(key=lambda s: -abs(s['delta_t']))

    print(f"\n=== {pk.upper()} ===")
    print("old_rev:", old_rev, "old_units:", old_units)
    print("new_rev (crosscheck sum):", new_rev, "new_units:", new_units)
    print("cross-check rows:", len(cc))
    print("missing count (old products not in crosscheck):", len(missing))
    for m in missing:
        print("  MISSING:", m)
    print("cac:", cac)
    print("grand:", grand)
    print("swings matched:", len(swings), "of", len(old_products), "old products")
    print("unmatched old products (no row in Summary sheet, by disp):", len(unmatched_old))
    for u in unmatched_old[:20]:
        print("  UNMATCHED:", u)
    extra_in_grand = set(grand_products.keys()) - set(p['disp'] for p in old_products)
    print("extra disp names in Summary sheet not in old products:", len(extra_in_grand))
    for e in list(extra_in_grand)[:10]:
        print("  EXTRA:", e)

    return {
        'recon': {'old_rev': old_rev, 'new_rev': new_rev, 'old_units': old_units, 'new_units': new_units,
                  'sum_rev': grand['rev'], 'sum_units': grand['units'], 'sum_profit': grand['profit'], 'missing': missing},
        'grand': grand, 'cac': cac, 'new_rev': new_rev, 'new_units': new_units, 'swings': swings,
        'matched_new_profit': matched_new_profit,
    }

jan_result = process('jan', 'Jan-Mar Unit cost Economic', 'Jan-Mar Summary')
apr_result = process('apr', 'April-June Unit Cost Ecomomic', 'Apr-Jun Summary')

with open('website_new_summary_computed.json', 'w', encoding='utf-8') as f:
    json.dump({'jan': jan_result, 'apr': apr_result}, f)
print("\nwrote website_new_summary_computed.json")
