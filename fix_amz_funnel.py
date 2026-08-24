"""
Fix D_AMZ's broken organic traffic funnel (Section 5 crashes on Amazon tab:
"Cannot read properties of null (reading 'toFixed')").

Root cause: D_AMZ.products[].sessions/pageviews/units_ordered/s2p/p2u/s2u were
never populated (all null), so D_AMZ.fun (aggregate) was also all-null, and the
render code calls F.s2p.toFixed(1) etc. with no null guard.

Fix: merge real per-ASIN Sessions/Page Views/Units Ordered from the
"Raw data Apr-June" sheet in the Amazon workbook (the sheet the app's own
Section 5 note already cites as the source), joined on ASIN — matching what
Website's Section 5 already does with GA data.

Data quirk discovered: Amazon repeats the same Sessions/Page Views figure on
every SKU row for a given (parent) ASIN (traffic is tracked per-ASIN, not
per-SKU), while Units Ordered is genuinely per-SKU and additive. So per ASIN:
sessions/pageviews = max() across its rows, units_ordered = sum() across rows.
"""
import openpyxl, warnings, json, re
from collections import defaultdict
warnings.filterwarnings('ignore')

def flt(v):
    try: return float(v) if v is not None else 0.0
    except: return 0.0

# ---- 1. Read + aggregate the raw traffic sheet ----
wb = openpyxl.load_workbook('Unit cost economics AMAZON April-June (2).xlsx', data_only=True)
ws = wb['Raw data Apr-June']
hdrs = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]
col = {h: i+1 for i, h in enumerate(hdrs) if h}

by_asin = defaultdict(list)
for r in range(2, ws.max_row+1):
    asin = ws.cell(r, col['(Parent) ASIN']).value
    if not asin: continue
    by_asin[str(asin).strip()].append({
        'sessions': flt(ws.cell(r, col['Sessions - Total']).value),
        'pageviews': flt(ws.cell(r, col['Page Views - Total']).value),
        'units_ordered': flt(ws.cell(r, col['Units Ordered']).value),
    })

agg = {}
for asin, rows in by_asin.items():
    agg[asin] = {
        'sessions': max(r['sessions'] for r in rows),
        'pageviews': max(r['pageviews'] for r in rows),
        'units_ordered': sum(r['units_ordered'] for r in rows),
    }

# ---- 2. Load D_AMZ out of index.html ----
with open('index.html', encoding='utf-8') as f:
    html = f.read()

def extract_span(content, marker):
    i = content.index(marker)
    start = i + len(marker)
    depth = 0
    j = start
    in_str = False
    esc = False
    str_ch = ''
    while j < len(content):
        c = content[j]
        if in_str:
            if esc: esc = False
            elif c == '\\': esc = True
            elif c == str_ch: in_str = False
        else:
            if c == '"' or c == "'": in_str = True; str_ch = c
            elif c == '{': depth += 1
            elif c == '}':
                depth -= 1
                if depth == 0: j += 1; break
        j += 1
    return start, j

marker = 'const D_AMZ = '
s, e = extract_span(html, marker)
D_AMZ = json.loads(html[s:e])

# ---- 3. Merge onto each product by ASIN ----
matched, unmatched = 0, []
for p in D_AMZ['products']:
    asin = (p.get('asin') or '').strip()
    a = agg.get(asin)
    if not a:
        p['sessions'] = None
        p['pageviews'] = None
        p['units_ordered'] = None
        p['s2p'] = None
        p['p2u'] = None
        p['s2u'] = None
        unmatched.append(p.get('key'))
        continue
    matched += 1
    sess, pv, uo = a['sessions'], a['pageviews'], a['units_ordered']
    p['sessions'] = sess
    p['pageviews'] = pv
    p['units_ordered'] = uo
    p['s2p'] = round(pv/sess*100, 5) if sess else None
    p['p2u'] = round(uo/pv*100, 5) if pv else None
    p['s2u'] = round(uo/sess*100, 5) if sess else None

print(f'Matched {matched}/{len(D_AMZ["products"])} products by ASIN.')
if unmatched:
    print('Unmatched (left null):', unmatched)

# ---- 4. Recompute the fun aggregate from the merged products ----
tot_sessions = sum(p['sessions'] for p in D_AMZ['products'] if p.get('sessions'))
tot_pageviews = sum(p['pageviews'] for p in D_AMZ['products'] if p.get('pageviews'))
tot_units_ord = sum(p['units_ordered'] for p in D_AMZ['products'] if p.get('units_ordered'))

D_AMZ['fun'] = {
    'sessions': tot_sessions,
    'pageviews': tot_pageviews,
    'units_ordered': tot_units_ord,
    's2p': round(tot_pageviews/tot_sessions*100, 5) if tot_sessions else None,
    'p2u': round(tot_units_ord/tot_pageviews*100, 5) if tot_pageviews else None,
    's2u': round(tot_units_ord/tot_sessions*100, 5) if tot_sessions else None,
}
print('New D_AMZ.fun:', D_AMZ['fun'])

# ---- 5. Write back ----
new_js = json.dumps(D_AMZ, ensure_ascii=False, separators=(',', ':'))
new_block = marker + new_js
html = html[:s-len(marker)] + new_block + html[e:]
# strip the ; if the char right after was already consumed correctly (extract_span stops right after closing brace)

with open('index.html', 'w', encoding='utf-8') as f:
    f.write(html)
print('index.html written. New size:', len(html))
