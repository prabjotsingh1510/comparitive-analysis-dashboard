"""
Update D.periods.apr in the main website D object with new Excel data.
Also update the mktg spend comparison panel in the marketing tab.
"""
import openpyxl, json, re

wb = openpyxl.load_workbook('Channel Economics Dashboard.xlsx', data_only=True)

def num(v):
    if v is None: return 0.0
    try: return float(v)
    except: return 0.0

# ============================================================
# WEBSITE - Read from Excel
# ============================================================
ws_web = wb['Raw Data - Website']
web_headers = [ws_web.cell(row=4, column=c).value for c in range(1, 50)]

W = []
for r in range(5, 200):
    code = ws_web.cell(row=r, column=1).value
    if code is None: continue
    row = {}
    for c, h in enumerate(web_headers[:42], 1):
        if h:
            v = ws_web.cell(row=r, column=c).value
            try: row[h] = float(v) if v is not None else None
            except: row[h] = v
    W.append(row)

# Print headers for reference
print("Web headers:", [h for h in web_headers[:42] if h])

# Build products for apr period
web_prods = []
for p in W:
    code = str(p.get('Product Name (Code)', '') or '')
    disp = str(p.get('Product Name (Full)', code) or code)
    units = num(p.get('Units sold'))
    netrev_u = num(p.get('Net Revenue without GST'))
    gm_u = num(p.get('Gross margin'))
    cm1_u = num(p.get('CM 1'))
    cm2_u = num(p.get('CM2'))
    oh_u = num(p.get('Overhead Cost'))
    ebitda_u = num(p.get('EBITA'))
    mktg_u = num(p.get('Marketing Cost'))
    ship_u = num(p.get('Shipping to Customer'))
    mrp = num(p.get('MRP Including GST'))
    cat = str(p.get('Category', '') or '')
    rev_gst = num(p.get('3 month revenue with GST'))
    rev_no_gst = num(p.get('3 month revenue without GST'))

    netrev_t = netrev_u * units
    cm2_t = cm2_u * units
    ebitda_t = ebitda_u * units
    gm_t = gm_u * units
    cm1_t = cm1_u * units
    mktg_t = mktg_u * units

    if cm2_u > 0 and ebitda_u > 0:
        quad = 'star'; act = 'scale'
    elif cm2_u > 0:
        quad = 'overhead'; act = 'overhead'
    else:
        quad = 'loss'; act = 'cut'

    web_prods.append({
        'k': f"{code}||{disp}",
        'key': code,
        'disp': disp[:100],
        'cat': cat,
        'rev': round(rev_gst, 2),
        'units': units,
        'mrp': mrp if mrp else None,
        'netrev_u': round(netrev_u, 2),
        'netrev_t': round(netrev_t, 2),
        'gm_u': round(gm_u, 2),
        'gm_t': round(gm_t, 2),
        'gmpct': round(gm_u/netrev_u if netrev_u else 0, 4),
        'ship_u': round(ship_u, 2),
        'mktg_u': round(mktg_u, 2),
        'mktg_t': round(mktg_t, 2),
        'cm1_u': round(cm1_u, 2),
        'cm1_t': round(cm1_t, 2),
        'cm2_u': round(cm2_u, 2),
        'cm2_t': round(cm2_t, 2),
        'cm2pct': round(num(p.get('CM2%', 0))*100, 2),
        'oh_u': round(oh_u, 2),
        'oh_t': round(oh_u*units, 2),
        'ebitda_u': round(ebitda_u, 2),
        'ebitda_t': round(ebitda_t, 2),
        'ebitdapct': round(num(p.get('EBITA%', 0))*100, 2),
        'quad': quad, 'act': act,
        'rev_gst': round(rev_gst, 2),
        'rev_no_gst': round(rev_no_gst, 2)
    })

# Totals
T_units = sum(p['units'] for p in web_prods)
T_rev = sum(p['rev'] for p in web_prods)
T_netrev = sum(p['netrev_t'] for p in web_prods)
T_gm = sum(p['gm_t'] for p in web_prods)
T_cm1 = sum(p['cm1_t'] for p in web_prods)
T_cm2 = sum(p['cm2_t'] for p in web_prods)
T_mktg = sum(p['mktg_t'] for p in web_prods)
T_oh = sum(p['oh_t'] for p in web_prods)
T_ebitda = sum(p['ebitda_t'] for p in web_prods)
T_rev_gst = sum(p['rev_gst'] for p in web_prods)
T_rev_no_gst = sum(p['rev_no_gst'] for p in web_prods)

stars = [p for p in web_prods if p['quad']=='star']
overhead = [p for p in web_prods if p['quad']=='overhead']
loss = [p for p in web_prods if p['quad']=='loss']

print(f"\nWebsite Apr-Jun: {len(web_prods)} SKUs, {int(T_units)} units")
print(f"  Net Rev={T_netrev:,.0f}, CM2={T_cm2:,.0f} ({T_cm2/T_netrev*100:.1f}%), EBITDA={T_ebitda:,.0f}")
print(f"  Stars={len(stars)}, Overhead={len(overhead)}, Loss={len(loss)}")

# Category rollup
cats = {}
for p in web_prods:
    cat = p['cat'] or 'Other'
    if cat not in cats:
        cats[cat] = {'cat': cat, 'rev': 0, 'units': 0, 'cm2': 0, 'ebitda': 0, 'n': 0, 'netrev': 0, 'mktg': 0, 'cm2pct': 0, 'ebitdapct': 0}
    cats[cat]['rev'] += p['rev']; cats[cat]['units'] += p['units']
    cats[cat]['netrev'] += p['netrev_t']; cats[cat]['cm2'] += p['cm2_t']
    cats[cat]['ebitda'] += p['ebitda_t']; cats[cat]['n'] += 1
    cats[cat]['mktg'] += p['mktg_t']
cats_list = sorted(cats.values(), key=lambda x: -x['rev'])
for c in cats_list:
    c['cm2pct'] = round(c['cm2']/c['netrev']*100, 4) if c['netrev'] else 0
    c['ebitdapct'] = round(c['ebitda']/c['netrev']*100, 4) if c['netrev'] else 0

# Pareto
pareto = sorted(web_prods, key=lambda p: -p['rev_gst'])
cum = 0
pareto_list = []
for p in pareto:
    cum += p['rev_gst']
    pareto_list.append({'key': p['key'], 'disp': p['disp'][:80], 'rev': round(p['rev_gst'], 2), 'cum_share': round(cum/T_rev_gst*100, 4)})

# Build new D.periods.apr object (website Apr-Jun period data)
apr_period = {
    'label': 'Apr\u2013Jun 2026',
    'tot': {
        'rev': round(T_rev_gst, 2),
        'netrev': round(T_netrev, 2),
        'units': int(T_units),
        'cm2': round(T_cm2, 2),
        'ebitda': round(T_ebitda, 2),
        'cm2pct': round(T_cm2/T_netrev*100, 5) if T_netrev else 0,
        'ebitdapct': round(T_ebitda/T_netrev*100, 5) if T_netrev else 0,
        'cm2pct_rev': round(T_cm2/T_rev_gst*100, 5) if T_rev_gst else 0,
        'ebitdapct_rev': round(T_ebitda/T_rev_gst*100, 5) if T_rev_gst else 0,
        'mktg': round(T_mktg, 2),
        'gm': round(T_gm, 2),
        'oh': round(T_oh, 2),
        'n': len(web_prods),
        'gmpct': round(T_gm/T_netrev if T_netrev else 0, 4),
        'mktgpct': round(T_mktg/T_netrev if T_netrev else 0, 4)
    },
    'quad': {
        'star': {'n': len(stars), 'rev': round(sum(p['rev_gst'] for p in stars), 2), 'ebitda': round(sum(p['ebitda_t'] for p in stars), 2), 'cm2': round(sum(p['cm2_t'] for p in stars), 2), 'units': int(sum(p['units'] for p in stars))},
        'overhead': {'n': len(overhead), 'rev': round(sum(p['rev_gst'] for p in overhead), 2), 'ebitda': round(sum(p['ebitda_t'] for p in overhead), 2), 'cm2': round(sum(p['cm2_t'] for p in overhead), 2), 'units': int(sum(p['units'] for p in overhead))},
        'loss': {'n': len(loss), 'rev': round(sum(p['rev_gst'] for p in loss), 2), 'ebitda': round(sum(p['ebitda_t'] for p in loss), 2), 'cm2': round(sum(p['cm2_t'] for p in loss), 2), 'units': int(sum(p['units'] for p in loss))}
    },
    'cats': cats_list,
    'products': web_prods,
    'pareto': pareto_list
}

# ============================================================
# INJECT INTO HTML
# ============================================================
with open('index.html', encoding='utf-8') as f:
    html = f.read()

# Find D.periods.apr in the big D object
# The D object is at position 89445 (approx)
# We need to find "apr": inside D.periods
# Strategy: find D = {... "periods": { "jan": {...}, "apr": { ... } ... }}
# Replace just the "apr" value

# The D object starts at const D = {
d_start = html.find('const D = {')
d_end_search = d_start + len('const D = ')

# Count braces to find end of D
i = d_end_search
depth = 0; in_str = False; esc = False
while i < len(html):
    c = html[i]
    if esc: esc = False
    elif c == '\\' and in_str: esc = True
    elif c == '"' and not esc: in_str = not in_str
    elif not in_str:
        if c == '{': depth += 1
        elif c == '}':
            depth -= 1
            if depth == 0: d_end = i + 1; break
    i += 1

d_obj_str = html[d_end_search:d_end]
print(f"D object found, length={len(d_obj_str):,} chars")

# Parse the D object JSON
try:
    D = json.loads(d_obj_str)
    print("D object parsed successfully")
    print("D.periods keys:", list(D.get('periods', {}).keys()))
except json.JSONDecodeError as e:
    print(f"JSON parse error: {e}")
    # Need another approach - find apr by string position
    D = None

if D is not None:
    # Update apr period
    if 'periods' in D:
        D['periods']['apr'] = apr_period
        print(f"Updated D.periods.apr: {len(D['periods']['apr']['products'])} products")
    else:
        D['periods'] = {'apr': apr_period}
        print("Created D.periods.apr")
    
    # Serialize and inject
    new_d_str = json.dumps(D, ensure_ascii=False, default=str)
    html = html[:d_end_search] + new_d_str + html[d_end:]
    
    with open('index.html', 'w', encoding='utf-8') as f:
        f.write(html)
    print("index.html saved with updated website D.periods.apr")
else:
    print("WARNING: Could not update D.periods.apr - D object parse failed")
    # Try string-based replacement for the apr section
    apr_marker = '"apr":'
    apr_pos = d_obj_str.find(apr_marker)
    if apr_pos != -1:
        print(f"Found 'apr' at pos {apr_pos} in D object")
        # Find where apr value ends (by brace counting)
        val_start = apr_pos + len(apr_marker)
        # Skip whitespace
        while val_start < len(d_obj_str) and d_obj_str[val_start] in ' \r\n\t': val_start += 1
        i2 = val_start; depth2 = 0; in_str2 = False; esc2 = False
        while i2 < len(d_obj_str):
            c2 = d_obj_str[i2]
            if esc2: esc2 = False
            elif c2 == '\\' and in_str2: esc2 = True
            elif c2 == '"' and not esc2: in_str2 = not in_str2
            elif not in_str2:
                if c2 == '{': depth2 += 1
                elif c2 == '}':
                    depth2 -= 1
                    if depth2 == 0: val_end = i2 + 1; break
            i2 += 1
        
        # Replace apr value in d_obj_str
        new_apr_json = json.dumps(apr_period, ensure_ascii=False, default=str)
        new_d_obj = d_obj_str[:apr_pos + len(apr_marker)] + new_apr_json + d_obj_str[val_end:]
        html = html[:d_end_search] + new_d_obj + html[d_end:]
        
        with open('index.html', 'w', encoding='utf-8') as f:
            f.write(html)
        print("index.html saved with string-based apr replacement")
    else:
        print("ERROR: Could not find 'apr' in D object")
